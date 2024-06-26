from bdb import effective
import io
import os
import re
import xlrd
import shutil
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from pandas import Timestamp as ts
from datetime import datetime as dt
from zipfile import BadZipFile, ZipFile

from configuration_options import ConfigurationOptions
from ra_logging import TextLogger,EmailLogger,AttachmentLogger
from data_extraction import open_workbook,get_data_range

# 2021-11-04
# California Public Utilities Commission
# Robert Hansen, PE

class Organizer:
    '''
    a class to open excel files within a specified directory and copy files
    recognized as Resource Adequacy Monthly Filings to a separate directory with
    a new name according to the filing contents.
    '''
    def __init__(self,configuration_path:Path,filing_month:ts=None):
        '''
        initializes an object of class Organizer.

        parameters:
            configuration_path - path object pointing to a yaml file containing
                configuration options
            filing_month - an optional filing month timestamp to overwrite the
                date in the configuration options yaml file
        '''
        self.config = ConfigurationOptions(configuration_path,filing_month=filing_month)
        self.logger = TextLogger(
            self.config.get_option('cli_logging_criticalities'),
            self.config.get_option('file_logging_criticalities'),
            self.config.paths.get_path('log')
        )
        self.email_logger = EmailLogger(log_path=self.config.paths.get_path('email_log'))
        self.attachment_logger = AttachmentLogger(log_path=self.config.paths.get_path('attachment_log'))

    def validate_attachment(self,attachment_id:str):
        '''
        reviews an attachment linked by the input path object and categorizes
        it based on its relevance to the resource adequacy filing program.

        parameters:
            attachment_id - a string representing a single email attachment
                downloaded from kiteworks
        '''
        def set_attachment_value(column,value):
            attachment_selection = (self.attachment_logger.data.loc[:,'attachment_id']==attachment_id)
            self.attachment_logger.data.loc[attachment_selection,column] = value
        attachment = self.attachment_logger.data.loc[(self.attachment_logger.data.loc[:,'attachment_id']==attachment_id),:].iloc[0]
        download_path = Path(attachment.loc['download_path'])
        emails = self.email_logger.data
        email_information = emails.loc[(emails.loc[:,'email_id']==attachment.loc['email_id']),:].iloc[0]
        self.logger.log('Validating Attachment: {} ({})'.format(attachment_id,download_path),'INFORMATION')
        if download_path.is_file():
            if download_path.suffix in ('.xlsx','.xlsm'):
                with open(download_path,'rb') as f:
                    in_mem_file = io.BytesIO(f.read())
                    try:
                        sheetnames = {
                            'monthly_filing' : [
                                'Certification',
                                'LSE Allocations',
                                'I_Phys_Res_Import_RA_Res',
                                'III_Demand_Response',
                            ],
                            'cam_rmr_update' : ['Jun to Dec CAM Update','Diablo Canyon Credits'],
                            'incremental_local' : ['IncrementalLocal'],
                            'year_ahead' : [
                                'loadforecastinputdata',
                                'DRforAllocation',
                                'Flexrequirements',
                                'Flex RMR',
                                'Local RA-CAM-{}'.format(pd.to_datetime(self.config.filing_month).year),
                            ],
                            'month_ahead' : ['Monthly Tracking'],
                            'cam_rmr' : [
                                'CAMRMR',
                                'monthlytracking',
                            ],
                            'supply_plan': ['Export'],
                            'nqc_list' : r'(\d{4}).*NQC List',
                        }
                        wb = load_workbook(in_mem_file,data_only=True,read_only=True)
                        # monthly filing:
                        if all([sheetname in wb.sheetnames for sheetname in sheetnames['monthly_filing']]):
                            set_attachment_value('ra_category','ra_monthly_filing')
                            sheet = wb['Certification']
                            if self.config.organizations.lookup_id(sheet['B5'].value):
                                submittal_information = {
                                    'date' : sheet['B3'].value,
                                    'organization_full' : sheet['B5'].value,
                                    'organization_id' : self.config.organizations.lookup_id(sheet['B5'].value),
                                    'organization_representative' : {
                                        'name' : sheet['B21'].value,
                                        'title' : sheet['B23'].value,
                                        'email' : sheet['B22'].value,
                                        'sign_date' : sheet['B24'].value,
                                    },
                                    'organization_contact' : {
                                        'name' : sheet['B28'].value,
                                        'title' : sheet['B29'].value,
                                        'address' : '{}\n{}\n{}, {} {}'.format(
                                            sheet['B30'].value,
                                            sheet['B31'].value,
                                            sheet['B32'].value,
                                            sheet['B33'].value,
                                            sheet['B34'].value
                                        ),
                                        'phone' : sheet['B35'].value,
                                        'email' : sheet['B36'].value,
                                    },
                                    'organization_backup_contact' : {
                                        'name' : sheet['B40'].value,
                                        'title' : sheet['B41'].value,
                                        'phone' : sheet['B42'].value,
                                        'email' : sheet['B43'].value,
                                    },
                                    'compliance_period' : sheet['B3'].value,
                                    'submittal_date' : sheet['B7'].value,
                                }
                                effective_date = submittal_information['date']
                                if not isinstance(effective_date,dt):
                                    effective_date = self.config.filing_month
                                else:
                                    pass
                                set_attachment_value('effective_date',effective_date)
                                set_attachment_value('organization_id',submittal_information['organization_id'])
                            else:
                                self.logger.log('Load Serving Entity Alias Not Recognized: \'{}\'\tAdd ID and Aliases to {}'.format(sheet['B5'].value,self.config.paths.get_path('organizations')),'WARNING')
                                effective_date = self.config.filing_month
                                set_attachment_value('effective_date',effective_date)
                                set_attachment_value('organization_id','[LSE Alias Not Recognized]')
                        # CAM, RMR, and Diablo Canyon credit true-up:
                        elif all([sheetname in wb.sheetnames for sheetname in sheetnames['cam_rmr_update']]) and email_information.loc['group']=='internal':
                            set_attachment_value('ra_category','cam_rmr_update')
                            set_attachment_value('organization_id','CEC')
                            if re.match(r'.*(\d{4}).*',download_path.name):
                                effective_date = pd.to_datetime(re.match(r'.*(\d{4}).*',download_path.name).groups()[0]).replace(month=6)
                                set_attachment_value('effective_date',effective_date)
                            else:
                                effective_date = self.config.filing_month.replace(month=6)
                                set_attachment_value('effective_date',effective_date)
                        # incremental local:
                        elif all([sheetname in wb.sheetnames for sheetname in sheetnames['incremental_local']]) and email_information.loc['group']=='internal':
                            set_attachment_value('ra_category','incremental_local')
                            set_attachment_value('organization_id','CEC')
                            if re.match(r'.*(\d{4}).*',download_path.name):
                                effective_date = pd.to_datetime(re.match(r'.*(\d{4}).*',download_path.name).groups()[0]).replace(month=7)
                                set_attachment_value('effective_date',effective_date)
                            else:
                                effective_date = self.config.filing_month.replace(month=7)
                                set_attachment_value('effective_date',effective_date)
                        # year ahead:
                        elif all([sheetname in wb.sheetnames for sheetname in sheetnames['year_ahead']]) and email_information.loc['group']=='internal':
                            set_attachment_value('ra_category','year_ahead')
                            set_attachment_value('organization_id','CEC')
                            if re.match(r'.*(\d{4}).*',download_path.name):
                                effective_date = pd.to_datetime(re.match(r'.*(\d{4}).*',download_path.name).groups()[0])
                                set_attachment_value('effective_date',effective_date)
                            else:
                                effective_date = self.config.filing_month
                                set_attachment_value('effective_date',effective_date)
                        # month ahead:
                        elif all([sheetname in wb.sheetnames for sheetname in sheetnames['month_ahead']]) and email_information.loc['group']=='internal':
                            set_attachment_value('ra_category','month_ahead')
                            set_attachment_value('organization_id','CPUC')
                            if re.match(r'.*(\d{4}).*',download_path.name):
                                effective_date = pd.to_datetime(re.match(r'.*(\d{4}).*',download_path.name).groups()[0])
                                set_attachment_value('effective_date',effective_date)
                            else:
                                effective_date = self.config.filing_month
                                set_attachment_value('effective_date',effective_date)
                        # cam-rmr:
                        elif all([sheetname in wb.sheetnames for sheetname in sheetnames['cam_rmr']]) and email_information.loc['group']=='internal':
                            effective_date = pd.to_datetime(dt.strptime(' '.join(re.match(r'(\w{3})MA(\d{2})',wb['CAMRMR']['A1'].value).groups()),'%b %y'))
                            set_attachment_value('ra_category','cam_rmr')
                            set_attachment_value('organization_id','CPUC')
                            set_attachment_value('effective_date',effective_date)
                        # nqc list:
                        elif any([re.match(sheetnames['nqc_list'], s) for s in wb.sheetnames]):
                            matching_sheetnames = filter(lambda s: s,[re.match(sheetnames['nqc_list'], s) for s in wb.sheetnames])
                            effective_date = ts(min([int(x.groups()[0]) for x in matching_sheetnames]),1,1)
                            set_attachment_value('ra_category','nqc_list')
                            set_attachment_value('organization_id','CPUC')
                            set_attachment_value('effective_date',effective_date)
                        # system and flexible supply plans:
                        elif all(sheetname in wb.sheetnames for sheetname in sheetnames['supply_plan']):
                            sheet = wb[sheetnames['supply_plan'][0]]
                            columns = [sheet['{}1'.format(get_column_letter(i+1))].value for i in range(sheet.max_column)]
                            system_columns = [
                                r'\s*validation\s*',
                                r'\s*scid\s*',
                                r'.*id\s*',
                                r'\s*local.*',
                                r'\s*system.*',
                                r'.*total.*',
                                r'.*start.*',
                                r'.*end.*',
                                r'.*lse.*',
                                r'\s*errors.*warnings\s*',
                            ]
                            local_columns = [
                                r'\s*validation\s*',
                                r'\s*supplier\s*',
                                r'.*id\s*',
                                r'\s*category\s*',
                                r'\s*flex.*',
                                r'.*start.*',
                                r'.*end.*',
                                r'\s*lse\s*',
                                r'\s*errors.*warnings\s*',
                            ]
                            effective_date_elements = re.match(r'.*\s(\d{1,2})[_\W](\d{1,2})[_\W](\d{2,4}).*',download_path.name).groups()
                            if len(effective_date_elements[2])==2:
                                effective_date = pd.to_datetime(dt.strptime('/'.join(effective_date_elements),'%m/%d/%y'))
                            else:
                                effective_date = pd.to_datetime(dt.strptime('/'.join(effective_date_elements),'%m/%d/%Y'))
                            effective_date = effective_date.replace(effective_date.year+int((effective_date.month+1)/12),(effective_date.month+1)%12+1,1)
                            if len(columns)>=len(system_columns) and all([re.match(s,columns[i].lower()) for i,s in enumerate(system_columns)]):
                                set_attachment_value('ra_category','supply_plan_system')
                                set_attachment_value('effective_date',effective_date)
                                set_attachment_value('organization_id','CAISO')
                            elif len(columns)>=len(local_columns) and all([re.match(s,columns[i].lower()) for i,s in enumerate(local_columns)]):
                                set_attachment_value('ra_category','supply_plan_flexible')
                                set_attachment_value('effective_date',effective_date)
                                set_attachment_value('organization_id','CAISO')
                            else:
                                self.logger.log('Input Excel File Does Not Match Templates: '.format(download_path),'INFORMATION')
                                set_attachment_value('ra_category','none')
                                set_attachment_value('effective_date','NaT')
                                set_attachment_value('organization_id','n/a')
                            # system supply plan:
                            # flexible supply plan:
                        else:
                            set_attachment_value('ra_category','none')
                            set_attachment_value('organization_id','n/a')
                            set_attachment_value('effective_date','NAT')
                        wb.close()
                    except BadZipFile:
                        self.logger.log('Unable to open Excel file: {}'.format(download_path),'WARNING')
                        set_attachment_value('ra_category','none')
                        set_attachment_value('organization_id','n/a')
                        set_attachment_value('effective_date','NAT')
            # system and flexible supply plans (old format):
            elif download_path.suffix=='.xls' and email_information.loc['group']=='internal':
                wb = xlrd.open_workbook(download_path)
                sheet = wb.sheet_by_index(0)
                columns = [sheet.cell_value(rowx=0,colx=column_number) for column_number in range(sheet.ncols)]
                system_columns = [
                    r'\s*validation\s*',
                    r'\s*scid\s*',
                    r'.*id\s*',
                    r'\s*local.*',
                    r'\s*system.*',
                    r'.*total.*',
                    r'.*start.*',
                    r'.*end.*',
                    r'.*lse.*',
                    r'\s*errors.*warnings\s*',
                ]
                local_columns = [
                    r'\s*validation\s*',
                    r'\s*supplier\s*',
                    r'.*id\s*',
                    r'\s*category\s*',
                    r'\s*flex.*',
                    r'.*start.*',
                    r'.*end.*',
                    r'\s*lse\s*',
                    r'\s*errors.*warnings\s*',
                ]
                effective_date_elements = re.match(r'.*\s(\d{1,2})[_\W](\d{1,2})[_\W](\d{2,4}).*',download_path.name).groups()
                if len(effective_date_elements[2])==2:
                    effective_date = pd.to_datetime(dt.strptime('/'.join(effective_date_elements),'%m/%d/%y'))
                else:
                    effective_date = pd.to_datetime(dt.strptime('/'.join(effective_date_elements),'%m/%d/%Y'))
                effective_date = effective_date.replace(effective_date.year+int((effective_date.month+1)/12),(effective_date.month+1)%12+1,1)
                if len(columns)>=len(system_columns) and all([re.match(s,columns[i].lower()) for i,s in enumerate(system_columns)]):
                    set_attachment_value('ra_category','supply_plan_system')
                    set_attachment_value('effective_date',effective_date)
                    set_attachment_value('organization_id','CAISO')
                elif len(columns)>=len(local_columns) and all([re.match(s,columns[i].lower()) for i,s in enumerate(local_columns)]):
                    set_attachment_value('ra_category','supply_plan_flexible')
                    set_attachment_value('effective_date',effective_date)
                    set_attachment_value('organization_id','CAISO')
                else:
                    self.logger.log('Input Excel File Does Not Match Templates: '.format(download_path),'INFORMATION')
                    set_attachment_value('ra_category','none')
                    set_attachment_value('effective_date','NaT')
                    set_attachment_value('organization_id','n/a')
            else:
                self.logger.log('Skipping {} File: {}'.format(download_path.suffix,download_path),'INFORMATION')
                set_attachment_value('ra_category','none')
                set_attachment_value('effective_date','NaT')
                set_attachment_value('organization_id','n/a')
        else:
            self.logger.log('Input File Not Found: {}'.format(download_path),'WARNING')
            set_attachment_value('ra_category','none')
            set_attachment_value('effective_date','NaT')
            set_attachment_value('organization_id','n/a')
        self.attachment_logger.commit()

    def ingest_manual_downloads(self):
        '''
        checks for files in the download directories that do not appear in the
        attachment log and adds them as manual overrides.
        '''
        ingest_timestamp = ts.now()
        downloads_internal = self.config.paths.get_path('downloads_internal').iterdir()
        downloads_external = self.config.paths.get_path('downloads_external').iterdir()
        for path,sender_group in [(path,'internal') for path in downloads_internal] + [(path,'external') for path in downloads_external]:
            if path.is_file() and str(path) not in self.attachment_logger.data.loc[:,'download_path'].values:
                internal_bit = int(sender_group=='internal')
                email_id = '00000000-0000-0000-0000-{}000{}'.format(ingest_timestamp.strftime('%Y%m%d'),internal_bit)
                if email_id not in self.email_logger.data.loc[:,'email_id'].values:
                    email_information = pd.Series({
                        'email_id' : email_id,
                        'sender' : 'manual_download',
                        'subject' : 'n/a',
                        'receipt_date' : ingest_timestamp,
                        'included' : True,
                        'group' : sender_group,
                    })
                    self.email_logger.log(email_information)
                attachment_index = len(self.attachment_logger.data.loc[(self.attachment_logger.data.loc[:,'email_id']==email_id),'attachment_id'])
                attachment_id = '{}000{}{:020.0f}'.format(ingest_timestamp.strftime('%Y%m%d'),internal_bit,attachment_index)
                attachment_information = pd.Series({
                    'email_id' : email_id,
                    'attachment_id' : attachment_id,
                    'download_path' : str(path),
                    'ra_category' : 'not_validated',
                    'organization_id' : '',
                    'effective_date' : '',
                    'archive_path' : '',
                })
                self.attachment_logger.log(attachment_information)
        self.attachment_logger.commit()
        self.email_logger.commit()

    def validate_all(self):
        '''
        validates each entry in the attachments_log and fills additional
        information based on attachment contents.
        '''
        for _,attachment in self.attachment_logger.data.loc[(self.attachment_logger.data.loc[:,'ra_category']=='not_validated'),:].iterrows():
            attachment_id = attachment.loc['attachment_id']
            self.validate_attachment(attachment_id)

    def set_versions(self):
        '''
        determines the version numbers for each attachment in the attachment
        log and sets the archive path for all entries in the attachment_log
        which don't yet have one.
        '''
        columns = self.attachment_logger.data.columns
        self.attachment_logger.data = self.attachment_logger.data.merge(self.email_logger.data.loc[:,['email_id','receipt_date']],on='email_id')
        self.attachment_logger.data.loc[:,'version'] = \
            self.attachment_logger.data.sort_values(['ra_category','organization_id','effective_date','receipt_date']) \
            .groupby(['ra_category','organization_id','effective_date']).cumcount()
        def get_archive_path(r):
            ra_category = r.loc['ra_category']
            organization = self.config.organizations.get_organization(r.loc['organization_id'])
            if organization:
                if r.loc['effective_date'] not in ('','n/a'):
                    effective_date = r.loc['effective_date']
                else:
                    effective_date = self.config.filing_month
                version = r.loc['version']
                archive_path = str(self.config.paths.get_path(ra_category,organization=organization,date=effective_date,version=version))
            else:
                archive_path = ''
            return archive_path
        self.attachment_logger.data.loc[:,'archive_path'] = self.attachment_logger.data.apply(get_archive_path,axis='columns')
        self.attachment_logger.data = self.attachment_logger.data.loc[:,columns]
        self.attachment_logger.commit()

    def copy_rename(self,attachment_id:str):
        '''
        copies a single attachment identified by its attachment_id to the archive path if the file doesn't already exist.

        parameters:
            attachment_id - a string representing a single email attachment
                downloaded from kiteworks
        '''
        attachment = self.attachment_logger.data.loc[(self.attachment_logger.data.loc[:,'attachment_id']==attachment_id),:].iloc[0]
        root_directory = Path(self.config.get_option('archive_root_directory'))
        download_path = Path(attachment.loc['download_path'])
        archive_path = Path(attachment.loc['archive_path'])
        if archive_path.name!='':
            if download_path.is_file() and not archive_path.is_file():
                archive_path.parent.mkdir(parents=True,exist_ok=True)
                shutil.copyfile(download_path,root_directory/archive_path)
                self.logger.log('Copying {} to {}'.format(download_path.relative_to(root_directory),archive_path),'INFORMATION')
            elif Path(archive_path).is_file():
                self.logger.log('Skipping File {} -- Already Exists in Archive as {}'.format(download_path.relative_to(root_directory),archive_path),'INFORMATION')
            else:
                self.logger.log('Unable to Copy {} to Archive'.format(download_path.relative_to(root_directory)),'INFORMATION')
        else:
            self.logger.log('Skipping File {} -- Does Not Match Any RA Templates'.format(download_path.relative_to(root_directory)),'INFORMATION')

    def copy_rename_all(self):
        '''
        copies all current attachments to their archive locations
        '''
        self.attachment_logger.load_log()
        filing_month = pd.to_datetime(self.config.filing_month)
        current_attachment_ids = self.attachment_logger.data.loc[
            (
                (
                    (self.attachment_logger.data.loc[:,'ra_category']=='ra_monthly_filing') | \
                    (self.attachment_logger.data.loc[:,'ra_category']=='supply_plan_system') | \
                    (self.attachment_logger.data.loc[:,'ra_category']=='supply_plan_flexible') | \
                    (self.attachment_logger.data.loc[:,'ra_category']=='cam_rmr')
                ) & \
                (self.attachment_logger.data.loc[:,'effective_date']==filing_month)
            ) | \
            (
                (
                    (self.attachment_logger.data.loc[:,'ra_category']=='year_ahead') | \
                    (self.attachment_logger.data.loc[:,'ra_category']=='month_ahead') | \
                    (self.attachment_logger.data.loc[:,'ra_category']=='nqc_list')
                ) & \
                (self.attachment_logger.data.loc[:,'effective_date']==filing_month.replace(month=1))
            ) | \
            (
                (self.attachment_logger.data.loc[:,'ra_category']=='cam_rmr_update') & \
                (self.attachment_logger.data.loc[:,'effective_date']==filing_month.replace(month=6))
            ) | \
            (
                (self.attachment_logger.data.loc[:,'ra_category']=='incremental_local') & \
                (self.attachment_logger.data.loc[:,'effective_date']==filing_month.replace(month=7))
            ),'attachment_id']
        for attachment_id in current_attachment_ids:
            self.copy_rename(attachment_id)

    def unzip(self,path:Path):
        '''
        expands a single zip archive into its parent directory.

        parameters:
            path - the path of a zip file to decompress
        '''
        if path.is_file():
            try:
                with ZipFile(str(path),'r') as z:
                    z.extractall(path.parent)
                self.logger.log('Decompressing {} Archive: {}'.format(path.suffix,path),'INFORMATION')
            except:
                self.logger.log('Unable to Decompress Archive','WARNING')
        else:
            self.logger.log('File Not Found: {}'.format(path),'WARNING')

    # traverse directory tree, opening sub-directories recursively and copying/renaming files:
    def traverse(self,directory_path: Path,download_source: str):
        '''
        loops through contents of a directory, copying files to their archive
        directory based on the download source and recursively opening
        sub-directories.

        parameters:
            directory_path - path object pointing to the directory to traverse
            download_source - string indicating whether to treat directory
                contents as internal or external
        '''
        if directory_path.is_dir():
            for item in directory_path.iterdir():
                if item.is_dir():
                    self.traverse(item,download_source)
                elif item.is_file():
                    self.copy_rename(item,download_source)
                else:
                    self.logger.log('Unknown Item: {}'.format(item),'WARNING')

    def cleanup(self,directory_path: Path):
        '''
        deletes all contents of the specified directory.

        parameters:
            directory_path - path object pointing to the directory to clean up
        '''
        if directory_path.is_dir():
            for item in directory_path.iterdir():
                if item.is_dir():
                    shutil.rmtree(item)
                elif item.is_file():
                    os.remove(item)
        elif directory_path.is_file():
            os.remove(directory_path)
        else:
            pass

    def organize(self):
        '''
        calls each of the main methods of the ra_organizer method in sequence
        to handle downloaded and manually placed input files and prepare for
        consolidation and summarization.
        '''
        # for item in itertools.chain(self.config.paths.get_path('downloads_internal').iterdir(),self.config.paths.get_path('downloads_external').iterdir()):
        #     if item.suffix=='.zip':
        #         self.unzip(item)
        #     else:
        #         pass
        # self.traverse(self.config.paths.get_path('downloads_internal'),'internal')
        # self.traverse(self.config.paths.get_path('downloads_external'),'external')
        # self.cleanup(self.config.paths.get_path('download_directory'))
        self.ingest_manual_downloads()
        self.validate_all()
        self.set_versions()
        self.copy_rename_all()

    def compress_archive(self):
        '''
        creates a zip archive and populates it with compressed copies of the
        files specified in the Organizer instance's Paths object.
        '''
        # get list of current lses from summary template file:
        path = self.config.paths.get_path('ra_summary_template')
        ra_summary = open_workbook(path,data_only=True,read_only=True)
        data_range = get_data_range(ra_summary['Summary'],'A','',self.config)
        active_lses = [row[0].value for row in data_range]
        path_ids = [(path_id,None) for path_id in filter(lambda s: s!='ra_monthly_filing',self.config.paths.files_for_archive)] + \
            list(zip(['ra_monthly_filing'] * len(active_lses),active_lses))
        paths = [path_id for path_ids in [self.config.paths.get_all_versions(path_id[0],organization=self.config.organizations.get_organization(path_id[1])) for path_id in path_ids] for path_id in path_ids]
        if self.config.paths.get_path('results_archive').exists():
            self.config.paths.get_path('results_archive').unlink()
        else:
            pass
        with ZipFile(self.config.paths.get_path('results_archive'),'x') as archive:
            for path in paths:
                if path is not None and path.is_file():
                    archive.write(path,arcname=str(path))
                else:
                    pass