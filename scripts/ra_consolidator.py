from os import rename
import xlrd
import numpy as np
import pandas as pd
from pathlib import Path
from pandas import Timestamp as ts
from openpyxl.styles import PatternFill,Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.cell import get_column_letter

from ra_logging import TextLogger,EmailLogger,AttachmentLogger,ConsolidationLogger
from configuration_options import ConfigurationOptions
from data_extraction import open_workbook,get_data_range,data_range_to_dataframe,read_ra_monthly_filing,get_year_ahead_tables,get_month_ahead_tables,get_cam_rmr_tables,get_incremental_local_tables,get_nqc_list,read_supply_plan,rename_locality

class WorkbookConsolidator:
    '''
    this class contains methods to collect data from allocations, filings, and
    supply plans into monthly summary and cross-check files.
    '''
    def __init__(self,configuration_path:Path,filing_month:ts=None):
        '''
        initializes an instance of the WorkbookConsolidator class

        parameters:
            configuration_path - path object pointing to a yaml file containing
                configuration options for the WorkbookConsolidator object
            filing_month - an optional filing month timestamp to overwrite the
                date in the configuration options yaml file
        '''
        self.config = ConfigurationOptions(configuration_path,filing_month=filing_month)
        self.logger = TextLogger(
            self.config.get_option('cli_logging_criticalities'),
            self.config.get_option('file_logging_criticalities'),
            self.config.paths.get_path('log')
        )
        self.email_logger = EmailLogger(self.config.paths.get_path('email_log'))
        self.attachment_logger = AttachmentLogger(self.config.paths.get_path('attachment_log'))
        self.consolidation_logger = ConsolidationLogger(self.config.paths.get_path('consolidation_log'))

    def initialize_ra_summary(self):
        '''
        creates or overwrites an ra summary workbook for the filing month
        defined in the WorkbookConsolidator instance's ConfigurationOptions
        object at the location defined in its Paths object.
        '''
        self.logger.log('Creating New Monthly RA Summary File from Template: {}'.format(self.config.paths.get_path('ra_summary_template').name),'INFORMATION')

        # open ra_summary_starter file into memory:
        try:
            path = self.config.paths.get_path('ra_summary_template')
            ra_summary = open_workbook(path,data_only=False,read_only=False,in_mem=True)
            ra_summary.save(str(self.config.paths.get_path('ra_summary')))
            ra_summary.close()
            self.logger.log('Created New Monthly RA Summary File: {}'.format(self.config.paths.get_path('ra_summary').name),'INFORMATION')
        except:
            self.logger.log('Unable to Create New Monthly RA Summary File','ERROR')

    def initialize_caiso_cross_check(self):
        '''
        creates or overwrites a caiso cross-check file for the filing month
        defined in the WorkbookConsolidator instance's ConfigurationOptions
        object at the location defined in its Paths object.
        '''
        self.logger.log('Creating New CAISO Supply Plan Cross-Check File from Template: {}'.format(self.config.paths.get_path('caiso_cross_check_template').name),'INFORMATION')
        try:
            path = self.config.paths.get_path('caiso_cross_check_template')
            caiso_cross_check = open_workbook(path,data_only=False,read_only=False,in_mem=True)
            caiso_cross_check.save(str(self.config.paths.get_path('caiso_cross_check')))
            caiso_cross_check.close()
            self.logger.log('Created New CAISO Supply Plan Cross-Check File: {}'.format(self.config.paths.get_path('caiso_cross_check').name),'INFORMATION')
        except:
            self.logger.log('Unable to Create New CAISO Supply Plan Cross-Check File','ERROR')

    def consolidate_allocations(self):
        '''
        loads data from various regulatory workbook and populates the summary
        and caiso cross-check workbooks with data and calculation results.
        '''
        self.logger.log('Consolidating Allocation Data','INFORMATION')

        # start timer:
        init_time = ts.now()

        filing_month = self.config.filing_month

        # get source data from year ahead file:
        path = self.config.paths.get_path('year_ahead')
        year_ahead = open_workbook(path)
        year_ahead_tables = get_year_ahead_tables(year_ahead,self.config)
        load_forecast_input_data = year_ahead_tables[0]
        demand_response_allocation = year_ahead_tables[1]
        flexibility_requirements = year_ahead_tables[3]
        flexibility_rmr = year_ahead_tables[4]
        flexibility_cme = year_ahead_tables[5]
        local_rar = year_ahead_tables[6]
        cam_system = year_ahead_tables[8]
        irp_system = year_ahead_tables[9]
        year_ahead.close()

        # get source data from month ahead file:
        path = self.config.paths.get_path('month_ahead')
        month_ahead = open_workbook(path,in_mem=False)
        (month_ahead_forecasts,monthly_tracking) = get_month_ahead_tables(month_ahead,self.config)
        month_ahead.close()

        # get source data from cam-rmr file:
        if filing_month.year < 2024:
            path = self.config.paths.get_path('cam_rmr')
            cam_rmr = open_workbook(path,in_mem=False)
            (cam_rmr_monthly_tracking,total_cam_rmr) = get_cam_rmr_tables(cam_rmr)
            cam_rmr.close()
        else:
            (cam_rmr_monthly_tracking,total_cam_rmr) = (None,None)

        # get source data from incremental local workbook:
        if filing_month.month >= 7:
            path = self.config.paths.get_path('incremental_local')
            incremental_local = open_workbook(path)
            (incremental_flex,incremental_local_load,local_rar_trueup) = get_incremental_local_tables(incremental_local,self.config)
            incremental_local.close()
        else:
            (incremental_flex,incremental_local_load,local_rar_trueup) = (None,None,None)

        # open summary file and initialize summary table:
        path = self.config.paths.get_path('ra_summary')
        ra_summary = open_workbook(path,data_only=False,read_only=False)
        columns = [
            'organization_id',
        ]
        data_range = get_data_range(ra_summary['Summary'],'A','',self.config)
        summary = data_range_to_dataframe(columns,data_range)

        # open caiso supply plan cross-check file:
        path = self.config.paths.get_path('caiso_cross_check')
        caiso_cross_check = open_workbook(path,data_only=False,read_only=False)

        # create function to use in dataframe.apply():
        def calculate_summary(row):
            organization_id = row.loc['organization_id']
            if organization_id in ('PGE','SCE','SDGE'):
                inverse_organization_selection = list(dict.fromkeys(filter(lambda idx: idx!=organization_id,monthly_tracking.index.get_level_values(0))).keys())
            else:
                inverse_organization_selection = []
            all_lses = list(map(lambda d: d['id'],self.config.organizations.list_load_serving_entities()))
            # NP26 summary sheet:
            if organization_id=='PGE' and filing_month.year<2024:
                cam_load_share_pge = -monthly_tracking.loc[(inverse_organization_selection,filing_month),'pge_revised_nonjurisdictional_load_share'].sum()
            elif filing_month.year<2024:
                cam_load_share_pge = monthly_tracking.loc[(organization_id,filing_month),'pge_revised_nonjurisdictional_load_share']
            else:
                cam_load_share_pge = 0
            if filing_month.year<2024:
                np26_cam = total_cam_rmr.loc['np26_cam'] * cam_load_share_pge
                np26_rmr = total_cam_rmr.loc['np26_rmr'] * monthly_tracking.loc[(organization_id,filing_month),'pge_revised_nonjurisdictional_load_share'] + total_cam_rmr.loc['system_rmr'] * monthly_tracking.loc[(organization_id,filing_month),'total_revised_jurisdictional_load_share']
            else:
                np26_cam = cam_rmr.loc[(organization_id,'north','cam'),filing_month]
                np26_rmr = cam_rmr.loc[(organization_id,'north','rmr'),filing_month]
            np26_cpe_system_cam = np.round(
                cam_system.loc[(cam_system.loc[:,'path_26_region']=='north'),filing_month.to_numpy().astype('datetime64[M]')].sum() * \
                load_forecast_input_data.loc[filter(lambda i: i[0]=='PGE' and i[1]==organization_id and i[2]==filing_month,load_forecast_input_data.index),'final_coincident_peak_forecast'].sum() / \
                load_forecast_input_data.loc[filter(lambda i: i[0]=='PGE' and i[1] in all_lses and i[2]==filing_month,load_forecast_input_data.index),'final_coincident_peak_forecast'].sum(),
                2
            )
            np26_irp_system_cam = irp_system.loc[(irp_system.loc[:,'path_26_region']=='north')&(irp_system.loc[:,'lse']==organization_id),filing_month]
            np26_ra_obligation = np.round(
                (
                    # PGEload:
                    (1 + self.config.get_option('planning_reserve_margin')) * month_ahead_forecasts.loc[(organization_id,filing_month),'pge_revised_monthly_forecast']
                    # NP26CAM:
                    -np26_cam
                    # NP26RMR:
                    -np26_rmr
                    # CPE System CAM:
                    -np26_cpe_system_cam
                    # IRP System CAM:
                    -np26_irp_system_cam
                ),
                0
            )
            sn_path26_allocation = 0
            # SP26 summary sheet:
            if organization_id=='SCE' and filing_month.year<2024:
                cam_load_share_sce = -cam_rmr_monthly_tracking.loc[(inverse_organization_selection,filing_month),'sce_revised_nonjurisdictional_load_share'].sum()
                cam_load_share_sdge = cam_rmr_monthly_tracking.loc[(organization_id,filing_month),'sdge_revised_nonjurisdictional_load_share']
            elif organization_id=='SDGE' and filing_month.year<2024:
                cam_load_share_sce = cam_rmr_monthly_tracking.loc[(organization_id,filing_month),'sce_revised_nonjurisdictional_load_share']
                cam_load_share_sdge = -cam_rmr_monthly_tracking.loc[(inverse_organization_selection,filing_month),'sdge_revised_nonjurisdictional_load_share'].sum()
            elif filing_month<2024:
                cam_load_share_sce = cam_rmr_monthly_tracking.loc[(organization_id,filing_month),'sce_revised_nonjurisdictional_load_share']
                cam_load_share_sdge = cam_rmr_monthly_tracking.loc[(organization_id,filing_month),'sdge_revised_nonjurisdictional_load_share']
            else:
                cam_load_share_sce = 0
                cam_load_share_sdge = 0
            if filing_month.year<2024:
                sp26_cam = total_cam_rmr.loc['sce_cam'] * cam_load_share_sce + total_cam_rmr.loc['sdge_cam'] * cam_load_share_sdge
                sp26_rmr = total_cam_rmr.loc['sp26_rmr'] * cam_rmr_monthly_tracking.loc[(organization_id,filing_month),'sce_revised_nonjurisdictional_load_share']
                sce_lcr = total_cam_rmr.loc['sce_preferred_lcr_credit'] * cam_rmr_monthly_tracking.loc[(organization_id,filing_month),'sce_revised_jurisdictional_load_share']
            else:
                sp26_cam = cam_rmr.loc[(cam_rmr.loc[:,'lse']==organization_id)&(cam_rmr.loc[:,'path_26_region']=='south')&(cam_rmr.loc[:,'type']=='cam'),filing_month]
                sp26_rmr = cam_rmr.loc[(cam_rmr.loc[:,'lse']==organization_id)&(cam_rmr.loc[:,'path_26_region']=='south')&(cam_rmr.loc[:,'type']=='rmr'),filing_month]
                sce_lcr = 0
            sp26_cpe_system_cam = np.round(
                cam_system.loc[(cam_system.loc[:,'path_26_region']=='south'),filing_month.to_numpy().astype('datetime64[M]')].sum() * \
                load_forecast_input_data.loc[filter(lambda i:i[0]=='SCE' and i[1]==organization_id and i[2]==filing_month,load_forecast_input_data.index),'final_coincident_peak_forecast'].sum() / \
                load_forecast_input_data.loc[filter(lambda i:i[0]=='SCE' and i[1] in all_lses and i[2]==filing_month,load_forecast_input_data.index),'final_coincident_peak_forecast'].sum(),
                2
            )
            sp26_irp_system_cam = irp_system.loc[(irp_system.loc[:,'path_26_region']=='south')&(irp_system.loc[:,'lse']==organization_id),filing_month]
            sp26_ra_obligation = np.round(
                (
                    (1 + self.config.get_option('planning_reserve_margin')) * (
                        # SCEload:
                        month_ahead_forecasts.loc[(organization_id,filing_month),'sce_revised_monthly_forecast']
                        # SDGEload:
                        +month_ahead_forecasts.loc[(organization_id,filing_month),'sdge_revised_monthly_forecast']
                    )
                    # SP26CAM:
                    -sp26_cam
                    # SP26RMR:
                    -sp26_rmr
                    # SCELCR:
                    -sce_lcr
                    # CPE System CAM:
                    -sp26_cpe_system_cam
                    # IRP System CAM:
                    -sp26_irp_system_cam
                ) ,
                0
            )
            ns_path26_allocation = 0
            if filing_month.year < 2024:
                system_rmr_credit = total_cam_rmr.loc['system_rmr'] * cam_rmr_monthly_tracking.loc[(organization_id,filing_month),'total_revised_jurisdictional_load_share']
            else:
                system_rmr_credit = cam_rmr.loc[(organization_id,'system','rmr'),filing_month]
            def incremental_flex_by_category(category:int):
                if incremental_flex is not None:
                    flex = incremental_flex.loc[(organization_id,category),'flexibility_requirement']
                else:
                    flex = 0
                return flex
            def incremental_load_by_area(area:str):
                if incremental_local_load is not None:
                    incremental_load = incremental_local_load.loc[(organization_id,area),'incremental_load']
                else:
                    incremental_load = 0
                return incremental_load
            def august_demand_response(iou_territory:str,location:str,allocation_type:str):
                month = ts(filing_month.year,8,1)
                if (iou_territory,organization_id,month) in load_forecast_input_data.index and (location,allocation_type,month) in demand_response_allocation.index:
                    august_forecast_lse = load_forecast_input_data.loc[(iou_territory,organization_id,month),'final_coincident_peak_forecast']
                    organization_id_indices = list(dict.fromkeys(load_forecast_input_data.loc[(iou_territory),:].index.get_level_values(0)))
                    august_forecast_local = load_forecast_input_data.loc[(iou_territory,organization_id_indices,month),'final_coincident_peak_forecast'].sum()
                    august_demand_response_allocation = demand_response_allocation.loc[(location,allocation_type,month),'allocation'].sum()
                    if august_forecast_local>0:
                        demand_response = np.round(august_forecast_lse / august_forecast_local * august_demand_response_allocation,decimals=2)
                    else:
                        demand_response = 0
                else:
                    demand_response = 0
                return demand_response

            return pd.Series({
                'organization_id' : organization_id,
                'np26_ra_obligation' : np26_ra_obligation,
                'sn_path26_allocation' : sn_path26_allocation,
                'sp26_ra_obligation' : sp26_ra_obligation,
                'ns_path26_allocation' : ns_path26_allocation,
                'system_rmr_credit' : total_cam_rmr.loc['system_rmr'] * cam_rmr_monthly_tracking.loc[(organization_id,filing_month),'total_revised_jurisdictional_load_share'],
                'np26_cpe_system_cam' : np26_cpe_system_cam,
                'sp26_cpe_system_cam' : sp26_cpe_system_cam,
                'year_ahead_flex_rar_category1' : flexibility_requirements.loc[(organization_id,1,filing_month),'flexibility_requirement'] - flexibility_rmr.loc[(organization_id,filing_month),'flexibility_rmr'] - flexibility_cme.loc[(organization_id,filing_month),'flexibility_cme'],
                'year_ahead_flex_rar_category2' : flexibility_requirements.loc[(organization_id,2,filing_month),'flexibility_requirement'],
                'year_ahead_flex_rar_category3' : flexibility_requirements.loc[(organization_id,3,filing_month),'flexibility_requirement'],
                'year_ahead_flex_incremental_category1' : incremental_flex_by_category(1),
                'year_ahead_flex_incremental_category2' : incremental_flex_by_category(2),
                'year_ahead_flex_incremental_category3' : incremental_flex_by_category(3),
                'los_angeles_local_rar' : local_rar.loc[organization_id,'los_angeles'],
                'ventura_local_rar' : local_rar.loc[organization_id,'ventura'],
                'san_diego_local_rar' : local_rar.loc[organization_id,'san_diego'],
                'bay_area_local_rar' : local_rar.loc[organization_id,'bay_area'],
                'fresno_local_rar' : local_rar.loc[organization_id,'fresno'],
                'sierra_local_rar' : local_rar.loc[organization_id,'sierra'],
                'stockton_local_rar' : local_rar.loc[organization_id,'stockton'],
                'kern_local_rar' : local_rar.loc[organization_id,'kern'],
                'humboldt_local_rar' : local_rar.loc[organization_id,'humboldt'],
                'northern_california_local_rar' : local_rar.loc[organization_id,'northern_california'],
                'los_angeles_august_demand_response' : august_demand_response('SCE','los_angeles','prorated'),
                'ventura_august_demand_response' : august_demand_response('SCE','ventura','prorated'),
                'san_diego_august_demand_response' : august_demand_response('SDGE','san_diego','prorated'),
                'bay_area_august_demand_response' : august_demand_response('PGE','bay_area','prorated'),
                'fresno_august_demand_response' : august_demand_response('PGE','fresno','base'),
                'sierra_august_demand_response' : august_demand_response('PGE','sierra','base'),
                'stockton_august_demand_response' : august_demand_response('PGE','stockton','base'),
                'kern_august_demand_response' : august_demand_response('PGE','kern','base'),
                'humboldt_august_demand_response' : august_demand_response('PGE','humboldt','base'),
                'northern_california_august_demand_response' : august_demand_response('PGE','northern_california','base'),
                'los_angeles_incremental_load' : incremental_load_by_area('los_angeles'),
                'ventura_incremental_load' : incremental_load_by_area('ventura'),
                'san_diego_incremental_load' : incremental_load_by_area('san_diego'),
                'bay_area_incremental_load' : incremental_load_by_area('bay_area'),
                'fresno_incremental_load' : incremental_load_by_area('fresno'),
                'sierra_incremental_load' : incremental_load_by_area('sierra'),
                'stockton_incremental_load' : incremental_load_by_area('stockton'),
                'kern_incremental_load' : incremental_load_by_area('kern'),
                'humboldt_incremental_load' : incremental_load_by_area('humboldt'),
                'northern_california_incremental_load' : incremental_load_by_area('northern_california'),
            })
        summary = summary.merge(summary.apply(calculate_summary,axis='columns'),on='organization_id')

        #  copy summary dataframe to worksheet:
        row_number = 2
        def procurement(local_area:str,organization_id:str):
            if self.config.organizations.get_type(organization_id)=='investor-owned utility':
                procurement_str = '=SUMIFS(PhysicalResources!E:E,PhysicalResources!A:A,@INDIRECT("A"&ROW()),' + \
                    'PhysicalResources!K:K,"{0}")'
                return procurement_str.format(local_area)
            else:
                procurement_str = '=SUMIFS(PhysicalResources!E:E,PhysicalResources!A:A,@INDIRECT("A"&ROW()),' + \
                    'PhysicalResources!K:K,"{0}")+' + \
                    '{1}*SUMIFS(PhysicalResources!E:E,' + \
                    'PhysicalResources!A:A,@INDIRECT("A"&ROW()),' + \
                    'PhysicalResources!K:K,"{0}",' + \
                    'PhysicalResources!F:F,"DR")'
                if local_area=='San Diego-IV':
                    transmission_loss_adder = self.config.get_option('transmission_loss_adder_sdge')
                elif local_area in ['LA Basin','Big Creek-Ventura']:
                    transmission_loss_adder = self.config.get_option('transmission_loss_adder_sce')
                else:
                    transmission_loss_adder = self.config.get_option('transmission_loss_adder_pge')
                return procurement_str.format(local_area,round(transmission_loss_adder-1,4))
        compliance_check = '=IF(@INDIRECT(ADDRESS(ROW(),COLUMN()-1))+' + \
            '@INDIRECT(ADDRESS(ROW(),COLUMN()-2))-' + \
            '@INDIRECT(ADDRESS(ROW(),COLUMN()-4))-' + \
            '@INDIRECT(ADDRESS(ROW(),COLUMN()-3))>=0,' + \
            '"compliant",' + \
            '@INDIRECT(ADDRESS(ROW(),COLUMN()-1))+' + \
            '@INDIRECT(ADDRESS(ROW(),COLUMN()-2))-' + \
            '@INDIRECT(ADDRESS(ROW(),COLUMN()-4))-' + \
            '@INDIRECT(ADDRESS(ROW(),COLUMN()-3)))'
        for _,s in summary.iterrows():
            organization_id = s.loc['organization_id']
            ra_summary['Summary'][f'H{row_number}'].value = s.loc['system_rmr_credit']
            ra_summary['NP26'][f'A{row_number}'].value = organization_id
            ra_summary['NP26'][f'B{row_number}'].value = s.loc['np26_ra_obligation']
            ra_summary['NP26'][f'B{row_number}'].number_format = '0'
            ra_summary['NP26'][f'E{row_number}'].value = s.loc['sn_path26_allocation']
            ra_summary['NP26'][f'J{row_number}'].value = s.loc['np26_cpe_system_cam']
            ra_summary['SP26'][f'A{row_number}'].value = organization_id
            ra_summary['SP26'][f'B{row_number}'].value = s.loc['sp26_ra_obligation']
            ra_summary['SP26'][f'B{row_number}'].number_format = '0'
            ra_summary['SP26'][f'E{row_number}'].value = s.loc['ns_path26_allocation']
            ra_summary['SP26'][f'J{row_number}'].value = s.loc['sp26_cpe_system_cam']
            ra_summary['FlexRAR'][f'A{row_number}'].value = organization_id
            ra_summary['FlexRAR'][f'B{row_number}'] = '=@INDIRECT("E"&ROW())+@INDIRECT("G"&ROW())+@INDIRECT("I"&ROW())'
            ra_summary['FlexRAR'][f'C{row_number}'] = '=@INDIRECT("F"&ROW())+@INDIRECT("H"&ROW())+@INDIRECT("J"&ROW())'
            ra_summary['FlexRAR'][f'D{row_number}'] = '=IFERROR(@INDIRECT("C"&ROW())/@INDIRECT("B"&ROW()),0)'
            ra_summary['FlexRAR'][f'D{row_number}'].number_format = '0.00%'
            ra_summary['FlexRAR'][f'E{row_number}'] = '=ROUND(@INDIRECT("L"&ROW())+@INDIRECT("Q"&ROW()),0)'
            ra_summary['FlexRAR'][f'G{row_number}'] = '=ROUND(@INDIRECT("M"&ROW())+@INDIRECT("R"&ROW()),0)'
            ra_summary['FlexRAR'][f'H{row_number}'] = '=MIN(@INDIRECT("G"&ROW())+@INDIRECT("I"&ROW()),SUMIFS(PhysicalResources!$H:$H,PhysicalResources!$A:$A,@INDIRECT("A"&ROW()), PhysicalResources!$I:$I,2))'
            ra_summary['FlexRAR'][f'I{row_number}'] = '=ROUND(@INDIRECT("N"&ROW())+@INDIRECT("S"&ROW()),0)'
            ra_summary['FlexRAR'][f'J{row_number}'] = '=MIN(@INDIRECT("I"&ROW()),SUMIFS(PhysicalResources!$H:$H,PhysicalResources!$A:$A,@INDIRECT("A"&ROW()), PhysicalResources!$I:$I,3))'
            ra_summary['FlexRAR'][f'L{row_number}'].value = s.loc['year_ahead_flex_rar_category1']
            ra_summary['FlexRAR'][f'M{row_number}'].value = s.loc['year_ahead_flex_rar_category2']
            ra_summary['FlexRAR'][f'N{row_number}'].value = s.loc['year_ahead_flex_rar_category3']
            ra_summary['FlexRAR'][f'O{row_number}'] = '=@INDIRECT("L"&ROW())+@INDIRECT("M"&ROW())+@INDIRECT("N"&ROW())'
            ra_summary['FlexRAR'][f'Q{row_number}'].value = s.loc['year_ahead_flex_incremental_category1']
            ra_summary['FlexRAR'][f'R{row_number}'].value = s.loc['year_ahead_flex_incremental_category2']
            ra_summary['FlexRAR'][f'S{row_number}'].value = s.loc['year_ahead_flex_incremental_category3']
            ra_summary['FlexRAR'][f'T{row_number}'] = '=@INDIRECT("Q"&ROW())+@INDIRECT("R"&ROW())+@INDIRECT("S"&ROW())'
            ra_summary['MCC_Check'][f'A{row_number}'] = organization_id
            ra_summary['CertifyingOfficers'][f'A{row_number}'] = organization_id
            ra_summary['LocalTrueUp'][f'A{row_number}'].value = organization_id
            # los angeles basin local area:
            ra_summary['LocalTrueUp'][f'B{row_number}'].value = s.loc['los_angeles_local_rar']
            ra_summary['LocalTrueUp'][f'C{row_number}'].value = s.loc['los_angeles_incremental_load']
            ra_summary['LocalTrueUp'][f'D{row_number}'].value = s.loc['los_angeles_august_demand_response']
            ra_summary['LocalTrueUp'][f'E{row_number}'] = procurement('LA Basin',organization_id)
            ra_summary['LocalTrueUp'][f'F{row_number}'] = compliance_check
            # big creek and ventura local area:
            ra_summary['LocalTrueUp'][f'G{row_number}'].value = s.loc['ventura_local_rar']
            ra_summary['LocalTrueUp'][f'H{row_number}'].value = s.loc['ventura_incremental_load']
            ra_summary['LocalTrueUp'][f'I{row_number}'].value = s.loc['ventura_august_demand_response']
            ra_summary['LocalTrueUp'][f'J{row_number}'] = procurement('Big Creek-Ventura',organization_id)
            ra_summary['LocalTrueUp'][f'K{row_number}'] = compliance_check
            # san diego local area:
            ra_summary['LocalTrueUp'][f'L{row_number}'].value = s.loc['san_diego_local_rar']
            ra_summary['LocalTrueUp'][f'M{row_number}'].value = s.loc['san_diego_incremental_load']
            ra_summary['LocalTrueUp'][f'N{row_number}'].value = s.loc['san_diego_august_demand_response']
            ra_summary['LocalTrueUp'][f'O{row_number}'] = procurement('San Diego-IV',organization_id)
            ra_summary['LocalTrueUp'][f'P{row_number}'] = compliance_check
            # san francisco bay local area:
            ra_summary['LocalTrueUp'][f'Q{row_number}'].value = s.loc['bay_area_local_rar']
            ra_summary['LocalTrueUp'][f'R{row_number}'].value = s.loc['bay_area_incremental_load']
            ra_summary['LocalTrueUp'][f'S{row_number}'].value = s.loc['bay_area_august_demand_response']
            ra_summary['LocalTrueUp'][f'T{row_number}'] = procurement('Bay Area',organization_id)
            ra_summary['LocalTrueUp'][f'U{row_number}'] = compliance_check
            # humboldt county local area:
            ra_summary['LocalTrueUp'][f'V{row_number}'].value = s.loc['humboldt_local_rar']
            ra_summary['LocalTrueUp'][f'W{row_number}'].value = s.loc['humboldt_incremental_load']
            ra_summary['LocalTrueUp'][f'X{row_number}'].value = s.loc['humboldt_august_demand_response']
            ra_summary['LocalTrueUp'][f'Y{row_number}'] = procurement('Humboldt',organization_id)
            ra_summary['LocalTrueUp'][f'Z{row_number}'] = compliance_check
            # sierra local area:
            ra_summary['LocalTrueUp'][f'AA{row_number}'].value = s.loc['sierra_local_rar']
            ra_summary['LocalTrueUp'][f'AB{row_number}'].value = s.loc['sierra_incremental_load']
            ra_summary['LocalTrueUp'][f'AC{row_number}'].value = s.loc['sierra_august_demand_response']
            ra_summary['LocalTrueUp'][f'AD{row_number}'] = procurement('Sierra',organization_id)
            ra_summary['LocalTrueUp'][f'AE{row_number}'] = compliance_check
            # stockton local area:
            ra_summary['LocalTrueUp'][f'AF{row_number}'].value = s.loc['stockton_local_rar']
            ra_summary['LocalTrueUp'][f'AG{row_number}'].value = s.loc['stockton_incremental_load']
            ra_summary['LocalTrueUp'][f'AH{row_number}'].value = s.loc['stockton_august_demand_response']
            ra_summary['LocalTrueUp'][f'AI{row_number}'] = procurement('Stockton',organization_id)
            ra_summary['LocalTrueUp'][f'AJ{row_number}'] = compliance_check
            # northern california and north bay local area:
            ra_summary['LocalTrueUp'][f'AK{row_number}'].value = s.loc['northern_california_local_rar']
            ra_summary['LocalTrueUp'][f'AL{row_number}'].value = s.loc['northern_california_incremental_load']
            ra_summary['LocalTrueUp'][f'AM{row_number}'].value = s.loc['northern_california_august_demand_response']
            ra_summary['LocalTrueUp'][f'AN{row_number}'] = procurement('NCNB',organization_id)
            ra_summary['LocalTrueUp'][f'AO{row_number}'] = compliance_check
            # fresno local area:
            ra_summary['LocalTrueUp'][f'AP{row_number}'].value = s.loc['fresno_local_rar']
            ra_summary['LocalTrueUp'][f'AQ{row_number}'].value = s.loc['fresno_incremental_load']
            ra_summary['LocalTrueUp'][f'AR{row_number}'].value = s.loc['fresno_august_demand_response']
            ra_summary['LocalTrueUp'][f'AS{row_number}'] = procurement('Fresno',organization_id)
            ra_summary['LocalTrueUp'][f'AT{row_number}'] = compliance_check
            # kern county local area:
            ra_summary['LocalTrueUp'][f'AU{row_number}'].value = s.loc['kern_local_rar']
            ra_summary['LocalTrueUp'][f'AV{row_number}'].value = s.loc['kern_incremental_load']
            ra_summary['LocalTrueUp'][f'AW{row_number}'].value = s.loc['kern_august_demand_response']
            ra_summary['LocalTrueUp'][f'AX{row_number}'] = procurement('Kern',organization_id)
            ra_summary['LocalTrueUp'][f'AY{row_number}'] = compliance_check
            # pge other aggregated:
            ra_summary['LocalTrueUp'][f'AZ{row_number}'] = '=@INDIRECT("V"&ROW())+@INDIRECT("AA"&ROW())+@INDIRECT("AF"&ROW())+@INDIRECT("AK"&ROW())+@INDIRECT("AP"&ROW())+@INDIRECT("AU"&ROW())'
            ra_summary['LocalTrueUp'][f'BA{row_number}'] = '=@INDIRECT("W"&ROW())+@INDIRECT("AB"&ROW())+@INDIRECT("AG"&ROW())+@INDIRECT("AL"&ROW())+@INDIRECT("AQ"&ROW())+@INDIRECT("AV"&ROW())'
            ra_summary['LocalTrueUp'][f'BB{row_number}'] = '=@INDIRECT("X"&ROW())+@INDIRECT("AC"&ROW())+@INDIRECT("AH"&ROW())+@INDIRECT("AM"&ROW())+@INDIRECT("AR"&ROW())+@INDIRECT("AW"&ROW())'
            ra_summary['LocalTrueUp'][f'BC{row_number}'] = '=@INDIRECT("Y"&ROW())+@INDIRECT("AD"&ROW())+@INDIRECT("AI"&ROW())+@INDIRECT("AN"&ROW())+@INDIRECT("AS"&ROW())+@INDIRECT("AX"&ROW())'
            ra_summary['LocalTrueUp'][f'BD{row_number}'] = compliance_check
            # sce service territory:
            ra_summary['LocalTrueUp'][f'BE{row_number}'] = '=@INDIRECT("B"&ROW())+@INDIRECT("G"&ROW())'
            ra_summary['LocalTrueUp'][f'BF{row_number}'] = '=@INDIRECT("C"&ROW())+@INDIRECT("H"&ROW())'
            ra_summary['LocalTrueUp'][f'BG{row_number}'] = '=@INDIRECT("D"&ROW())+@INDIRECT("I"&ROW())'
            ra_summary['LocalTrueUp'][f'BH{row_number}'] = '=@INDIRECT("E"&ROW())+@INDIRECT("J"&ROW())'
            ra_summary['LocalTrueUp'][f'BI{row_number}'] = compliance_check
            # sdge service territory:
            ra_summary['LocalTrueUp'][f'BJ{row_number}'] = '=@INDIRECT("L"&ROW())'
            ra_summary['LocalTrueUp'][f'BK{row_number}'] = '=@INDIRECT("M"&ROW())'
            ra_summary['LocalTrueUp'][f'BL{row_number}'] = '=@INDIRECT("N"&ROW())'
            ra_summary['LocalTrueUp'][f'BM{row_number}'] = '=@INDIRECT("O"&ROW())'
            ra_summary['LocalTrueUp'][f'BN{row_number}'] = compliance_check
            # pge service territory:
            ra_summary['LocalTrueUp'][f'BO{row_number}'] = '=@INDIRECT("Q"&ROW())+@INDIRECT("V"&ROW())+@INDIRECT("AA"&ROW())+@INDIRECT("AF"&ROW())+@INDIRECT("AK"&ROW())+@INDIRECT("AP"&ROW())+@INDIRECT("AU"&ROW())'
            ra_summary['LocalTrueUp'][f'BP{row_number}'] = '=@INDIRECT("R"&ROW())+@INDIRECT("W"&ROW())+@INDIRECT("AB"&ROW())+@INDIRECT("AG"&ROW())+@INDIRECT("AL"&ROW())+@INDIRECT("AQ"&ROW())+@INDIRECT("AV"&ROW())'
            ra_summary['LocalTrueUp'][f'BQ{row_number}'] = '=@INDIRECT("S"&ROW())+@INDIRECT("X"&ROW())+@INDIRECT("AC"&ROW())+@INDIRECT("AH"&ROW())+@INDIRECT("AM"&ROW())+@INDIRECT("AR"&ROW())+@INDIRECT("AW"&ROW())'
            ra_summary['LocalTrueUp'][f'BR{row_number}'] = '=@INDIRECT("T"&ROW())+@INDIRECT("Y"&ROW())+@INDIRECT("AD"&ROW())+@INDIRECT("AI"&ROW())+@INDIRECT("AN"&ROW())+@INDIRECT("AS"&ROW())+@INDIRECT("AX"&ROW())'
            ra_summary['LocalTrueUp'][f'BS{row_number}'] = compliance_check

            # caiso supply plan cross-check:
            caiso_cross_check['Requirements'][f'A{row_number+1}'] = organization_id
            caiso_cross_check['Requirements'][f'B{row_number+1}'].value = s.loc[['np26_ra_obligation','sp26_ra_obligation']].sum()
            caiso_cross_check['Requirements'][f'B{row_number+1}'].number_format = '0'
            caiso_cross_check['Requirements'][f'G{row_number+1}'] = s.loc['np26_cpe_system_cam'] + s.loc['sp26_cpe_system_cam']
            caiso_cross_check['Requirements'][f'H{row_number+1}'] = s.loc['system_rmr_credit']
            caiso_cross_check['Requirements'][f'J{row_number+1}'] = '=@INDIRECT("M"&ROW())+@INDIRECT("O"&ROW())+@INDIRECT("Q"&ROW())'
            caiso_cross_check['Requirements'][f'K{row_number+1}'] = '=@INDIRECT("N"&ROW())+@INDIRECT("P"&ROW())+@INDIRECT("R"&ROW())'
            caiso_cross_check['Requirements'][f'L{row_number+1}'] = '=IFERROR(@INDIRECT("K"&ROW())/@INDIRECT("J"&ROW()),0)'
            caiso_cross_check['Requirements'][f'L{row_number+1}'].number_format = '0.00%'
            caiso_cross_check['Requirements'][f'M{row_number+1}'].value = s.loc[['year_ahead_flex_rar_category1','year_ahead_flex_incremental_category1']].sum()
            caiso_cross_check['Requirements'][f'O{row_number+1}'].value = s.loc[['year_ahead_flex_rar_category2','year_ahead_flex_incremental_category2']].sum()
            caiso_cross_check['Requirements'][f'Q{row_number+1}'].value = s.loc[['year_ahead_flex_rar_category3','year_ahead_flex_incremental_category3']].sum()
            caiso_cross_check['Requirements'][f'S{row_number+1}'].value = s.loc['year_ahead_flex_rar_category1']
            caiso_cross_check['Requirements'][f'T{row_number+1}'].value = s.loc['year_ahead_flex_rar_category2']
            caiso_cross_check['Requirements'][f'U{row_number+1}'].value = s.loc['year_ahead_flex_rar_category3']
            caiso_cross_check['Requirements'][f'V{row_number+1}'].value = s.loc[['year_ahead_flex_rar_category1','year_ahead_flex_rar_category2','year_ahead_flex_rar_category3']].sum()
            # los angeles basin local area:
            caiso_cross_check['Requirements'][f'X{row_number+1}'].value = s.loc['los_angeles_local_rar']
            caiso_cross_check['Requirements'][f'Y{row_number+1}'].value = s.loc['los_angeles_incremental_load']
            caiso_cross_check['Requirements'][f'Z{row_number+1}'].value = s.loc['los_angeles_august_demand_response']
            caiso_cross_check['Requirements'][f'AB{row_number+1}'] = compliance_check
            # big creek and ventura local area:
            caiso_cross_check['Requirements'][f'AC{row_number+1}'].value = s.loc['ventura_local_rar']
            caiso_cross_check['Requirements'][f'AD{row_number+1}'].value = s.loc['ventura_incremental_load']
            caiso_cross_check['Requirements'][f'AE{row_number+1}'].value = s.loc['ventura_august_demand_response']
            caiso_cross_check['Requirements'][f'AG{row_number+1}'] = compliance_check
            # san diego local area:
            caiso_cross_check['Requirements'][f'AH{row_number+1}'].value = s.loc['san_diego_local_rar']
            caiso_cross_check['Requirements'][f'AI{row_number+1}'].value = s.loc['san_diego_incremental_load']
            caiso_cross_check['Requirements'][f'AJ{row_number+1}'].value = s.loc['san_diego_august_demand_response']
            caiso_cross_check['Requirements'][f'AL{row_number+1}'] = compliance_check
            # san francisco bay local area:
            caiso_cross_check['Requirements'][f'AM{row_number+1}'].value = s.loc['bay_area_local_rar']
            caiso_cross_check['Requirements'][f'AN{row_number+1}'].value = s.loc['bay_area_incremental_load']
            caiso_cross_check['Requirements'][f'AO{row_number+1}'].value = s.loc['bay_area_august_demand_response']
            caiso_cross_check['Requirements'][f'AQ{row_number+1}'] = compliance_check
            # humboldt county local area:
            caiso_cross_check['Requirements'][f'AR{row_number+1}'].value = s.loc['humboldt_local_rar']
            caiso_cross_check['Requirements'][f'AS{row_number+1}'].value = s.loc['humboldt_incremental_load']
            caiso_cross_check['Requirements'][f'AT{row_number+1}'].value = s.loc['humboldt_august_demand_response']
            caiso_cross_check['Requirements'][f'AV{row_number+1}'] = compliance_check
            # sierra local area:
            caiso_cross_check['Requirements'][f'AW{row_number+1}'].value = s.loc['sierra_local_rar']
            caiso_cross_check['Requirements'][f'AX{row_number+1}'].value = s.loc['sierra_incremental_load']
            caiso_cross_check['Requirements'][f'AY{row_number+1}'].value = s.loc['sierra_august_demand_response']
            caiso_cross_check['Requirements'][f'BA{row_number+1}'] = compliance_check
            # stockton local area:
            caiso_cross_check['Requirements'][f'BB{row_number+1}'].value = s.loc['stockton_local_rar']
            caiso_cross_check['Requirements'][f'BC{row_number+1}'].value = s.loc['stockton_incremental_load']
            caiso_cross_check['Requirements'][f'BD{row_number+1}'].value = s.loc['stockton_august_demand_response']
            caiso_cross_check['Requirements'][f'BF{row_number+1}'] = compliance_check
            # northern california and north bay local area:
            caiso_cross_check['Requirements'][f'BG{row_number+1}'].value = s.loc['northern_california_local_rar']
            caiso_cross_check['Requirements'][f'BH{row_number+1}'].value = s.loc['northern_california_incremental_load']
            caiso_cross_check['Requirements'][f'BI{row_number+1}'].value = s.loc['northern_california_august_demand_response']
            caiso_cross_check['Requirements'][f'BK{row_number+1}'] = compliance_check
            # fresno local area:
            caiso_cross_check['Requirements'][f'BL{row_number+1}'].value = s.loc['fresno_local_rar']
            caiso_cross_check['Requirements'][f'BM{row_number+1}'].value = s.loc['fresno_incremental_load']
            caiso_cross_check['Requirements'][f'BN{row_number+1}'].value = s.loc['fresno_august_demand_response']
            caiso_cross_check['Requirements'][f'BP{row_number+1}'] = compliance_check
            # kern county local area:
            caiso_cross_check['Requirements'][f'BQ{row_number+1}'].value = s.loc['kern_local_rar']
            caiso_cross_check['Requirements'][f'BR{row_number+1}'].value = s.loc['kern_incremental_load']
            caiso_cross_check['Requirements'][f'BS{row_number+1}'].value = s.loc['kern_august_demand_response']
            caiso_cross_check['Requirements'][f'BU{row_number+1}'] = compliance_check
            # pge other aggregated:
            caiso_cross_check['Requirements'][f'BV{row_number+1}'].value = s.loc[['humboldt_local_rar','sierra_local_rar','stockton_local_rar','northern_california_local_rar','fresno_local_rar','kern_local_rar']].sum()
            caiso_cross_check['Requirements'][f'BW{row_number+1}'].value = s.loc[['humboldt_incremental_load','sierra_incremental_load','stockton_incremental_load','northern_california_incremental_load','fresno_incremental_load','kern_incremental_load']].sum()
            caiso_cross_check['Requirements'][f'BX{row_number+1}'].value = s.loc[['humboldt_august_demand_response','sierra_august_demand_response','stockton_august_demand_response','northern_california_august_demand_response','fresno_august_demand_response','kern_august_demand_response']].sum()
            caiso_cross_check['Requirements'][f'BZ{row_number+1}'] = compliance_check
            # sce service territory:
            caiso_cross_check['Requirements'][f'CA{row_number+1}'] = s.loc[['los_angeles_local_rar','ventura_local_rar']].sum()
            caiso_cross_check['Requirements'][f'CB{row_number+1}'] = s.loc[['los_angeles_incremental_load','ventura_incremental_load']].sum()
            caiso_cross_check['Requirements'][f'CC{row_number+1}'] = s.loc[['los_angeles_august_demand_response','ventura_august_demand_response']].sum()
            caiso_cross_check['Requirements'][f'CE{row_number+1}'] = compliance_check
            # sdge service territory:
            caiso_cross_check['Requirements'][f'CF{row_number+1}'] = s.loc['san_diego_local_rar']
            caiso_cross_check['Requirements'][f'CG{row_number+1}'] = s.loc['san_diego_incremental_load']
            caiso_cross_check['Requirements'][f'CH{row_number+1}'] = s.loc['san_diego_august_demand_response']
            caiso_cross_check['Requirements'][f'CJ{row_number+1}'] = compliance_check
            # pge service territory:
            caiso_cross_check['Requirements'][f'CK{row_number+1}'] = s.loc[['bay_area_local_rar','humboldt_local_rar','sierra_local_rar','stockton_local_rar','northern_california_local_rar','fresno_local_rar','kern_local_rar']].sum()
            caiso_cross_check['Requirements'][f'CL{row_number+1}'] = s.loc[['bay_area_incremental_load','humboldt_incremental_load','sierra_incremental_load','stockton_incremental_load','northern_california_incremental_load','fresno_incremental_load','kern_incremental_load']].sum()
            caiso_cross_check['Requirements'][f'CM{row_number+1}'] = s.loc[['bay_area_august_demand_response','humboldt_august_demand_response','sierra_august_demand_response','stockton_august_demand_response','northern_california_august_demand_response','fresno_august_demand_response','kern_august_demand_response']].sum()
            caiso_cross_check['Requirements'][f'CO{row_number+1}'] = compliance_check

            row_number += 1
        self.consolidation_logger.commit()

        # total rows on local true-up sheet:
        ra_summary['LocalTrueUp'][f'A{row_number}'] = 'Total:'
        for col in map(get_column_letter,range(2,72)):
            ra_summary['LocalTrueUp'][f'{col}{row_number}'] = '=SUM(INDIRECT(ADDRESS(2,COLUMN())&":"&ADDRESS(ROW()-1,COLUMN())))'

        # save and close summary and caiso supply plan cross-check files:
        ra_summary.save(str(self.config.paths.get_path('ra_summary')))
        ra_summary.close()
        caiso_cross_check.save(str(self.config.paths.get_path('caiso_cross_check')))
        caiso_cross_check.close()

        # check time and report:
        run_time = (ts.now() - init_time).total_seconds()
        self.logger.log('Retrieved Allocations in {:02.0f}:{:02.0f}:{:05.2f}'.format(int(run_time/3600),int((run_time%3600)/60),run_time%60),'INFORMATION')

    # collect data from each of the monthly lse filings:
    def consolidate_filings(self):
        '''
        loads data from each ra monthly filing workbook and populates the
        summary and caiso cross-check workbooks with data and calculation
        results.
        '''
        # start timer:
        init_time = ts.now()
        self.logger.log('Consolidating Data from LSE Filings','INFORMATION')

        filing_month = self.config.filing_month

        # get list of active load serving entities from summary sheet:
        path = self.config.paths.get_path('ra_summary')
        ra_summary = open_workbook(path,data_only=False,read_only=False)
        data_range = get_data_range(ra_summary['Summary'],'A','',self.config)
        active_organizations = [row[0].value for row in data_range]

        # open summary from previous month:
        path = self.config.paths.get_path('ra_summary_previous_month')
        ra_summary_previous_month = open_workbook(path,data_only=False,read_only=True)

        # open caiso supply plan cross-check file:
        path = self.config.paths.get_path('caiso_cross_check')
        caiso_cross_check = open_workbook(path,data_only=False,read_only=False)

        # load nqc list from most recent file for current year:
        path = self.config.paths.get_path('nqc_list')
        nqc_workbook = open_workbook(path,data_only=True,in_mem=True)
        nqc_list = get_nqc_list(nqc_workbook,self.config)

        # set the nqc list index:
        nqc_list.set_index('resource_id',inplace=True)
        nqc_list.sort_index(inplace=True)

        # initialize tables:
        summary = pd.DataFrame()
        physical_resources = pd.DataFrame()
        demand_response = pd.DataFrame()

        # combine data tables from each lse filing:
        for organization_id in active_organizations:
            organization = self.config.organizations.get_organization(organization_id)
            monthly_filing_tables = read_ra_monthly_filing(organization,self.config,self.logger)
            ra_monthly_filing_summary = monthly_filing_tables[0]
            ra_monthly_filing_physical_resources = monthly_filing_tables[1]
            ra_monthly_filing_demand_response = monthly_filing_tables[2]

            # append summary, physical resources, and demand response tables with lse-specific data:
            summary = pd.concat([summary,ra_monthly_filing_summary],axis='index',ignore_index=True)
            physical_resources = pd.concat([physical_resources,ra_monthly_filing_physical_resources],axis='index',ignore_index=True)
            demand_response = pd.concat([demand_response,ra_monthly_filing_demand_response],axis='index',ignore_index=True)

        # set summary table index:
        summary.set_index('organization_id',inplace=True)
        summary.sort_index(inplace=True)

        # helper function to retrieve zone for a given resource id:
        def get_zone(resource_id: str):
            if resource_id in nqc_list.index:
                zone = nqc_list.loc[resource_id,'zone']
            else:
                zone = 'Unknown'
            return zone

        # helper function to retrieve local area for a given resource id:
        def get_local_area(resource_id: str):
            if resource_id in nqc_list.index:
                local_area = nqc_list.loc[resource_id,'local_area']
            else:
                local_area = 'Unknown'
            return local_area

        # lookup local areas for physical resources and demand response programs:
        physical_resources.loc[:,'local_area'] = physical_resources.loc[:,'resource_id'].map(get_local_area)
        physical_resources.loc[:,'locality'] = physical_resources.loc[:,'local_area'].map(rename_locality)
        demand_response.loc[:,'local_area'] = demand_response.loc[:,'program_id'].map(get_local_area)
        demand_response.loc[:,'locality'] = demand_response.loc[:,'local_area'].map(rename_locality)

        # get certifications from previous month:
        columns = [
            'organization_id',
            'organization_officer_name',
            'organization_officer_title',
            'filename'
        ]
        try:
            data_range = get_data_range(ra_summary_previous_month['CertifyingOfficers'],'A','BCH',self.config)
            previous_certifications = data_range_to_dataframe(columns,data_range)
        except:
            previous_certifications = pd.DataFrame(columns=columns)
        previous_certifications.set_index('organization_id',inplace=True)
        previous_certifications.sort_index(inplace=True)

        # write nqc list to summary file:
        row_number = 2
        for resource_id,r in nqc_list.iterrows():
            ra_summary['NQC_List'][f'A{row_number}'].value = r.loc['generator_name']
            ra_summary['NQC_List'][f'B{row_number}'].value = resource_id
            ra_summary['NQC_List'][f'C{row_number}'].value = r.loc['zone']
            ra_summary['NQC_List'][f'D{row_number}'].value = r.loc['local_area']
            ra_summary['NQC_List'][f'E{row_number}'].value = r.loc[ts(filing_month.year,1,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'F{row_number}'].value = r.loc[ts(filing_month.year,2,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'G{row_number}'].value = r.loc[ts(filing_month.year,3,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'H{row_number}'].value = r.loc[ts(filing_month.year,4,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'I{row_number}'].value = r.loc[ts(filing_month.year,5,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'J{row_number}'].value = r.loc[ts(filing_month.year,6,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'K{row_number}'].value = r.loc[ts(filing_month.year,7,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'L{row_number}'].value = r.loc[ts(filing_month.year,8,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'M{row_number}'].value = r.loc[ts(filing_month.year,9,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'N{row_number}'].value = r.loc[ts(filing_month.year,10,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'O{row_number}'].value = r.loc[ts(filing_month.year,11,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'P{row_number}'].value = r.loc[ts(filing_month.year,12,1).to_numpy().astype('datetime64[M]')]
            ra_summary['NQC_List'][f'Q{row_number}'].value = r.loc['dispatchable']
            ra_summary['NQC_List'][f'R{row_number}'].value = r.loc['deliverability_status']
            ra_summary['NQC_List'][f'S{row_number}'].value = r.loc['deliverable']
            ra_summary['NQC_List'][f'T{row_number}'].value = r.loc['comments']
            row_number+=1

        # write consolidated data and formulas to summary and caiso check files:
        first_row_number_summary = 2
        row_number_summary = first_row_number_summary
        first_row_number_physical_resources = 2
        row_number_physical_resources = first_row_number_physical_resources
        for organization_id in active_organizations:
            ra_summary['NP26'][f'C{row_number_summary}'] = '=SUMIFS(PhysicalResources!D:D,PhysicalResources!A:A,@INDIRECT("A"&ROW()),PhysicalResources!J:J,"North",PhysicalResources!F:F,"<>DR")'
            ra_summary['NP26'][f'C{row_number_summary}'].fill = PatternFill(start_color='FFCC99',end_color='FFCC99',fill_type='solid')
            ra_summary['NP26'][f'C{row_number_summary}'].number_format = '0.00'
            ra_summary['NP26'][f'D{row_number_summary}'] = '=ROUND(@INDIRECT("H"&ROW()),2)'
            ra_summary['NP26'][f'D{row_number_summary}'].number_format = '0.00'
            ra_summary['NP26'][f'E{row_number_summary}'].value = 0
            ra_summary['NP26'][f'E{row_number_summary}'].number_format = '0.00'
            ra_summary['NP26'][f'F{row_number_summary}'] = '=MAX(@INDIRECT("B"&ROW())-@INDIRECT("C"&ROW())-@INDIRECT("D"&ROW())-@INDIRECT("E"&ROW()),0)'
            ra_summary['NP26'][f'F{row_number_summary}'].fill = PatternFill(start_color='CC99FF',end_color='CC99FF',fill_type='solid')
            ra_summary['NP26'][f'F{row_number_summary}'].number_format = '0.00'
            ra_summary['NP26'][f'G{row_number_summary}'].value = summary.loc[organization_id,'np26dr']
            ra_summary['NP26'][f'G{row_number_summary}'].number_format = '0.00'
            ra_summary['NP26'][f'H{row_number_summary}'] = '={}*@INDIRECT("G"&ROW())'.format(self.config.get_option('demand_response_multiplier'))
            ra_summary['NP26'][f'H{row_number_summary}'].number_format = '0.00'
            ra_summary['NP26'][f'I{row_number_summary}'] = '=@INDIRECT("C"&ROW())+@INDIRECT("D"&ROW())'
            ra_summary['NP26'][f'I{row_number_summary}'].number_format = '0.00'
            ra_summary['NP26'][f'L{row_number_summary}'] = '=@INDIRECT("C"&ROW())+@INDIRECT("D"&ROW())-@INDIRECT("B"&ROW())'
            ra_summary['NP26'][f'L{row_number_summary}'].number_format = '0.00'
            ra_summary['NP26'][f'M{row_number_summary}'] = '=@INDIRECT("K"&ROW())+@INDIRECT("SP26!K"&ROW())'
            ra_summary['NP26'][f'M{row_number_summary}'].number_format = '0.00'
            ra_summary['NP26'][f'N{row_number_summary}'] = '=@INDIRECT("L"&ROW())-@INDIRECT("SP26!L"&ROW())'
            ra_summary['NP26'][f'N{row_number_summary}'].number_format = '0.00'
            ra_summary['SP26'][f'C{row_number_summary}'] = '=SUMIFS(PhysicalResources!D:D,PhysicalResources!A:A,@INDIRECT("A"&ROW()),PhysicalResources!J:J,"South",PhysicalResources!F:F, "<>DR")'
            ra_summary['SP26'][f'C{row_number_summary}'].fill = PatternFill(start_color='FFCC99',end_color='FFCC99',fill_type='solid')
            ra_summary['SP26'][f'C{row_number_summary}'].number_format = '0.00'
            ra_summary['SP26'][f'D{row_number_summary}'] = '=ROUND(@INDIRECT("H"&ROW()),2)'
            ra_summary['SP26'][f'D{row_number_summary}'].number_format = '0.00'
            ra_summary['SP26'][f'E{row_number_summary}'].value = 0
            ra_summary['SP26'][f'E{row_number_summary}'].number_format = '0.00'
            ra_summary['SP26'][f'F{row_number_summary}'] = '=MAX(@INDIRECT("B"&ROW())-@INDIRECT("C"&ROW())-@INDIRECT("D"&ROW())-@INDIRECT("E"&ROW()),0)'
            ra_summary['SP26'][f'F{row_number_summary}'].fill = PatternFill(start_color='CC99FF',end_color='CC99FF',fill_type='solid')
            ra_summary['SP26'][f'F{row_number_summary}'].number_format = '0.00'
            ra_summary['SP26'][f'G{row_number_summary}'] = summary.loc[organization_id,'sp26dr']
            ra_summary['SP26'][f'G{row_number_summary}'].number_format = '0.00'
            ra_summary['SP26'][f'H{row_number_summary}'] = '={}*@INDIRECT("G"&ROW())'.format(self.config.get_option('demand_response_multiplier'))
            ra_summary['SP26'][f'H{row_number_summary}'].number_format = '0.00'
            ra_summary['SP26'][f'I{row_number_summary}'] = '=@INDIRECT("C"&ROW())+@INDIRECT("D"&ROW())'
            ra_summary['SP26'][f'I{row_number_summary}'].number_format = '0.00'
            ra_summary['SP26'][f'L{row_number_summary}'] = '=@INDIRECT("C"&ROW())+@INDIRECT("D"&ROW())-@INDIRECT("B"&ROW())'
            ra_summary['SP26'][f'L{row_number_summary}'].number_format = '0.00'
            ra_summary['SP26'][f'M{row_number_summary}'] = '=@INDIRECT("K"&ROW())+@INDIRECT("NP26!K"&ROW())'
            ra_summary['SP26'][f'M{row_number_summary}'].number_format = '0.00'
            ra_summary['MCC_Check'][f'B{row_number_summary}'] = '=VLOOKUP(@INDIRECT("A"&ROW()),Summary!$A:$B,2,FALSE)'
            ra_summary['MCC_Check'][f'C{row_number_summary}'] = '=@INDIRECT("F"&ROW())+@INDIRECT("L"&ROW())+@INDIRECT("N"&ROW())'
            ra_summary['MCC_Check'][f'D{row_number_summary}'] = '=IFERROR(@INDIRECT("C"&ROW())/@INDIRECT("B"&ROW()),"")'
            ra_summary['MCC_Check'][f'E{row_number_summary}'] = '=@INDIRECT("B"&ROW())*MCC_Parameters!$B$2'
            ra_summary['MCC_Check'][f'F{row_number_summary}'] = '=MIN(@INDIRECT("E"&ROW()),@INDIRECT("NP26!H"&ROW())+@INDIRECT("SP26!H"&ROW()))'
            ra_summary['MCC_Check'][f'G{row_number_summary}'] = '=@INDIRECT("B"&ROW())*MCC_Parameters!$B$3'
            ra_summary['MCC_Check'][f'H{row_number_summary}'] = '=MIN(@INDIRECT("G"&ROW()),SUMIFS(PhysicalResources!$D:$D,PhysicalResources!$A:$A,@INDIRECT("A"&ROW()),PhysicalResources!$F:$F,1))'
            ra_summary['MCC_Check'][f'I{row_number_summary}'] = '=@INDIRECT("B"&ROW())*MCC_Parameters!$B$4'
            ra_summary['MCC_Check'][f'J{row_number_summary}'] = '=MIN(@INDIRECT("I"&ROW()),SUMIFS(PhysicalResources!$D:$D,PhysicalResources!$A:$A,@INDIRECT("A"&ROW()),PhysicalResources!$F:$F,2)+@INDIRECT("H"&ROW()))'
            ra_summary['MCC_Check'][f'K{row_number_summary}'] = '=@INDIRECT("B"&ROW())*MCC_Parameters!$B$5'
            ra_summary['MCC_Check'][f'L{row_number_summary}'] = '=MIN(@INDIRECT("K"&ROW()),SUMIFS(PhysicalResources!$D:$D,PhysicalResources!$A:$A,@INDIRECT("A"&ROW()),PhysicalResources!$F:$F,3)+@INDIRECT("J"&ROW()))'
            ra_summary['MCC_Check'][f'M{row_number_summary}'] = '=@INDIRECT("B"&ROW())*MCC_Parameters!$B$6'
            ra_summary['MCC_Check'][f'N{row_number_summary}'] = '=SUMIFS(PhysicalResources!$D:$D,PhysicalResources!$A:$A,@INDIRECT("A"&ROW()),PhysicalResources!$F:$F,4)'
            ra_summary['MCC_Check'][f'O{row_number_summary}'] = '=@INDIRECT("B"&ROW())*MCC_Parameters!$B$7'
            ra_summary['MCC_Check'][f'P{row_number_summary}'] = '=SUMIFS(PhysicalResources!$D:$D,PhysicalResources!$A:$A,@INDIRECT("A"&ROW()),PhysicalResources!$F:$F,4,PhysicalResources!$G:$G,TRUE)'
            ra_summary['FlexRAR'][f'F{row_number_summary}'] = '=SUMIFS(PhysicalResources!$H:$H,PhysicalResources!$A:$A,@INDIRECT("A"&ROW()),PhysicalResources!$I:$I,1)'
            ra_summary['CertifyingOfficers'][f'B{row_number_summary}'] = summary.loc[organization_id,'organization_officer_name']
            ra_summary['CertifyingOfficers'][f'C{row_number_summary}'] = summary.loc[organization_id,'organization_officer_title']
            ra_summary['CertifyingOfficers'][f'E{row_number_summary}'] = '=IF(@INDIRECT("D"&ROW())=@INDIRECT("B"&ROW()),"-","Yes")'
            ra_summary['CertifyingOfficers'][f'G{row_number_summary}'] = '=IF(@INDIRECT("F"&ROW())=@INDIRECT("C"&ROW()),"-","Yes")'
            if organization_id in previous_certifications.index:
                ra_summary['CertifyingOfficers'][f'D{row_number_summary}'] = previous_certifications.loc[organization_id,'organization_officer_name']
                ra_summary['CertifyingOfficers'][f'F{row_number_summary}'] = previous_certifications.loc[organization_id,'organization_officer_title']
                ra_summary['CertifyingOfficers'][f'I{row_number_summary}'] = previous_certifications.loc[organization_id,'filename']
            else:
                ra_summary['CertifyingOfficers'][f'D{row_number_summary}'] = '[Record Not Found]'
                ra_summary['CertifyingOfficers'][f'F{row_number_summary}'] = '[Record Not Found]'
                ra_summary['CertifyingOfficers'][f'I{row_number_summary}'] = '[Record Not Found]'
            if self.config.paths.get_path('ra_monthly_filing',self.config.organizations.get_organization(organization_id)):
                filename = self.config.paths.get_path('ra_monthly_filing',self.config.organizations.get_organization(organization_id)).name
            else:
                filename = 'Monthly Filing Not Found'
            ra_summary['Summary'][f'J{row_number_summary}'] = filename
            ra_summary['NP26'][f'N{row_number_summary}'] = filename
            ra_summary['SP26'][f'M{row_number_summary}'] = filename
            ra_summary['CertifyingOfficers'][f'H{row_number_summary}'] = filename
            caiso_cross_check['Requirements'][f'C{row_number_summary+1}'].value = physical_resources.loc[(physical_resources.loc[:,'organization_id']==organization_id),'resource_adequacy_system'].sum() - physical_resources.loc[(physical_resources.loc[:,'organization_id']==organization_id)&(physical_resources.loc[:,'resource_mcc_bucket']=='DR'),'resource_adequacy_system'].sum()
            caiso_cross_check['Requirements'][f'C{row_number_summary+1}'].number_format = '0.00'
            caiso_cross_check['Requirements'][f'D{row_number_summary+1}'].value = (self.config.get_option('demand_response_multiplier')) * (summary.loc[organization_id,'np26dr'] + summary.loc[organization_id,'sp26dr'])
            caiso_cross_check['Requirements'][f'D{row_number_summary+1}'].number_format = '0.00'
            caiso_cross_check['Requirements'][f'E{row_number_summary+1}'] = '=@INDIRECT("C"&ROW())+@INDIRECT("D"&ROW())'
            caiso_cross_check['Requirements'][f'E{row_number_summary+1}'].number_format = '0.00'
            caiso_cross_check['Requirements'][f'F{row_number_summary+1}'] = '=IFERROR(ROUND(@INDIRECT("E"&ROW())/@INDIRECT("B"&ROW()),2),"")'
            caiso_cross_check['Requirements'][f'F{row_number_summary+1}'].number_format = '0.00%'
            caiso_cross_check['Requirements'][f'N{row_number_summary+1}'].value = physical_resources.loc[(physical_resources.loc[:,'organization_id']==organization_id)&(physical_resources.loc[:,'resource_adequacy_flexibility_category']==1),'resource_adequacy_committed_flexible'].sum()
            caiso_cross_check['Requirements'][f'P{row_number_summary+1}'].value = min(
                caiso_cross_check['Requirements'][f'O{row_number_summary+1}'].value +
                caiso_cross_check['Requirements'][f'Q{row_number_summary+1}'].value,
                physical_resources.loc[(physical_resources.loc[:,'organization_id']==organization_id)&(physical_resources.loc[:,'resource_adequacy_flexibility_category']==2),'resource_adequacy_committed_flexible'].sum()
            )
            caiso_cross_check['Requirements'][f'R{row_number_summary+1}'].value = min(
                caiso_cross_check['Requirements'][f'Q{row_number_summary+1}'].value,
                physical_resources.loc[(physical_resources.loc[:,'organization_id']==organization_id)&(physical_resources.loc[:,'resource_adequacy_flexibility_category']==3),'resource_adequacy_committed_flexible'].sum()
            )
            # calculate values for caiso cross-check file because excel
            # calculations aren't available from calculated fields in summary file:
            if self.config.organizations.get_organization(organization_id)['type']=='investor-owned utility':
                demand_response_coefficient = 1
            else:
                demand_response_coefficient = self.config.get_option('demand_response_multiplier')
            localities = [
                ['AA{}','los_angeles','sce'],
                ['AF{}','ventura','sce'],
                ['AK{}','san_diego','sdge'],
                ['AP{}','bay_area','pge'],
                ['AU{}','humboldt','pge'],
                ['AZ{}','sierra','pge'],
                ['BE{}','stockton','pge'],
                ['BJ{}','northern_california','pge'],
                ['BO{}','fresno','pge'],
                ['BT{}','kern','pge'],
            ]
            for cell_address,locality,service_territory in localities:
                if organization_id in ['pge','sce','sdge']:
                    transmission_loss_adder = 1
                else:
                    transmission_loss_adder = round(self.config.get_option(f'transmission_loss_adder_{service_territory}'),5)
                caiso_cross_check['Requirements'][cell_address.format(row_number_summary+1)].value = \
                    physical_resources.loc[
                        (physical_resources.loc[:,'organization_id']==organization_id) &
                        (physical_resources.loc[:,'locality']==locality),
                        'resource_adequacy_local'
                    ].sum() + \
                    transmission_loss_adder * \
                    demand_response.loc[
                        (demand_response.loc[:,'organization_id']==organization_id) &
                        (demand_response.loc[:,'locality']==locality),
                        'resource_adequacy_local'
                    ].sum()
            regions = [
                ['BY{}',['humboldt','sierra','stockton','northern_california','fresno','kern'],'pge'],
                ['CD{}',['los_angeles','ventura'],'sce'],
                ['CI{}',['san_diego'],'sdge'],
                ['CN{}',['bay_area','humboldt','sierra','stockton','northern_california','fresno','kern'],'pge'],
            ]
            for cell_address,localities,service_territory in regions:
                if organization_id in ['pge','sce','sdge']:
                    transmission_loss_adder = 1
                else:
                    transmission_loss_adder = round(self.config.get_option(f'transmission_loss_adder_{service_territory}'),5)
                caiso_cross_check['Requirements'][cell_address.format(row_number_summary+1)].value = \
                    physical_resources.loc[
                        (physical_resources.loc[:,'organization_id']==organization_id) &
                        (physical_resources.loc[:,'locality'].isin(localities)),
                        'resource_adequacy_local'
                    ].sum() + \
                    transmission_loss_adder * \
                    demand_response.loc[
                        (demand_response.loc[:,'organization_id']==organization_id) &
                        (demand_response.loc[:,'locality'].isin(localities)),
                        'resource_adequacy_local'
                    ].sum()
            caiso_cross_check['FilingSummary'][f'A{row_number_summary}'] = organization_id
            caiso_cross_check['FilingSummary'][f'B{row_number_summary}'] = '=SUMIF(Filings!$A:$A,@INDIRECT("A"&ROW()),Filings!$G:$G)'
            caiso_cross_check['FilingSummary'][f'C{row_number_summary}'] = '=SUMIF(CAISO_Sys_SP!$C:$C,@INDIRECT("A"&ROW()),CAISO_Sys_SP!$I:$I)'
            caiso_cross_check['FilingSummary'][f'D{row_number_summary}'] = '=@INDIRECT("C"&ROW())-@INDIRECT("B"&ROW())'
            caiso_cross_check['FilingSummary'][f'D{row_number_summary}'].fill = PatternFill(start_color='FFF2CC',end_color='FFF2CC',fill_type='solid')
            caiso_cross_check['FilingSummary'][f'E{row_number_summary}'] = '=SUMIFS(Filings!$M:$M,Filings!$A:$A,@INDIRECT("A"&ROW()),Filings!$Q:$Q,1)'
            caiso_cross_check['FilingSummary'][f'F{row_number_summary}'] = '=SUMIFS(CAISO_Flex_SP!$G:$G,CAISO_Flex_SP!$B:$B,@INDIRECT("A"&ROW()),CAISO_Flex_SP!$F:$F,1)'
            caiso_cross_check['FilingSummary'][f'G{row_number_summary}'] = '=@INDIRECT("F"&ROW())-@INDIRECT("E"&ROW())'
            caiso_cross_check['FilingSummary'][f'G{row_number_summary}'].fill = PatternFill(start_color='FFF2CC',end_color='FFF2CC',fill_type='solid')
            caiso_cross_check['FilingSummary'][f'H{row_number_summary}'] = '=SUMIFS(Filings!$M:$M,Filings!$A:$A,@INDIRECT("A"&ROW()),Filings!$Q:$Q,2)'
            caiso_cross_check['FilingSummary'][f'I{row_number_summary}'] = '=SUMIFS(CAISO_Flex_SP!$G:$G,CAISO_Flex_SP!$B:$B,@INDIRECT("A"&ROWC()),CAISO_Flex_SP!$F:$F,2)'
            caiso_cross_check['FilingSummary'][f'J{row_number_summary}'] = '=@INDIRECT("I"&ROW())-@INDIRECT("H"&ROW())'
            caiso_cross_check['FilingSummary'][f'J{row_number_summary}'].fill = PatternFill(start_color='FFF2CC',end_color='FFF2CC',fill_type='solid')
            caiso_cross_check['FilingSummary'][f'K{row_number_summary}'] = '=SUMIFS(Filings!$M:$M,Filings!$A:$A,@INDIRECT("A"&ROW()),Filings!$Q:$Q,3)'
            caiso_cross_check['FilingSummary'][f'L{row_number_summary}'] = '=SUMIFS(CAISO_Flex_SP!$G:$G,CAISO_Flex_SP!$B:$B,@INDIRECT("A"&ROW()),CAISO_Flex_SP!$F:$F,3)'
            caiso_cross_check['FilingSummary'][f'M{row_number_summary}'] = '=@INDIRECT("L"&ROW())-@INDIRECT("K"&ROW())'
            caiso_cross_check['FilingSummary'][f'M{row_number_summary}'].fill = PatternFill(start_color='FFF2CC',end_color='FFF2CC',fill_type='solid')
            caiso_cross_check['FilingSummary'][f'Y{row_number_summary}'] = '=@INDIRECT("L"&ROW())-@INDIRECT("K"&ROW())'

            # update log with compliance:
            requirements = caiso_cross_check['Requirements'][f'B{row_number_summary+1}'].value
            resources = caiso_cross_check['Requirements'][f'C{row_number_summary+1}'].value + \
                caiso_cross_check['Requirements'][f'D{row_number_summary+1}'].value
            if resources>=requirements:
                compliance = 'Compliant'
            else:
                compliance = 'Noncompliant'
            self.consolidation_logger.data.loc[
                (self.consolidation_logger.data.loc[:,'ra_category']=='ra_monthly_filing') & \
                (self.consolidation_logger.data.loc[:,'organization_id']==organization_id),
                'compliance'
            ] = compliance
            self.consolidation_logger.commit()
            row_number_summary += 1
            for _,row in physical_resources.loc[(physical_resources.loc[:,'organization_id']==organization_id),:].iterrows():
                ra_summary['PhysicalResources']['A{}'.format(row_number_physical_resources)] = organization_id
                ra_summary['PhysicalResources']['B{}'.format(row_number_physical_resources)] = row.loc['contract_id']
                ra_summary['PhysicalResources']['C{}'.format(row_number_physical_resources)] = row.loc['resource_id']
                ra_summary['PhysicalResources']['D{}'.format(row_number_physical_resources)].value = row.loc['resource_adequacy_system']
                ra_summary['PhysicalResources']['E{}'.format(row_number_physical_resources)].value = row.loc['resource_adequacy_local']
                ra_summary['PhysicalResources']['F{}'.format(row_number_physical_resources)].value = row.loc['resource_mcc_bucket']
                ra_summary['PhysicalResources']['G{}'.format(row_number_physical_resources)].value = row.loc['continuous_availability']
                ra_summary['PhysicalResources']['H{}'.format(row_number_physical_resources)].value = row.loc['resource_adequacy_committed_flexible']
                ra_summary['PhysicalResources']['I{}'.format(row_number_physical_resources)].value = row.loc['resource_adequacy_flexibility_category']
                ra_summary['PhysicalResources']['J{}'.format(row_number_physical_resources)] = '=VLOOKUP(@INDIRECT("C"&ROW()),\'NQC_List\'!$B:$D,2,FALSE)'
                ra_summary['PhysicalResources']['K{}'.format(row_number_physical_resources)] = '=VLOOKUP(@INDIRECT("C"&ROW()),\'NQC_List\'!$B:$D,3,FALSE)'
                caiso_cross_check['Filings']['A{}'.format(row_number_physical_resources+1)] = organization_id
                caiso_cross_check['Filings']['B{}'.format(row_number_physical_resources+1)] = row.loc['contract_id']
                caiso_cross_check['Filings']['C{}'.format(row_number_physical_resources+1)] = row.loc['resource_id']
                caiso_cross_check['Filings']['D{}'.format(row_number_physical_resources+1)] = '=CONCATENATE(@INDIRECT("A"&ROW()),@INDIRECT("C"&ROW()),@INDIRECT("G"&ROW()))'
                caiso_cross_check['Filings']['E{}'.format(row_number_physical_resources+1)] = '=CONCATENATE(@INDIRECT("A"&ROW()),@INDIRECT("C"&ROW()))'
                caiso_cross_check['Filings']['F{}'.format(row_number_physical_resources+1)] = '=CONCATENATE(@INDIRECT("A"&ROW()),@INDIRECT("C"&ROW()),@INDIRECT("M"&ROW()))'
                caiso_cross_check['Filings']['G{}'.format(row_number_physical_resources+1)].value = row.loc['resource_adequacy_system']
                caiso_cross_check['Filings']['H{}'.format(row_number_physical_resources+1)].value = row.loc['resource_adequacy_local']
                caiso_cross_check['Filings']['I{}'.format(row_number_physical_resources+1)] = '=SUMIFS($G:$G,$A:$A,@INDIRECT("A"&ROW()),$C:$C,@INDIRECT("C"&ROW()))'
                caiso_cross_check['Filings']['J{}'.format(row_number_physical_resources+1)] = '=SUMIFS(CAISO_Sys_SP!$I:$I,CAISO_Sys_SP!$C:$C,@INDIRECT("A"&ROW()),CAISO_Sys_SP!$F:$F,@INDIRECT("C"&ROW()))'
                caiso_cross_check['Filings']['K{}'.format(row_number_physical_resources+1)] = '=IF(@INDIRECT("J"&ROW())<>@INDIRECT("I"&ROW()),"Y: "&@INDIRECT("J"&ROW())-@INDIRECT("I"&ROW())&" MW","-")'
                caiso_cross_check['Filings']['L{}'.format(row_number_physical_resources+1)].value = row.loc['resource_mcc_bucket']
                caiso_cross_check['Filings']['M{}'.format(row_number_physical_resources+1)].value = row.loc['resource_adequacy_committed_flexible']
                caiso_cross_check['Filings']['N{}'.format(row_number_physical_resources+1)] = '=SUMIFS($M:$M,$A:$A,@INDIRECT("A"&ROW()),$C:$C,@INDIRECT("C"&ROW()))'
                caiso_cross_check['Filings']['O{}'.format(row_number_physical_resources+1)] = '=SUMIFS(CAISO_Flex_SP!$G:$G,CAISO_Flex_SP!$B:$B,@INDIRECT("A"&ROW()),CAISO_Flex_SP!$E:$E,@INDIRECT("C"&ROW()))'
                caiso_cross_check['Filings']['P{}'.format(row_number_physical_resources+1)] = '=IF(@INDIRECT("O"&ROW())<>@INDIRECT("N"&ROW()),"Y: "&@INDIRECT("O"&ROW())-@INDIRECT("N"&ROW())&" MW","-")'
                caiso_cross_check['Filings']['Q{}'.format(row_number_physical_resources+1)].value = row.loc['resource_adequacy_flexibility_category']
                caiso_cross_check['Filings']['R{}'.format(row_number_physical_resources+1)].value = get_zone(row.loc['resource_id'])
                caiso_cross_check['Filings']['S{}'.format(row_number_physical_resources+1)] = '=IF(NOT(ISBLANK(@INDIRECT("G"&ROW()))),IF(ISNA(VLOOKUP(@INDIRECT("A"&ROW()),CAISO_Sys_SP!$A:$A,1,FALSE)),"N","-"),"")'
                caiso_cross_check['Filings']['T{}'.format(row_number_physical_resources+1)] = '=IF(AND(NOT(ISBLANK(@INDIRECT("H"&ROW()))),NOT(@INDIRECT("H"&ROW())=0)),IF(ISNA(VLOOKUP(@INDIRECT("B"&ROW()),CAISO_Sys_SP!$B:$B,1,FALSE)),"N","-"),"")'
                caiso_cross_check['Filings']['U{}'.format(row_number_physical_resources+1)] = '=IF(NOT(ISBLANK(@INDIRECT("M"&ROW()))),IF(ISNA(VLOOKUP(@INDIRECT("C"&ROW()),CAISO_Flex_SP!$A:$A,1,FALSE)),"N","-"),"")'
                for col in 'DEFIJKNOPSTU':
                    caiso_cross_check['Filings']['{}{}'.format(col,row_number_physical_resources+1)].fill = PatternFill(start_color='DDEBF7',end_color='DDEBF7',fill_type='solid')
                row_number_physical_resources += 1
            for _,row in demand_response.loc[(demand_response.loc[:,'organization_id']==organization_id),:].iterrows():
                ra_summary['PhysicalResources']['A{}'.format(row_number_physical_resources)] = organization_id
                ra_summary['PhysicalResources']['B{}'.format(row_number_physical_resources)] = row.loc['contract_id']
                ra_summary['PhysicalResources']['C{}'.format(row_number_physical_resources)] = row.loc['program_id']
                ra_summary['PhysicalResources']['D{}'.format(row_number_physical_resources)].value = row.loc['resource_adequacy_system']
                ra_summary['PhysicalResources']['E{}'.format(row_number_physical_resources)].value = row.loc['resource_adequacy_local']
                ra_summary['PhysicalResources']['F{}'.format(row_number_physical_resources)].value = row.loc['resource_mcc_bucket']
                ra_summary['PhysicalResources']['G{}'.format(row_number_physical_resources)].value = False
                ra_summary['PhysicalResources']['H{}'.format(row_number_physical_resources)].value = row.loc['resource_adequacy_committed_flexible']
                ra_summary['PhysicalResources']['I{}'.format(row_number_physical_resources)].value = row.loc['resource_adequacy_flexibility_category']
                ra_summary['PhysicalResources']['J{}'.format(row_number_physical_resources)] = '=VLOOKUP(@INDIRECT("C"&ROW()),\'NQC_List\'!$B:$D,2,FALSE)'
                ra_summary['PhysicalResources']['K{}'.format(row_number_physical_resources)] = '=VLOOKUP(@INDIRECT("C"&ROW()),\'NQC_List\'!$B:$D,3,FALSE)'
                caiso_cross_check['Filings']['A{}'.format(row_number_physical_resources+1)] = organization_id
                caiso_cross_check['Filings']['B{}'.format(row_number_physical_resources+1)] = row.loc['contract_id']
                caiso_cross_check['Filings']['C{}'.format(row_number_physical_resources+1)] = row.loc['program_id']
                caiso_cross_check['Filings']['D{}'.format(row_number_physical_resources+1)] = '=CONCATENATE(@INDIRECT("A"&ROW()),@INDIRECT("C"&ROW()),@INDIRECT("G"&ROW()))'
                caiso_cross_check['Filings']['E{}'.format(row_number_physical_resources+1)] = '=CONCATENATE(@INDIRECT("A"&ROW()),@INDIRECT("C"&ROW()))'
                caiso_cross_check['Filings']['F{}'.format(row_number_physical_resources+1)] = '=CONCATENATE(@INDIRECT("A"&ROW()),@INDIRECT("C"&ROW()),@INDIRECT("M"&ROW()))'
                caiso_cross_check['Filings']['G{}'.format(row_number_physical_resources+1)].value = row.loc['resource_adequacy_system']
                caiso_cross_check['Filings']['H{}'.format(row_number_physical_resources+1)].value = row.loc['resource_adequacy_local']
                caiso_cross_check['Filings']['I{}'.format(row_number_physical_resources+1)] = '=SUMIFS($G:$G,$A:$A,@INDIRECT("A"&ROW()),$C:$C,@INDIRECT("C"&ROW()))'
                caiso_cross_check['Filings']['J{}'.format(row_number_physical_resources+1)] = '=SUMIFS(CAISO_Sys_SP!$I:$I,CAISO_Sys_SP!$C:$C,@INDIRECT("A"&ROW()),CAISO_Sys_SP!$F:$F,@INDIRECT("C"&ROW()))'
                caiso_cross_check['Filings']['K{}'.format(row_number_physical_resources+1)] = '=IF(@INDIRECT("J"&ROW())<>@INDIRECT("I"&ROW()),"Y: "&@INDIRECT("J"&ROW())-@INDIRECT("I"&ROW())&" MW","-")'
                caiso_cross_check['Filings']['L{}'.format(row_number_physical_resources+1)].value = row.loc['resource_mcc_bucket']
                caiso_cross_check['Filings']['M{}'.format(row_number_physical_resources+1)].value = row.loc['resource_adequacy_committed_flexible']
                caiso_cross_check['Filings']['N{}'.format(row_number_physical_resources+1)] = '=SUMIFS($M:$M,$A:$A,@INDIRECT("A"&ROW()),$C:$C,@INDIRECT("C"&ROW()))'
                caiso_cross_check['Filings']['O{}'.format(row_number_physical_resources+1)] = '=SUMIFS(CAISO_Flex_SP!$G:$G,CAISO_Flex_SP!$B:$B,@INDIRECT("A"&ROW()),CAISO_Flex_SP!$E:$E,@INDIRECT("C"&ROW()))'
                caiso_cross_check['Filings']['P{}'.format(row_number_physical_resources+1)] = '=IF(@INDIRECT("O"&ROW())<>@INDIRECT("N"&ROW()),"Y: "&@INDIRECT("O"&ROW())-@INDIRECT("N"&ROW())&" MW","-")'
                caiso_cross_check['Filings']['Q{}'.format(row_number_physical_resources+1)].value = row.loc['resource_adequacy_flexibility_category']
                caiso_cross_check['Filings']['R{}'.format(row_number_physical_resources+1)].value = get_zone(row.loc['program_id'])
                caiso_cross_check['Filings']['S{}'.format(row_number_physical_resources+1)] = '=IF(NOT(ISBLANK(@INDIRECT("G"&ROW()))),IF(ISNA(VLOOKUP(@INDIRECT("A"&ROW()),CAISO_Sys_SP!$A:$A,1,FALSE)),"N","-"),"")'
                caiso_cross_check['Filings']['T{}'.format(row_number_physical_resources+1)] = '=IF(AND(NOT(ISBLANK(@INDIRECT("H"&ROW()))),NOT(@INDIRECT("H"&ROW())=0)),IF(ISNA(VLOOKUP(@INDIRECT("B"&ROW()),CAISO_Sys_SP!$B:$B,1,FALSE)),"N","-"),"")'
                caiso_cross_check['Filings']['U{}'.format(row_number_physical_resources+1)] = '=IF(NOT(ISBLANK(@INDIRECT("M"&ROW()))),IF(ISNA(VLOOKUP(@INDIRECT("C"&ROW()),CAISO_Flex_SP!$A:$A,1,FALSE)),"N","-"),"")'
                for col in 'DEFIJKNOPSTU':
                    caiso_cross_check['Filings']['{}{}'.format(col,row_number_physical_resources+1)].fill = PatternFill(start_color='DDEBF7',end_color='DDEBF7',fill_type='solid')
                row_number_physical_resources += 1

        # apply conditional formatting to flex-rar sheet:
        ra_summary['FlexRAR'].conditional_formatting.add(
            'C{}:C{}'.format(first_row_number_summary,row_number_summary-1),
            CellIsRule(
                operator='lessThan',
                formula=['@INDIRECT("B"&ROW())'],
                stopIfTrue=False,
                fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                font=Font(color='9C0006')
            )
        )
        ra_summary['FlexRAR'].conditional_formatting.add(
            'D{}:D{}'.format(first_row_number_summary,row_number_summary-1),
            CellIsRule(
                operator='lessThan',
                formula=[1.0],
                stopIfTrue=False,
                fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                font=Font(color='9C0006')
            )
        )
        ra_summary['FlexRAR'].conditional_formatting.add(
            'F{}:F{}'.format(first_row_number_summary,row_number_summary-1),
            CellIsRule(
                operator='lessThan',
                formula=['@INDIRECT("E"&ROW())'],
                stopIfTrue=False,
                fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                font=Font(color='9C0006')
            )
        )

        # create totals row in MCC_Check worksheet:
        ra_summary['MCC_Check'][f'A{row_number_summary}'] = 'Total:'
        for col in 'BCDEFGHIJKLMNOP':
            ra_summary['MCC_Check']['{}{}'.format(col,row_number_summary)] = '=SUM({0}{1}:{0}{2})'.format(col,first_row_number_summary,row_number_summary-1)

        # create totals row in FlexRAR worksheet:
        ra_summary['FlexRAR'][f'A{row_number_summary}'] = 'Total:'
        ra_summary['FlexRAR'][f'D{row_number_summary}'] = '=IFERROR(INDIRECT("C"&ROW())/@INDIRECT("B"&ROW()),0)'
        for col in 'BCEFGHIJLMNOQRST':
            ra_summary['FlexRAR']['{}{}'.format(col,row_number_summary)] = '=SUM({0}6:{0}{1})'.format(col,row_number_summary-1)
        for col in 'ABCDEFGHIJKLMNOPQRST':
            ra_summary['FlexRAR']['{}{}'.format(col,row_number_summary)].fill = PatternFill(start_color='00FF00',end_color='00FF00',fill_type='solid')

        # apply conditional formatting to local true-up sheet:
        for col in ['F','K','P','U','Z','AE','AJ','AO','AT','AY','BD','BI','BN','BS']:
            ra_summary['LocalTrueUp'].conditional_formatting.add(
                '{0}{1}:{0}{2}'.format(col,first_row_number_summary,row_number_summary-1),
                CellIsRule(
                    operator='lessThan',
                    formula=[0],
                    stopIfTrue=False,
                    fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                    font=Font(color='9C0006')
                )
            )

        # apply conditional formatting to certifying officer sheet:
        ra_summary['CertifyingOfficers'].conditional_formatting.add(
            'E{}:E{}'.format(first_row_number_summary,row_number_summary-1),
            CellIsRule(
                operator='equal',
                formula=['"Yes"'],
                stopIfTrue=False,
                fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                font=Font(color='9C0006')
            )
        )
        ra_summary['CertifyingOfficers'].conditional_formatting.add(
            'G{}:G{}'.format(first_row_number_summary,row_number_summary-1),
            CellIsRule(
                operator='equal',
                formula=['"Yes"'],
                stopIfTrue=False,
                fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                font=Font(color='9C0006')
            )
        )

        # write total lines in np and sp worksheets:
        for sheet_name in ['NP26','SP26']:
            ra_summary[sheet_name][f'A{row_number_summary}'] = 'Total:'
            ra_summary[sheet_name][f'A{row_number_summary}'].fill = PatternFill(start_color='00FF00',end_color='00FF00',fill_type='solid')
            for col in 'BCDEFGHIJK':
                ra_summary[sheet_name][f'{col}{row_number_summary}'] = '=SUM({0}{1}:{0}{2})'.format(col,first_row_number_summary,row_number_summary-1)
                ra_summary[sheet_name][f'{col}{row_number_summary}'].fill = PatternFill(start_color='00FF00',end_color='00FF00',fill_type='solid')
            ra_summary[sheet_name][f'I{row_number_summary}'] = '=@INDIRECT("C"&ROW())+@INDIRECT("E"&ROW())'
            ra_summary[sheet_name][f'A{row_number_summary+1}'].value = 'Net Long/Short:'
            ra_summary[sheet_name][f'C{row_number_summary+1}'].value = 'Excess in Zone:'
            ra_summary[sheet_name][f'A{row_number_summary+1}'].fill = PatternFill(start_color='CCFFCC',end_color='CCFFCC',fill_type='solid')
            ra_summary[sheet_name][f'E{row_number_summary+1}'] = '=@INDIRECT("D"&ROW()-1)+@INDIRECT("C"&ROW()-1)-@INDIRECT("B"&ROW()-1)'

            # apply conditional formatting to columns:
            ra_summary[sheet_name].conditional_formatting.add(
                f'F{first_row_number_summary}:F{row_number_summary-1}',
                CellIsRule(
                    operator='greaterThan',
                    formula=[0],
                    stopIfTrue=False,
                    fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                    font=Font(color='9C0006')
                )
            )
            ra_summary[sheet_name].conditional_formatting.add(
                '{0}{1}:{0}{2}'.format({'NP26':'M','SP26':'L'}[sheet_name],first_row_number_summary,row_number_summary-1),
                CellIsRule(
                    operator='equal',
                    formula=['"Monthly Filing Not Found"'],
                    stopIfTrue=False,
                    fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                    font=Font(color='9C0006')
                )
            )

        # apply conditional formatting to requirements sheet in caiso supply plan cross-check file:
        caiso_cross_check['Requirements'].conditional_formatting.add(
            f'E{first_row_number_summary+1}:E{row_number_summary}',
            CellIsRule(
                operator='lessThan',
                formula=['@INDIRECT("B"&ROW())'],
                stopIfTrue=False,
                fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                font=Font(color='9C0006')
            )
        )
        caiso_cross_check['Requirements'].conditional_formatting.add(
            f'F{first_row_number_summary+1}:F{row_number_summary}',
            CellIsRule(
                operator='lessThan',
                formula=[1.0],
                stopIfTrue=False,
                fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                font=Font(color='9C0006')
            )
        )
        caiso_cross_check['Requirements'].conditional_formatting.add(
            f'K{first_row_number_summary+1}:K{row_number_summary}',
            CellIsRule(
                operator='lessThan',
                formula=['@INDIRECT("J"&ROW())'],
                stopIfTrue=False,
                fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                font=Font(color='9C0006')
            )
        )
        caiso_cross_check['Requirements'].conditional_formatting.add(
            f'L{first_row_number_summary+1}:L{row_number_summary}',
            CellIsRule(
                operator='lessThan',
                formula=[1.0],
                stopIfTrue=False,
                fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                font=Font(color='9C0006')
            )
        )
        caiso_cross_check['Requirements'].conditional_formatting.add(
            f'N{first_row_number_summary+1}:N{row_number_summary}',
            CellIsRule(
                operator='lessThan',
                formula=['@INDIRECT("M"&ROW())'],
                stopIfTrue=False,
                fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                font=Font(color='9C0006')
            )
        )
        for col in ['AB','AG','AL','AQ','AV','BA','BF','BK','BP','BU','BZ','CE','CJ','CO']:
            caiso_cross_check['Requirements'].conditional_formatting.add(
                '{0}{1}:{0}{2}'.format(col,first_row_number_summary+1,row_number_summary),
                CellIsRule(
                    operator='lessThan',
                    formula=[0],
                    stopIfTrue=False,
                    fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
                    font=Font(color='9C0006')
                )
            )

        # save outputs and add to logs:
        save_timestamp = ts.now()
        email_id = '00000000-0000-0000-0000-{}0002'.format(save_timestamp.strftime('%Y%m%d'))
        if email_id not in self.email_logger.data.loc[:,'email_id'].values:
            email_information = pd.Series({
                'email_id' : email_id,
                'sender' : 'results',
                'subject' : '',
                'receipt_date' : save_timestamp,
                'included' : True,
                'group' : 'internal',
            })
            self.email_logger.log(email_information)
        attachment_index = len(self.attachment_logger.data.loc[(self.attachment_logger.data.loc[:,'email_id']==email_id),'attachment_id'])
        attachment_id = '{}0002{:020.0f}'.format(save_timestamp.strftime('%Y%m%d'),attachment_index)
        archive_path = self.config.paths.get_path('ra_summary')
        if attachment_id not in self.attachment_logger.data.loc[:,'attachment_id'].values and str(archive_path) not in self.attachment_logger.data.loc[:,'archive_path'].values:
            attachment_information = pd.Series({
                'email_id' : email_id,
                'attachment_id' : attachment_id,
                'download_path' : '',
                'ra_category' : 'ra_summary',
                'organization_id' : 'CPUC',
                'effective_date' : filing_month,
                'archive_path' : str(archive_path),
            })
            self.attachment_logger.log(attachment_information)
        ra_summary.save(str(self.config.paths.get_path('ra_summary')))
        ra_summary.close()

        attachment_index = len(self.attachment_logger.data.loc[(self.attachment_logger.data.loc[:,'email_id']==email_id),'attachment_id'])
        attachment_id = '{}0002{:020.0f}'.format(save_timestamp.strftime('%Y%m%d'),attachment_index)
        archive_path = self.config.paths.get_path('caiso_cross_check')
        if attachment_id not in self.attachment_logger.data.loc[:,'attachment_id'].values and str(archive_path) not in self.attachment_logger.data.loc[:,'archive_path'].values:
            attachment_information = pd.Series({
                'email_id' : email_id,
                'attachment_id' : attachment_id,
                'download_path' : '',
                'ra_category' : 'caiso_cross_check',
                'organization_id' : 'CPUC',
                'effective_date' : filing_month,
                'archive_path' : str(archive_path),
            })
            self.attachment_logger.log(attachment_information)
        self.attachment_logger.commit()
        self.email_logger.commit()
        caiso_cross_check.save(str(archive_path))
        caiso_cross_check.close()

        # check time and report:
        run_time = (ts.now() - init_time).total_seconds()
        self.logger.log('Consolidated LSE Filings in {:02.0f}:{:02.0f}:{:05.2f}'.format(int(run_time/3600),int((run_time%3600)/60),run_time%60),'INFORMATION')

    # consolidate caiso supply plans into cross-check file:
    def consolidate_supply_plans(self):
        '''
        loads data from caiso supply plan workbooks and populates the caiso
        cross-check workbooks with data and calculation results.
        '''
        # start timer:
        init_time = ts.now()
        self.logger.log('Copying CAISO Supply Plan into Cross-Check File','INFORMATION')

        supply_plan_system = read_supply_plan(self.config,'supply_plan_system')
        supply_plan_flexible = read_supply_plan(self.config,'supply_plan_flexible')

        path = self.config.paths.get_path('caiso_cross_check')
        caiso_cross_check = open_workbook(path,data_only=False,read_only=False)

        # copy supply plan data into cross-check file:
        row_number = 2
        for _,row in supply_plan_system.iterrows():
            caiso_cross_check['CAISO_Sys_SP'][f'A{row_number}'] = '=CONCATENATE(INDIRECT("C"&ROW()),INDIRECT("F"&ROW()),INDIRECT("I"&ROW()))'
            caiso_cross_check['CAISO_Sys_SP'][f'B{row_number}'] = '=CONCATENATE(INDIRECT("C"&ROW()),INDIRECT("F"&ROW()))'
            caiso_cross_check['CAISO_Sys_SP'][f'C{row_number}'] = '=VLOOKUP(INDIRECT("N"&ROW()),LoadServingEntities!B:C,2,FALSE)'
            caiso_cross_check['CAISO_Sys_SP'][f'J{row_number}'] = '=SUMIFS(I:I,C:C,INDIRECT("C"&ROW()),F:F,INDIRECT("F"&ROW()))'
            caiso_cross_check['CAISO_Sys_SP'][f'K{row_number}'] = '=SUMIFS(G:G,C:C,INDIRECT("C"&ROW()),F:F,INDIRECT("F"&ROW()))'
            caiso_cross_check['CAISO_Sys_SP'][f'P{row_number}'] = '=IF(ISNA(VLOOKUP(INDIRECT("A"&ROW()),Filings!D:D,1,FALSE)),"N","-")'
            for column_letter in 'ABCJKP':
                caiso_cross_check['CAISO_Sys_SP'][f'{column_letter}{row_number}'].fill = PatternFill(start_color='DDEBF7',end_color='DDEBF7',fill_type='solid')
            for column_index,column in enumerate(supply_plan_system.columns):
                caiso_cross_check['CAISO_Sys_SP']['{}{}'.format('DEFGHILMNO'[column_index],row_number)].value = row.loc[column]
            row_number += 1

        # copy supply plan data into cross-check file:
        row_number = 2
        for _,row in supply_plan_flexible.iterrows():
            caiso_cross_check['CAISO_Flex_SP'][f'A{row_number}'] = '=CONCATENATE(INDIRECT("B"&ROW()),INDIRECT("E"&ROW()),INDIRECT("G"&ROW()))'
            caiso_cross_check['CAISO_Flex_SP'][f'B{row_number}'] = '=VLOOKUP(INDIRECT("K"&ROW()),LoadServingEntities!B:C,2,FALSE)'
            caiso_cross_check['CAISO_Flex_SP'][f'H{row_number}'] = '=SUMIFS(G:G,B:B,INDIRECT("B"&ROW()),E:E,INDIRECT("E"&ROW()),F:F,INDIRECT("F"&ROW()))'
            caiso_cross_check['CAISO_Flex_SP'][f'M{row_number}'] = '=IF(ISNA(VLOOKUP(INDIRECT("A"&ROW()),Filings!F:F,1,FALSE)),"N","-")'
            for column_letter in 'ABHM':
                caiso_cross_check['CAISO_Flex_SP'][f'{column_letter}{row_number}'].fill = PatternFill(start_color='DDEBF7',end_color='DDEBF7',fill_type='solid')
            for column_index,column in enumerate(supply_plan_flexible.columns):
                caiso_cross_check['CAISO_Flex_SP']['{}{}'.format('CDEFGIJKL'[column_index],row_number)].value = row.loc[column]
            row_number += 1

        # save and close:
        caiso_cross_check.save(str(self.config.paths.get_path('caiso_cross_check')))
        caiso_cross_check.close()

        # check time and report:
        run_time = (ts.now() - init_time).total_seconds()
        self.logger.log('Copied CAISO Supply Plan to Cross-Check File in {:02.0f}:{:02.0f}:{:05.2f}'.format(int(run_time/3600),int((run_time%3600)/60),run_time%60),'INFORMATION')

    def check_files(self):
        '''
        Checks whether all files required for consolidation are available and provides a table of results

        returns:
            boolean - true if all files are available, false otherwise
        '''
        filing_month = self.config.filing_month
        self.attachment_logger.load_log()
        def set_file_status(ra_category,organization_id):
            attachments = self.attachment_logger.data
            if ra_category in ('ra_monthly_filing','supply_plan_system','supply_plan_flexible','cam_rmr'):
                effective_date = pd.to_datetime(filing_month)
            elif ra_category=='incremental_local':
                effective_date = pd.to_datetime(filing_month).replace(month=7)
            else:
                effective_date = pd.to_datetime(filing_month).replace(month=1)
            file_information = pd.Series({
                'filing_month' : filing_month,
                'ra_category' : ra_category,
                'effective_date' : effective_date,
                'organization_id' : organization_id,
                'attachment_id' : '',
                'archive_path' : '',
                'status' : '',
                'compliance' : '' if ra_category=='ra_monthly_filing' else 'n/a',
            })
            versions = attachments.loc[
                (attachments.loc[:,'ra_category']==ra_category) & \
                (attachments.loc[:,'effective_date']==effective_date) & \
                (attachments.loc[:,'organization_id']==organization_id), \
                ['email_id','attachment_id','archive_path']
            ].merge(self.email_logger.data.loc[:,['email_id','receipt_date']],on='email_id')
            versions.sort_values('receipt_date',ascending=True,inplace=True)
            if len(versions)>0 and Path(versions.iloc[0].loc['archive_path']).is_file():
                receipt_date = versions.iloc[0].loc['receipt_date']
                receipt_date = receipt_date.tz_localize(tz='UTC').tz_convert('US/Pacific').tz_localize(None)
                versions.sort_values('receipt_date',ascending=False,inplace=True)
                if ra_category=='ra_monthly_filing' and receipt_date>self.config.get_filing_due_date(filing_month):
                    file_information.loc['attachment_id'] = versions.iloc[0].loc['attachment_id']
                    file_information.loc['archive_path'] = versions.iloc[0].loc['archive_path']
                    file_information.loc['status'] = 'Late'
                else:
                    file_information.loc['attachment_id'] = versions.iloc[0].loc['attachment_id']
                    file_information.loc['archive_path'] = versions.iloc[0].loc['archive_path']
                    file_information.loc['status'] = 'Ready'
            elif ra_category=='incremental_local' and filing_month.month<=6:
                file_information.loc['attachment_id'] = ''
                file_information.loc['archive_path'] = ''
                file_information.loc['status'] = 'Not Required'
            elif ra_category=='cam_rmr' and filing_month.year>=2024:
                file_information.loc['attachment_id'] = ''
                file_information.loc['archive_path'] = ''
                file_information.loc['status'] = 'Not Required'
            elif ra_category=='ra_monthly_filing':
                if len(versions)>0:
                    file_information.loc['attachment_id'] = versions.iloc[0].loc['attachment_id']
                    file_information.loc['archive_path'] = ''
                    file_information.loc['status'] = 'Invalid File'
                else:
                    file_information.loc['attachment_id'] = ''
                    file_information.loc['archive_path'] = ''
                    file_information.loc['status'] = 'File Not Submitted'
            elif self.config.paths.get_path(ra_category) is not None:
                if self.config.paths.get_path(ra_category).is_file():
                    file_information.loc['attachment_id'] = versions.iloc[0].loc['attachment_id']
                    file_information.loc['archive_path'] = str(self.config.paths.get_path(ra_category))
                    file_information.loc['status'] = 'Ready'
                else:
                    file_information.loc['attachment_id'] = ''
                    file_information.loc['archive_path'] = ''
                    file_information.loc['status'] = 'File Not Found'
            else:
                file_information.loc['attachment_id'] = ''
                file_information.loc['archive_path'] = ''
                file_information.loc['status'] = 'File Not Found'
            # overwrite previous entries for a given file:
            check_columns = ['ra_category','effective_date','organization_id']
            previous_log_indices = (self.consolidation_logger.data.loc[:,check_columns]==file_information.loc[check_columns]).apply(all,axis='columns',result_type='reduce')
            if previous_log_indices.sum()>0:
                self.consolidation_logger.data.drop(index=self.consolidation_logger.data.loc[previous_log_indices,:].index,inplace=True)
            else:
                pass
            self.consolidation_logger.log(file_information)
        # get list of current lses from summary template file:
        path = self.config.paths.get_path('ra_summary_template')
        ra_summary = open_workbook(path,data_only=True,read_only=True)
        data_range = get_data_range(ra_summary['Summary'],'A','',self.config)
        active_lses = [row[0].value for row in data_range]
        ra_categories = ['year_ahead','incremental_local','month_ahead','cam_rmr','supply_plan_system','supply_plan_flexible'] + ['ra_monthly_filing'] * len(active_lses)
        organization_ids = ['CEC','CEC','CPUC','CPUC','CAISO','CAISO'] + active_lses
        # get list of active load serving entities from summary sheet:
        for ra_category,organization_id in zip(ra_categories,organization_ids):
            set_file_status(ra_category,organization_id)
        ready =  all([s in ('Ready','Late','Not Required') for s in self.consolidation_logger.data.loc[(self.consolidation_logger.data.loc[:,'ra_category']!='ra_monthly_filing'),'status']])
        self.consolidation_logger.commit()
        return ready