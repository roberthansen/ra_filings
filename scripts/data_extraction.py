import io
import re
import xlrd
import warnings
import pandas as pd
from pathlib import Path
from itertools import chain
from functools import reduce
from datetime import time
from pandas import Timestamp as ts, Timedelta as td
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ra_logging import TextLogger
from configuration_options import ConfigurationOptions

def open_workbook(path:Path,data_only:bool=True,read_only:bool=True,in_mem:bool=True):
    '''
    loads an excel workbook from a given path according to a set of parameters

    parameters:
        path - a Path object which points to an excel xlsx file
        data_only - openpyxl parameter to open workbook in data-only mode
            (i.e., evaluated values instead of formulas)
        read_only - openpyxl parameter to open workbook in read-only mode,
            which can improve load times
        in_mem - if true, loads workbook into python-managed memory rather of
            reading from disc, which can improve random access read times
    '''
    if path is not None:
        if path.is_file() and re.match(r'^.*\.xlsx$',path.name):
            with warnings.catch_warnings() as w:
                warnings.filterwarnings(action='ignore')
                if in_mem:
                    with path.open('rb') as f:
                        in_mem_file = io.BytesIO(f.read())
                        workbook = load_workbook(in_mem_file,data_only=data_only,read_only=read_only)
                else:
                    workbook = load_workbook(str(path),data_only=data_only,read_only=read_only)
        else:
            workbook = None
    else:
        workbook = None
    return workbook

def get_data_range(worksheet:Worksheet,lse_column:str,data_columns:str,config:ConfigurationOptions):
    '''
    extracts a table from a worksheet with rows corresponding to each load-serving entity
    parameters:
        worksheet - an excel worksheet expected to contain the data table
        lse_column - the column letter identifying the column in which load-
            serving entities are expected to be listed
        data_columns - a list of letters identifying columns from which to
            retrieve data matched to the load-serving entities identified in
            the lse_column
        config - an instance of the ConfigurationOptions class used for
            retrieving containing information about each load-serving entity
    '''
    first_row = 0
    last_row = 0
    organization_ids = [organization['id'].lower() for organization in config.organizations.list_load_serving_entities()]
    for row_number in range(1,worksheet.max_row+1):
        if first_row==0 and str(worksheet[f'{lse_column}{row_number}'].value).lower() in organization_ids:
            first_row = row_number
        if first_row>0 and last_row==0 and str(worksheet[f'{lse_column}{row_number+1}'].value).lower() in ('none','total','total:','total cpuc juris'):
            last_row = row_number
            break
    if first_row>0 and last_row>0:
        # return a list of lists containing excel cell objects:
        data_range = []
        for row_number in range(first_row,last_row+1):
            data_range.append(
                [worksheet[f'{lse_column}{row_number}']] + \
                [worksheet[f'{data_column}{row_number}'] for data_column in data_columns]
            )
    else:
        # return a list containing an empty list;
        data_range = [[]]
    return data_range

def data_range_to_dataframe(columns,data_range):
    '''
    converts a excel data range into a pandas dataframe
    parameters:
        columns - list of column labels to apply to the dataframe
        data_range - excel data range containing cells to insert into new
            dataframe
    '''
    data_array = []
    for data_range_row in data_range:
        data_array_row = []
        for data_range_cell in data_range_row:
            data_array_row.append(data_range_cell.value)
        data_array.append(data_array_row)
    return pd.DataFrame(data_array,columns=columns)

# find and return data range for the flex requirements net cam table:
def get_table(worksheet,table_header_text:str,table_header_offset:dict,columns:list):
    '''
    searches a given worksheet for table header text and returns a dataframe
    based on a table with an upper left corner determined by a header offset, a
    height defined by consecutive non-empty cells in the first column, and a
    width defined by the number of column labels.
    parameters:
        worksheet - an excel worksheet expected to contain the table header
        table_header_text - a string which, if found identically within the
            worksheet, will be used as the reference coordinates for the table
        table_header_offset - a dictionary containing coordinates for top left
            corner of range to extract relative to the cell containing
            table_header_text in the form {'rows':___,'columns':___}
        columns - a list which will be used as the column labels in the
            resulting dataframe
    '''
    table_bounds = {
        'top' : worksheet.max_row,
        'bottom' : worksheet.max_row,
        'left' : get_column_letter(worksheet.max_column),
        'right' : get_column_letter(worksheet.max_column),
    }
    for row_index,row in enumerate(worksheet.rows):
        if row_index+1<table_bounds['bottom']:
            for column_index,cell in enumerate(row):
                if table_header_text.lower() in str(cell.value).lower():
                    # found header, setting upper, left, and right boundaries of table (row and column indices are excel label-1):
                    table_bounds['top'] = row_index + 1 + table_header_offset['rows']
                    table_bounds['left'] = get_column_letter(column_index + table_header_offset['columns'] + 1)
                    table_bounds['right'] = get_column_letter(column_index + table_header_offset['columns'] + len(columns))
                if row_index+1>table_bounds['top'] and get_column_letter(column_index+1)==table_bounds['left'] and str(cell.value).replace(' ','') in ('None',''):
                    # found last row in table:
                    table_bounds['bottom'] = row_index
                    break
        else:
            break
    data_range = worksheet['{left}{top}:{right}{bottom}'.format(**table_bounds)]
    return data_range_to_dataframe(columns,data_range)

def read_ra_monthly_filing(organization:dict,config:ConfigurationOptions,logger:TextLogger,date:ts=None,version:int=None):
    '''
    extracts relevant data tables from a monthly filing workbook for a given
    lse and month

    parameters:
        organization - the abbreviated identifier string for an active load-
            serving entity, i.e., a key in the organizations dictionary
        config - an instance of the ConfigurationOptions class
        logger - text_logger object for logging errors and other information
        date - an optional datetime object representing the filing month; if
            left blank, the filing month defined in the paths object will be applied
        version - an optional integer version number; if left blank, the
            highest numbered version of a filing for the given load-serving
            entity and month will be used
    '''
    path = config.paths.get_path('ra_monthly_filing',organization=organization,date=date,version=version)
    try:
        ra_monthly_filing = open_workbook(path,data_only=True,read_only=True,in_mem=False)
    except UserWarning:
        pass
    physical_resources_columns=[
        'organization_id',
        'contract_id',
        'resource_id',
        'resource_adequacy_system',
        'resource_adequacy_local',
        'resource_mcc_bucket',
        'continuous_availability',
        'resource_adequacy_committed_flexible',
        'resource_adequacy_flexibility_category',
        'start_date',
        'end_date',
        'scid',
        'zone',
        'local_area',
    ]
    demand_response_columns = [
        'organization_id',
        'contract_id',
        'program_id',
        'resource_adequacy_system',
        'resource_adequacy_local',
        'resource_mcc_bucket',
        'third_party_program',
        'resource_adequacy_committed_flexible',
        'resource_adequacy_flexibility_category',
        'start_date',
        'end_date',
        'operator',
        'zone',
        'local_area',
    ]
    if ra_monthly_filing is not None:
        logger.log('Reading Filing Data for {name} ({id})'.format(**organization),'INFORMATION')
        sheet_names = [
            'Certification',
            'LSE Allocations',
            'ID and Local Area',
            'Summary Year Ahead',
            'Summary Month Ahead',
            'I_Phys_Res_Import_RA_Res',
            'II_Construc',
            'III_Demand_Response'
        ]
        sheet_check = all([sheet_name in ra_monthly_filing.sheetnames for sheet_name in sheet_names])
        if sheet_check:
            # check that each required column is present:
            column_checks = [
                r'.*contract\s+identifier.*',
                r'.*resource\s+id.*',
                r'.*system\s+ra.*',
                r'.*local\s+ra.*',
                r'.*mcc.*',
                r'.*available.*',
                r'.*flexible\s+ra.*',
                r'.*flexible\s+category.*',
                r'.*start\s+date.*',
                r'.*end\s+date.*',
                r'.*scid.*',
                r'.*zonal.*',
                r'.*area.*',
                r'north',
                r'south',
            ]
            cell_values = [cell.value for cell in ra_monthly_filing['I_Phys_Res_Import_RA_Res']['B3:P3'][0]]
            phys_res_check = all([re.match(column_check,str(cell_values[i]).lower()) for i,column_check in enumerate(column_checks)])
            column_checks = [
                r'.*identifier.*',
                r'.*program.*',
                r'.*system.*',
                r'.*local.*',
                r'.*mcc.*',
                r'.*third\s+party.*',
                r'.*flexible\s+ra.*',
                r'.*flexible\s+category.*',
                r'.*start\s+date.*',
                r'.*end\s+date.*',
                r'.*operator.*',
                r'.*zonal.*',
                r'.*area.*',
                r'.*do\s+not\s+delete.*',
                r'north',
                r'south',
                r'pg.*e',
                r'sce',
                r'sdge',
            ]
            worksheet = ra_monthly_filing['III_Demand_Response']
            cell_values = [cell.value for cell in chain(worksheet['B3:O3'][0],worksheet['O4:P4'][0],[worksheet['R3'],worksheet['T3'],worksheet['V3']])]
            demand_response_check = all([re.match(column_check,str(cell_values[i]).lower()) for i,column_check in enumerate(column_checks)])
        else:
            phys_res_check = False
            demand_response_check = False
        if all([sheet_check,phys_res_check,demand_response_check]):
            # retrieve values for summary table:
            pge_transmission_loss_coefficient = (config.get_option('transmission_loss_adder_pge')-1)/config.get_option('transmission_loss_adder_pge')
            sce_transmission_loss_coefficient = (config.get_option('transmission_loss_adder_sce')-1)/config.get_option('transmission_loss_adder_sce')
            sdge_transmission_loss_coefficient = (config.get_option('transmission_loss_adder_sdge')-1)/config.get_option('transmission_loss_adder_sdge')
            summary = pd.DataFrame({
                'organization_id' : [organization['id']],
                'organization_officer_name' : [ra_monthly_filing['Certification']['B21'].value],
                'organization_officer_title' : [ra_monthly_filing['Certification']['B22'].value],
                'np26dr' : [
                        ra_monthly_filing['III_Demand_Response']['O5'].value +
                        ra_monthly_filing['III_Demand_Response']['S13'].value * pge_transmission_loss_coefficient
                    ],
                'sp26dr' : [
                        ra_monthly_filing['III_Demand_Response']['P5'].value +
                        ra_monthly_filing['III_Demand_Response']['U8'].value * sce_transmission_loss_coefficient +
                        ra_monthly_filing['III_Demand_Response']['W6'].value * sdge_transmission_loss_coefficient
                    ],
            })

            # retrieve physical resources table:
            last_row = ra_monthly_filing['I_Phys_Res_Import_RA_Res'].max_row
            if last_row > 5:
                physical_resources = data_range_to_dataframe(physical_resources_columns,ra_monthly_filing[f'I_Phys_Res_Import_RA_Res'][f'A5:N{last_row}'])
                physical_resources.loc[:,'contract_id'] = physical_resources.loc[:,'contract_id'].map(lambda x: None if pd.isnull(x) or str(x).strip()=='' else x)
                physical_resources.loc[:,'resource_id'] = physical_resources.loc[:,'resource_id'].map(lambda x: None if pd.isnull(x) or str(x).strip()=='' else x)
                physical_resources.loc[:,'start_date'] = physical_resources.loc[:,'start_date'].map(parse_date)
                physical_resources.loc[:,'end_date'] = physical_resources.loc[:,'end_date'].map(parse_date)
                physical_resources.dropna(
                    axis='index',
                    how='all',
                    subset=[
                        'contract_id',
                        'resource_id'
                    ],
                    inplace=True
                )
            else:
                physical_resources = pd.DataFrame(columns=physical_resources_columns)
            physical_resources.loc[:,'organization_id'] = organization['id']
            physical_resources.replace('',0,inplace=True)
            physical_resources.loc[:,'continuous_availability'] = physical_resources.loc[:,'continuous_availability'].map(lambda s: True if s=='Y' else False)

            # retrieve demand response table:
            last_row = ra_monthly_filing['III_Demand_Response'].max_row
            if last_row > 17:
                demand_response = data_range_to_dataframe(demand_response_columns,ra_monthly_filing[f'III_Demand_Response'][f'A17:N{last_row}'])
                demand_response.loc[:,'contract_id'] = demand_response.loc[:,'contract_id'].map(lambda x: None if pd.isnull(x) or str(x).strip()=='' else x)
                demand_response.loc[:,'program_id'] = demand_response.loc[:,'program_id'].map(lambda x: None if pd.isnull(x) or str(x).strip()=='' else x)
                demand_response.loc[:,'start_date'] = demand_response.loc[:,'start_date'].map(parse_date)
                demand_response.loc[:,'end_date'] = demand_response.loc[:,'end_date'].map(parse_date)
                demand_response.dropna(
                    axis='index',
                    how='all',
                    subset=[
                        'contract_id',
                        'program_id'
                    ],
                    inplace=True
                )
            else:
                demand_response = pd.DataFrame(columns=demand_response_columns)
            demand_response.loc[:,'organization_id'] = organization['id']
            demand_response.replace('',0,inplace=True)
            demand_response.loc[:,'third_party_program'] = demand_response.loc[:,'third_party_program'].map(lambda s: True if s=='Y' else False)

        else:
            logger.log('Filing Did Not Pass Validation: {name} ({id})'.format(**organization),'WARNING')
            summary = pd.DataFrame({
                'organization_id' : [organization['id']],
                'organization_officer_name' : ['[Filing Did Not Pass Validation]'],
                'organization_officer_title' : ['[N/A]'],
                'np26dr' : [0],
                'sp26dr' : [0],
            })
            physical_resources = pd.DataFrame(columns=physical_resources_columns)
            demand_response = pd.DataFrame(columns=demand_response_columns)

        # close monthly filing workbook:
        ra_monthly_filing.close()

    # return void dataframe rows when monthly filing for lse is not found:
    else:
        logger.log('Unable to Find Filing Data for {name} ({id})'.format(**organization),'WARNING')
        summary = pd.DataFrame({
            'organization_id' : [organization['id']],
            'organization_officer_name' : ['[Monthly Filing Not Found]'],
            'organization_officer_title' : ['[N/A]'],
            'np26dr' : [0],
            'sp26dr' : [0],
        })
        physical_resources = pd.DataFrame(columns=physical_resources_columns)
        demand_response = pd.DataFrame(columns=demand_response_columns)
    return [summary,physical_resources,demand_response]

def get_year_ahead_tables(year_ahead,config:ConfigurationOptions):
    '''
    loads relevant data from the year-ahead workbook into dataframes
    parameters:
        year_ahead - year-ahead workbook
        config - an instance of the ConfigurationOptions class
    '''
    # get load forecast input table from the year ahead workbook:
    columns = [
        'iou_territory',
        'month',
        'organization_id',
        'lse_type',
        'submitted_forecast',
        'coincidence_adjustment',
        'coincident_peak_forecast',
        'lse_specific_total',
        'copkadj_with_lseadj',
        'eelmdr_adjustment',
        'adjusted_with_lmdr',
        'pro_rata_adjustment',
        'final_coincident_peak_forecast',
    ]
    data_range = year_ahead['loadforecastinputdata']['B2:N{}'.format(year_ahead['loadforecastinputdata'].max_row)]
    filing_month = config.filing_month
    load_forecast_input_data = data_range_to_dataframe(columns,data_range)
    load_forecast_input_data.dropna(axis='index',how='all',inplace=True)
    load_forecast_input_data.loc[:,'month'] = load_forecast_input_data.loc[:,'month'].map(lambda s: ts(filing_month.year,int(s),1)).astype('datetime64[M]')
    load_forecast_input_data.set_index(['iou_territory','organization_id','month'],inplace=True)
    load_forecast_input_data.sort_index(inplace=True)
    # accommodate errant entries:
    load_forecast_input_data.loc[:,columns[4:]] = load_forecast_input_data.loc[:,columns[4:]].applymap(lambda x: 0 if isinstance(x,str) else x)

    # demand response allocation table:
    month_columns = [ts(filing_month.year,month,1) for month in range(1,13)]
    columns = [
        'location',
    ] + month_columns
    data_range = year_ahead['DRforAllocation']['D2:P32']
    demand_response_allocation = data_range_to_dataframe(columns,data_range)
    demand_response_allocation.loc[:,'location'] = demand_response_allocation.loc[:,'location'].map(lambda x: str(x).lower().replace(' ','_').replace('&',''))
    demand_response_allocation.fillna(0,inplace=True)
    ignore_locations = ('','none','total_iou_service_area0','total_iou_service_area1','la_basin','big_creek/ventura','outside_lca','sce_total','other_pge_areas0','pge_dr','sdge','sdge_total','caiso_total')
    demand_response_allocation.drop(labels=demand_response_allocation.loc[(demand_response_allocation.loc[:,'location'].map(lambda s: s in ignore_locations)),:].index,inplace=True)
    demand_response_allocation.loc[:,'allocation_type'] = demand_response_allocation.loc[:,'location'].map(lambda s:'prorated' if s[-1]=='1' else 'base')
    demand_response_allocation.loc[:,'location'] = demand_response_allocation.loc[:,'location'].map(lambda s: s.replace('1','').replace('0',''))
    demand_response_allocation.loc[:,'location'] = demand_response_allocation.loc[:,'location'].replace('big_creek/ventura','ventura').replace('la_basin','los_angeles').replace('ncnb','northern_california').replace('sdge','san_diego')
    demand_response_allocation = demand_response_allocation.melt(id_vars=['location','allocation_type'],var_name='month',value_name='allocation')
    demand_response_allocation.loc[:,'month'] = demand_response_allocation.loc[:,'month'].astype('datetime64[M]')
    demand_response_allocation.set_index(['location','allocation_type','month'],inplace=True)
    demand_response_allocation.sort_index(inplace=True)

    # cam credits:
    columns = [
        'iou_territory',
        'category',
    ] + month_columns
    data_range = year_ahead['Flexrequirements']['R4:AE12']
    cam_credits = data_range_to_dataframe(columns,data_range)
    cam_credits = cam_credits.melt(id_vars=['iou_territory','category'],var_name='month',value_name='cam_credit')
    cam_credits.fillna(0,inplace=True)
    cam_credits.set_index(['iou_territory','category','month'],inplace=True)
    cam_credits.sort_index(inplace=True)

    # flexibility requirements:
    table_header_offset = {
        'rows' : 1,
        'columns' : -1,
    }
    table_header_text = 'Flex Requirements net CAM'
    month_columns = [ts(filing_month.year,month,1).to_numpy().astype('datetime64[M]') for month in range(1,13)]
    columns = [
        'organization_id',
        'flex_category',
    ] + month_columns
    flexibility_requirements = get_table(year_ahead['Flexrequirements'],table_header_text,table_header_offset,columns)
    flexibility_requirements['flex_category'] = flexibility_requirements.loc[:,'flex_category'].map(lambda s: int(s[-1]))
    flexibility_requirements = flexibility_requirements.melt(id_vars=['organization_id','flex_category'],var_name='month',value_name='flexibility_requirement')
    flexibility_requirements.loc[:,'month'] = flexibility_requirements.loc[:,'month'].astype('datetime64[M]')
    flexibility_requirements.set_index(['organization_id','flex_category','month'],inplace=True)
    flexibility_requirements.sort_index(inplace=True)

    # flex-rmr:
    columns = ['organization_id'] + month_columns
    data_range = get_data_range(year_ahead['Flex RMR'],'A','BCDEFGHIJKLM',config)
    flexibility_rmr = data_range_to_dataframe(columns,data_range)
    flexibility_rmr = flexibility_rmr.melt(id_vars=['organization_id'],var_name='month',value_name='flexibility_rmr')
    flexibility_rmr.set_index(['organization_id','month'],inplace=True)
    flexibility_rmr.sort_index(inplace=True)

    # flex-cme (new table for 2023):
    columns = ['organization_id'] + month_columns
    if filing_month.year>2022:
        data_range = get_data_range(year_ahead['CPE Flexible'],'A','BCDEFGHIJKLM',config)
        flexibility_cme = data_range_to_dataframe(columns,data_range)
        flexibility_cme = flexibility_cme.melt(id_vars=['organization_id'],var_name='month',value_name='flexibility_cme')
    else:
        lse_ids = [lse['id'] for lse in config.organizations.list_load_serving_entities()]
        flexibility_cme = pd.DataFrame({
            'organization_id':[lse_id_x for lse_id in lse_ids for lse_id_x in [lse_id]*12],'month':month_columns*len(lse_ids),'flexibility_cme':[0]*12*len(lse_ids)})
    flexibility_cme.set_index(['organization_id','month'],inplace=True)
    flexibility_cme.sort_index(inplace=True)

    # flex-irp (new table for 2024):
    columns = ['id'] + month_columns
    if filing_month.year>2023:
        flexibility_irp = get_table(year_ahead['IRP Flexible'],'Jan. EFC',{'rows':1,'columns':-1},columns)
    else:
        ids = [lse['id']+str(x) for lse in config.organizations.list_load_serving_entities() for x in range(1,4)]
        flexibility_irp = pd.DataFrame({
            **{'id':ids},
            **{month:[0]*len(ids) for month in month_columns}
        })
    flexibility_irp = flexibility_irp.melt(id_vars=['id'],var_name='month',value_name='flexibility_irp')
    id_re = re.compile(r'^(\w+)(\d)$')
    flexibility_irp.loc[:,'organization_id'] = flexibility_irp.loc[:,'id'].map(lambda s: id_re.match(s).groups()[0])
    flexibility_irp.loc[:,'flex_category'] = flexibility_irp.loc[:,'id'].map(lambda s: int(id_re.match(s).groups()[1]))
    flexibility_irp = flexibility_irp.loc[:,['organization_id','flex_category','month','flexibility_irp']]
    flexibility_irp.set_index(['organization_id','flex_category','month'],inplace=True)
    flexibility_irp.sort_index(inplace=True)

    # local cam:
    columns = [
        'organization_id',
        'los_angeles',
        'ventura',
        'san_diego',
        'bay_area',
        'fresno',
        'sierra',
        'stockton',
        'kern',
        'humboldt',
        'northern_california',
    ]
    data_range = get_data_range(year_ahead['Local RA-CAM-{}'.format(filing_month.year)],'B','CDEFGHIJKL',config)
    local_rar = data_range_to_dataframe(columns,data_range)
    local_rar.set_index('organization_id',inplace=True)
    local_rar.sort_index(inplace=True)
    local_rar.fillna(value=0,inplace=True)

    # total lcr:
    data_range = year_ahead['Local RA-CAM-{}'.format(filing_month.year)]['C2:L2']
    columns = list(map(location_renamer,[cell.value for cell in data_range[0]]))
    columns
    data_range = year_ahead['Local RA-CAM-{}'.format(filing_month.year)]['C4:L4']
    total_lcr = data_range_to_dataframe(columns,data_range)

    # cam system:
    columns = [
        'resource_name',
        'resource_id',
        'local_area',
        'procurement_status',
        'start_date',
        'end_date',
        'resource_mcc_bucket',
    ] + month_columns
    pge_sheet_re = re.compile('.*pge.*system.*')
    if len(list(filter(lambda s:pge_sheet_re.search(s.lower()),year_ahead.sheetnames)))>0:
        pge_system_sheetname = next(filter(lambda s:re.search('.*pge.*system.*',s.lower()),year_ahead.sheetnames))
        pge_cam_system = get_table(year_ahead[pge_system_sheetname],'CAM System RA Allocated (MW)',{'rows':2,'columns':-7},columns=columns)
        pge_cam_system.loc[:,'path_26_region'] = 'north'
    else:
        pge_cam_system = pd.DataFrame(columns=columns+['path_26_region'])
    sce_sheet_re = re.compile('.*sce.*system.*')
    if len(list(filter(lambda s:sce_sheet_re.search(s.lower()),year_ahead.sheetnames)))>0:
        sce_system_sheetname = next(filter(lambda s:re.search('.*sce.*system.*',s.lower()),year_ahead.sheetnames))
        sce_cam_system = get_table(year_ahead[sce_system_sheetname],'CAM System RA Allocated (MW)',{'rows':2,'columns':-7},columns=columns)
        sce_cam_system.loc[:,'path_26_region'] = 'south'
    else:
        sce_cam_system = pd.DataFrame(columns=columns+['path_26_region'])
    cam_system = pd.concat([pge_cam_system,sce_cam_system],ignore_index=True)

    # irp system:
    columns = ['id'] + month_columns
    id_re = re.compile(r'^(.+)([NS]P26)\s*IRP$')
    irp_system = get_table(year_ahead['IRP System'],'Jan. System NQC',{'rows':1,'columns':-1},columns)
    irp_system.loc[:,'lse'] = irp_system.loc[:,'id'].map(lambda s: id_re.match(s).groups()[0])
    irp_system.loc[(irp_system.loc[:,'id'].map(lambda s: id_re.match(s).groups()[1])=='NP26'),'path_26_region'] = 'north'
    irp_system.loc[(irp_system.loc[:,'id'].map(lambda s: id_re.match(s).groups()[1])=='SP26'),'path_26_region'] = 'south'
    irp_system = irp_system.loc[:,['lse','path_26_region']+month_columns]
    irp_system = irp_system.set_index(['lse','path_26_region']).sort_index()

    # year-ahead cam/rmr allocation:
    columns = ['id'] + month_columns
    if filing_month.year>=2024:
        id_re = re.compile(r'^(.+)([NS]P26|System)\s*(CAM|RMR)$')
        cam_rmr = get_table(year_ahead['Local RA-CAM-{}'.format(filing_month.year)],'YA CAM Allocatons (this flows into LSE table 8)',{'rows':13,'columns':-1},columns)
        cam_rmr.loc[:,'lse'] = cam_rmr.loc[:,'id'].map(lambda s: id_re.match(s).groups()[0] if id_re.match(s) else None)
        cam_rmr.loc[(cam_rmr.loc[:,'id'].map(lambda s: id_re.match(s).groups()[1])=='NP26'),'path_26_region'] = 'north'
        cam_rmr.loc[(cam_rmr.loc[:,'id'].map(lambda s: id_re.match(s).groups()[1])=='SP26'),'path_26_region'] = 'south'
        cam_rmr.loc[(cam_rmr.loc[:,'id'].map(lambda s: id_re.match(s).groups()[1])=='System'),'path_26_region'] = 'system'
        cam_rmr.loc[:,'type'] = cam_rmr.loc[:,'id'].map(lambda s: id_re.match(s).groups()[2].lower())
    else:
        cam_rmr = pd.DataFrame(columns=['lse','path_26_region','type']+columns)
    cam_rmr = cam_rmr.drop(columns=['id']).set_index(['lse','path_26_region','type']).sort_index()

    return (load_forecast_input_data,demand_response_allocation,cam_credits,flexibility_requirements,flexibility_rmr,flexibility_cme,flexibility_irp,local_rar,total_lcr,cam_system,irp_system,cam_rmr)

def get_incremental_local_tables(incremental_local,config:ConfigurationOptions):
    '''
    loads relevant data from the incremental local year-ahead adjustment
    workbook into dataframes

    parameters:
        incremental_local - incremental local year-ahead workbook
        config - an instance of the ConfigurationOptions class used for
            retrieving containing information about each load-serving entity
    '''
    columns = [
        'organization_id',
        1,
        2,
        3,
    ]
    data_range = get_data_range(incremental_local['IncrementalLocal'],'A','MNO',config)
    incremental_flex = data_range_to_dataframe(columns,data_range)
    incremental_flex = incremental_flex.melt(id_vars=['organization_id'],var_name='category',value_name='flexibility_requirement')
    incremental_flex.fillna({'flexibility_requirement':0},inplace=True)
    incremental_flex.set_index(['organization_id','category'],inplace=True)
    incremental_flex.sort_index(inplace=True)
    columns = [
        'organization_id',
        'los_angeles',
        'ventura',
        'san_diego',
        'bay_area',
        'fresno',
        'sierra',
        'stockton',
        'kern',
        'humboldt',
        'northern_california',
    ]
    data_range = get_data_range(incremental_local['IncrementalLocal'],'A','BCDEFGHIJK',config)
    incremental_local_load = data_range_to_dataframe(columns,data_range)
    incremental_local_load = incremental_local_load.melt(id_vars=['organization_id'],var_name='location',value_name='incremental_load')
    incremental_local_load.fillna({'incremental_load':0},inplace=True)
    incremental_local_load.set_index(['organization_id','location'],inplace=True)
    incremental_local_load.sort_index()
    columns = [
        'organization_id',
        'los_angeles',
        'ventura',
        'san_diego',
        'bay_area',
        'fresno',
        'sierra',
        'stockton',
        'kern',
        'humboldt',
        'northern_california',
        'sp26_cam_capacity',
        'sp26_condition2_rmr',
        'np26_condition2_rmr',
        'path26_ns',
        'path26_sn',
        'np26_cam_capacity',
    ]
    local_rar_trueup = get_table(incremental_local['Local Trueup'],'YA Local RAR Allocations',{'rows':6,'columns':1},columns)
    local_rar_trueup = local_rar_trueup.melt(id_vars=['organization_id'],var_name='location',value_name='local_rar_trueup')
    local_rar_trueup.fillna({'local_rar_trueup':0},inplace=True)
    local_rar_trueup.set_index(['organization_id','location'],inplace=True)
    local_rar_trueup.sort_index(inplace=True)


    return (incremental_flex,incremental_local_load,local_rar_trueup)

def get_cam_rmr_update_tables(cam_rmr_update,config:ConfigurationOptions):
    '''
    loads CAM, RMR, and Diablo Canyon credits from the mid-year credit true-up
    file into dataframes

    parameters:
        cam_rmr_update - mid-year credit true-up workbook
        config - an instance of the ConfigurationOptions class used for
            retrieving containing information about each load-serving entity
    '''
    n_organizations = len(config.organizations.list_load_serving_entities())
    month_columns = [ts(config.filing_month.year,month,1).to_numpy().astype('datetime64[M]') for month in range(1,13)]

    n_rows = n_organizations * 5
    data_range = cam_rmr_update['Jun to Dec CAM Update']['B14:N{}'.format(13+n_rows)]
    columns = ['row_id'] + month_columns
    id_parser = re.compile(r'(\w+)\s*(NP26|SP26|System)\s*(CAM|RMR)')
    cam_rmr_trueup = data_range_to_dataframe(columns,data_range)
    cam_rmr_trueup.loc[:,'row_id'] = cam_rmr_trueup.loc[:,'row_id'].fillna('')
    cam_rmr_trueup = cam_rmr_trueup.loc[(cam_rmr_trueup.loc[:,'row_id'].map(lambda s: id_parser.match(s) is not None)),:]
    columns = ['organization_id','path_26_region','type']
    cam_rmr_trueup = cam_rmr_trueup.join(
        cam_rmr_trueup.apply(
            lambda r: {k:v for k,v in zip(columns,id_parser.match(r.loc['row_id']).groups())},
            axis='columns',
            result_type='expand'
        )
    )
    cam_rmr_trueup = cam_rmr_trueup.fillna(0)
    cam_rmr_trueup.loc[:,'path_26_region'] = cam_rmr_trueup.loc[:,'path_26_region'].map(lambda s: 'north' if s=='NP26' else 'south' if s=='SP26' else 'system')
    cam_rmr_trueup.loc[:,'type'] = cam_rmr_trueup.loc[:,'type'].map(lambda s: s.lower())
    cam_rmr_trueup.set_index(['organization_id','path_26_region','type'],inplace=True)

    data_range = cam_rmr_update['Diablo Canyon Credits']['B7:N{}'.format(6+n_organizations)]
    columns = ['organization_id'] + month_columns
    diablo_canyon_credits = data_range_to_dataframe(columns,data_range)
    diablo_canyon_credits = diablo_canyon_credits.dropna(axis='index',how='any',subset=['organization_id'])
    diablo_canyon_credits = diablo_canyon_credits.fillna(0)
    diablo_canyon_credits.set_index('organization_id',inplace=True)

    return (cam_rmr_trueup,diablo_canyon_credits)

def get_month_ahead_tables(month_ahead,config:ConfigurationOptions):
    '''
    loads relevant data from the month-ahead forecasts workbook into dataframes
        parameters:
            month_ahead - month-ahead workbook
        returns:
            [month_ahead_forecasts,monthly_tracking] - list of extracted
                dataframes
    '''
    columns = [
        'organization_id',
        'lse_type',
        'jurisdiction',
        'lse_lu',
        'month',
        'id_and_date',
        'sce_year_ahead_forecast',
        'sdge_year_ahead_forecast',
        'pge_year_ahead_forecast',
        'total_year_ahead_forecast',
        'sce_esps_migrating_load',
        'sce_cca_migrating_load',
        'sdge_esps_migrating_load',
        'sdge_cca_migrating_load',
        'pge_esps_migrating_load',
        'pge_cca_migrating_load',
        'sce_migration_adjustment',
        'sdge_migration_adjustment',
        'pge_migration_adjustment',
        'total_migration_adjustment',
        'sce_revised_monthly_forecast',
        'sdge_revised_monthly_forecast',
        'pge_revised_monthly_forecast',
        'total_revised_monthly_forecast',
        'sce_revised_noncoincident_monthly_forecast',
        'sdge_revised_noncoincident_monthly_forecast',
        'pge_revised_noncoincident_monthly_forecast',
        'total_revised_noncoincident_monthly_forecast',
        'sce_revised_nonjurisdictional_load_share',
        'sdge_revised_nonjurisdictional_load_share',
        'pge_revised_nonjurisdictional_load_share',
        'total_revised_nonjurisdictional_load_share',
        'sce_revised_jurisdictional_load_share',
        'sdge_revised_jurisdictional_load_share',
        'pge_revised_jurisdictional_load_share',
        'total_revised_jurisdictional_load_share',
    ]
    if config.filing_month>ts(2022,12,31):
        data_range = month_ahead['Monthly Tracking']['B5:AK{}'.format(month_ahead['Monthly Tracking'].max_row)]
    else:
        data_range = month_ahead['Monthly Tracking for CPUC']['B5:AK{}'.format(month_ahead['Monthly Tracking for CPUC'].max_row)]
    month_ahead_forecasts = data_range_to_dataframe(columns,data_range)
    month_ahead_forecasts['organization_id'] = month_ahead_forecasts.loc[:,'organization_id'].map(lambda s: s.replace('Total','').strip() if isinstance(s,str) else s)
    month_ahead_forecasts.set_index(['organization_id','month'],inplace=True)
    month_ahead_forecasts.sort_index(inplace=True)

    # monthly tracking table
    columns = [
        'organization_id',
        'lse_type',
        'jurisdiction',
        'lse_lu',
        'month',
        'id_and_date',
        'sce_year_ahead_forecast',
        'sdge_year_ahead_forecast',
        'pge_year_ahead_forecast',
        'total_year_ahead_forecast',
        'sce_esps_migrating_load',
        'sce_cca_migrating_load',
        'sdge_esps_migrating_load',
        'sdge_cca_migrating_load',
        'pge_esps_migrating_load',
        'pge_cca_migrating_load',
        'sce_load_migration_adjustment',
        'sdge_load_migration_adjustment',
        'pge_load_migration_adjustment',
        'total_load_migration_adjustment',
        'sce_revised_monthly_forecast',
        'sdge_revised_monthly_forecast',
        'pge_revised_monthly_forecast',
        'total_revised_monthly_forecast',
        'sce_revised_noncoincident_monthly_forecast',
        'sdge_revised_noncoincident_monthly_forecast',
        'pge_revised_noncoincident_monthly_forecast',
        'total_revised_noncoincident_monthly_forecast',
        'sce_revised_nonjurisdictional_load_share',
        'sdge_revised_nonjurisdictional_load_share',
        'pge_revised_nonjurisdictional_load_share',
        'total_revised_nonjurisdictional_load_share',
        'sce_revised_jurisdictional_load_share',
        'sdge_revised_jurisdictional_load_share',
        'pge_revised_jurisdictional_load_share',
        'total_revised_jurisdictional_load_share',
    ]
    data_range = month_ahead['Monthly Tracking']['B5:AK{}'.format(month_ahead['Monthly Tracking'].max_row)]
    monthly_tracking = data_range_to_dataframe(columns,data_range)
    monthly_tracking.dropna(subset=['organization_id','month'],inplace=True)
    monthly_tracking['organization_id'] = monthly_tracking.loc[:,'organization_id'].map(lambda s: s.replace('Total','').strip() if isinstance(s,str) else s)
    monthly_tracking.set_index(['organization_id','month'],inplace=True)
    monthly_tracking.sort_index(inplace=True)

    return (month_ahead_forecasts,monthly_tracking)

def get_cam_rmr_tables(cam_rmr):
    '''
    loads relevant data from the cam-rmr workbook into dataframes
        parameters:
            cam_rmr - cam-rmr workbook
        returns:
            (cam_rmr_monthly_tracking,total_cam_rmr) - tuple of extracted
                dataframes
    '''
    columns = [
        'organization_id',
        'lse_type',
        'jurisdiction',
        'lse_lu',
        'month',
        'id_and_date',
        'sce_year_ahead_forecast',
        'sdge_year_ahead_forecast',
        'pge_year_ahead_forecast',
        'total_year_ahead_forecast',
        'sce_esps_migrating_load',
        'sce_cca_migrating_load',
        'sdge_esps_migrating_load',
        'sdge_cca_migrating_load',
        'pge_esps_migrating_load',
        'pge_cca_migrating_load',
        'sce_load_migration_adjustment',
        'sdge_load_migration_adjustment',
        'pge_load_migration_adjustment',
        'total_load_migration_adjustment',
        'sce_revised_monthly_forecast',
        'sdge_revised_monthly_forecast',
        'pge_revised_monthly_forecast',
        'total_revised_monthly_forecast',
        'sce_revised_noncoincident_monthly_forecast',
        'sdge_revised_noncoincident_monthly_forecast',
        'pge_revised_noncoincident_monthly_forecast',
        'total_revised_noncoincident_monthly_forecast',
        'sce_revised_nonjurisdictional_load_share',
        'sdge_revised_nonjurisdictional_load_share',
        'pge_revised_nonjurisdictional_load_share',
        'total_revised_nonjurisdictional_load_share',
        'blank',
        'sce_revised_jurisdictional_load_share',
        'sdge_revised_jurisdictional_load_share',
        'pge_revised_jurisdictional_load_share',
        'total_revised_jurisdictional_load_share',
    ]
    data_range = cam_rmr['monthlytracking']['B5:AL{}'.format(cam_rmr['monthlytracking'].max_row)]
    cam_rmr_monthly_tracking = data_range_to_dataframe(columns,data_range)
    cam_rmr_monthly_tracking.drop('blank',axis='columns',inplace=True)
    cam_rmr_monthly_tracking.dropna(subset=['organization_id','month'],inplace=True)
    cam_rmr_monthly_tracking['organization_id'] = cam_rmr_monthly_tracking.loc[:,'organization_id'].map(lambda s: s.replace('Total','').strip() if isinstance(s,str) else s)
    cam_rmr_monthly_tracking.set_index(['organization_id','month'],inplace=True)
    cam_rmr_monthly_tracking.sort_index(inplace=True)
    columns = [
        'np26_cam',
        'sp26_cam',
        'np26_rmr',
        'sp26_rmr',
        'system_rmr',
        'sce_preferred_lcr_credit',
        'sdge_cam',
        'sce_cam',
    ]
    data_range = cam_rmr['CAMRMR']['B4:I4']
    total_cam_rmr = pd.Series(data=[cell.value for cell in data_range[0]],index=columns)

    return (cam_rmr_monthly_tracking,total_cam_rmr)

def get_summary_tables(ra_summary,config:ConfigurationOptions):
    '''
    loads select data from a summary workbook into dataframes

    parameters:
        ra_summary - a pre-loaded resource adequacy monthly summary workbook
        config - an instance of the ConfigurationOptions class
    '''
    columns = [
        'organization_id',
        'resource_adequacy_obligation',
        'physical_resources',
        'demand_response_resources',
        'total_resources',
        'percent_obligation_available',
    ]
    data_range = get_data_range(ra_summary['Summary'],'A','BCDEF',config)
    summary = data_range_to_dataframe(columns,data_range)
    summary.set_index('organization_id',inplace=True)
    columns = [
        'organization_id',
        'resource_adequacy_obligation',
        'total_flex_capacity',
        'percent_obligation_available',
        'category_1_rar',
        'category_1_countable',
        'category_2_rar',
        'category_2_countable',
        'category_3_rar',
        'category_3_countable',
        'category_1_flex',
        'category_2_flex',
        'category_3_flex',
        'total_flex',
    ]
    data_range = get_data_range(ra_summary['FlexRAR'],'A','BCDEFGHIJLMNO',config)
    flex_rar = data_range_to_dataframe(columns,data_range)
    flex_rar.set_index('organization_id',inplace=True)
    columns = [
        'organization_id',
        'los_angeles_rar','los_angeles_incremental_adjustment','los_angeles_demand_response_allocation','los_angeles_procurement','los_angeles_compliance',
        'ventura_rar','ventura_incremental_adjustment','ventura_demand_response_allocation','ventura_procurement','ventura_compliance',
        'san_diego_rar','san_diego_incremental_adjustment','san_diego_demand_response_allocation','san_diego_procurement','san_diego_compliance',
        'bay_area_rar','bay_area_incremental_adjustment','bay_area_demand_response_allocation','bay_area_procurement','bay_area_compliance',
        'humboldt_rar','humboldt_incremental_adjustment','humboldt_demand_response_allocation','humboldt_procurement','humboldt_compliance',
        'sierra_rar','sierra_incremental_adjustment','sierra_demand_response_allocation','sierra_procurement','sierra_compliance',
        'stockton_rar','stockton_incremental_adjustment','stockton_demand_response_allocation','stockton_procurement','stockton_compliance',
        'northern_california_rar','northern_california_incremental_adjustment','northern_california_demand_response_allocation','northern_california_procurement','northern_california_compliance',
        'fresno_rar','fresno_incremental_adjustment','fresno_demand_response_allocation','fresno_procurement','fresno_compliance',
        'kern_rar','kern_incremental_adjustment','kern_demand_response_allocation','kern_procurement','kern_compliance',
        'pge_other_rar','pge_other_incremental_adjustment','pge_other_demand_response_allocation','pge_other_procurement','pge_other_compliance',
        'sce_tac_other_rar','sce_tac_other_incremental_adjustment','sce_tac_other_demand_response_allocation','sce_tac_other_procurement','sce_tac_other_compliance',
        'sdge_tac_other_rar','sdge_tac_other_incremental_adjustment','sdge_tac_other_demand_response_allocation','sdge_tac_other_procurement','sdge_tac_other_compliance',
        'pge_tac_other_rar','pge_tac_other_incremental_adjustment','pge_tac_other_demand_response_allocation','pge_tac_other_procurement','pge_tac_other_compliance',
    ]
    data_range = get_data_range(ra_summary['LocalTrueUp'],'A',list(map(get_column_letter,range(2,72))),config)
    local_trueup = data_range_to_dataframe(columns,data_range)
    local_trueup.set_index('organization_id',inplace=True)

    return (summary,flex_rar,local_trueup)

def get_cross_check_tables(caiso_cross_check,config:ConfigurationOptions):
    '''
    loads select data from a caiso cross-check workbook into dataframes.

    parameters:
        caiso_cross_check - a preloaded caiso cross-check workook
        config - an instance of the ConfigurationOptions class
    '''
    # system requirements and availability:
    columns = [
        'organization_id',
        'resources_required',
        'physical_resources_available',
        'demand_response_resources_available',
        'total_resources_available',
        'percent_required_resources_available',
    ]
    data_range = get_data_range(caiso_cross_check['Requirements'],'A','BCDEF',config)
    system = data_range_to_dataframe(columns,data_range)
    system.set_index('organization_id',inplace=True)

    # flexibility requirements and availability:
    columns = [
        'organization_id',
        'flexibility_requirements',
        'flexibility_available',
        'percent_flexibility_requirements_available',
        'flexibility_category_1',
        'flexibility_category_1_countable',
        'flexibility_category_2',
        'flexibility_category_2_countable',
        'flexibility_category_3',
        'flexibility_category_3_countable',
        'year_ahead_flexibility_category_1',
        'year_ahead_flexibility_category_2',
        'year_ahead_flexibility_category_3',
        'year_ahead_flex_total',

    ]
    data_range = get_data_range(caiso_cross_check['Requirements'],'A','HIJKLMNOPQRST',config)
    flexibility = data_range_to_dataframe(columns,data_range)
    flexibility.set_index('organization_id',inplace=True)

    return (system,flexibility)

def get_nqc_list(nqc_workbook,config:ConfigurationOptions):
    '''
    retrieves the monthly generation capacities from a given nqc workbook

    parameters:
        nqc_workbook - a workbook containing a valid nqc list
        config - an instance of the ConfigurationOptions class
    '''
    filing_month = config.filing_month
    current_nqc_worksheet_name = next(filter(lambda s: re.match(r'^.*{}.*NQC List.*$'.format(filing_month.year),s), nqc_workbook.sheetnames))
    month_columns = [ts(filing_month.year,month,1).to_numpy().astype('datetime64[M]') for month in range(1,13)]
    column_map = {
        'Generator Name' : 'generator_name',
        'Resource ID': 'resource_id',
        'Path Designation' : 'zone',
        'Local Area' : 'local_area',
        'Dispatchable' : 'dispatchable',
        'Deliverability Status' : 'deliverability_status',
        'Deliverability MW' : 'deliverable',
        'Comments' : 'comments',
    }
    for month in month_columns:
        month_abbreviation = ts(month).strftime('%b').upper()
        column_map[month_abbreviation] = month
    if current_nqc_worksheet_name in [ws.title for ws in nqc_workbook.worksheets]:
        last_column = get_column_letter(20)
        last_row = nqc_workbook[current_nqc_worksheet_name].max_row
        original_columns = nqc_workbook[current_nqc_worksheet_name][f'A1:{last_column}1'][0]
        mapped_columns = [column_map[cell.value] for cell in original_columns]
        month_columns = [ts(filing_month.year,month,1).to_numpy().astype('datetime64[M]') for month in range(1,13)]
        data_range = nqc_workbook[current_nqc_worksheet_name][f'A2:{last_column}{last_row}']
        nqc_list = data_range_to_dataframe(mapped_columns, data_range)

        # identify rows with duplicate resource ids:
        nqc_list.loc[:,'count'] = 1
        nqc_list = nqc_list.drop(columns='count').merge(
            right=nqc_list.loc[:,['resource_id','count']].groupby('resource_id').count(),
            how='left',
            on='resource_id'
        )

        # identify rows with empty data:
        nqc_list.loc[:,'empty'] = reduce(
            lambda s,c: c & s,
            [pd.isna(nqc_list.loc[:,s]) for s in month_columns]
        )

        # remove rows with duplicate resource ids and empty data:
        nqc_list = nqc_list.loc[~(
            (nqc_list.loc[:,'count']>1) &
            (nqc_list.loc[:,'empty'])
        ),:].drop(columns=['count','empty'])

        # sort table and use only the first rows:
        nqc_list = nqc_list.sort_values(['resource_id']+month_columns)
        nqc_list.groupby('resource_id').first()
    else:
        columns = [
            'generator_name',
            'resource_id',
            'zone',
            'local_area'
        ] + month_columns + [
            'dispatchable',
            'deliverability_status',
            'deliverable',
            'comments',
        ]
        nqc_list = pd.DataFrame(columns=columns)
    return nqc_list

def refresh_nqc_list_from_file(nqc,ra_summary,config:ConfigurationOptions):
    '''
    extracts updated generation capacities from an nqc file and replaces the
    data in the nqc list worksheet of the resource adequacy summary sheet.
    parameters:
        nqc - a workbook containing updated nqc values
        ra_summary - a workbook containing the monthly resource adequacy summary
            data
        config - an instance of the ConfigurationOptions class
    '''
    filing_month = config.filing_month
    current_nqc_worksheet_name = f'{filing_month.year} NQC List'
    if current_nqc_worksheet_name in [ws.title for ws in nqc.worksheets]:
        last_column = get_column_letter(nqc[current_nqc_worksheet_name].max_column)
        last_row = nqc[current_nqc_worksheet_name].max_row
        columns = nqc[current_nqc_worksheet_name][f'A1:{last_column}1']
        data_range = nqc[current_nqc_worksheet_name][f'A2:{last_column}{last_row}']
        current_nqc_list = data_range_to_dataframe(columns, data_range)

        # clear nqc list in summary workbook:
        last_row = ra_summary['NQC_List'].max_row
        ra_summary['NQC_List'].delete_rows(2,max(2,last_row-1))

        # input new nqc list in summary workbook:
        row_number = 2
        for _,s in current_nqc_list.iterrows():
            ra_summary['NQC_List'][f'A{row_number}'].value = s.loc['Generator Name']
            ra_summary['NQC_List'][f'B{row_number}'].value = s.loc['Resource ID']
            ra_summary['NQC_List'][f'C{row_number}'].value = s.loc['Path 26 Region']
            ra_summary['NQC_List'][f'D{row_number}'].value = s.loc['Local Area']
            ra_summary['NQC_List'][f'E{row_number}'].value = s.loc['JAN']
            ra_summary['NQC_List'][f'F{row_number}'].value = s.loc['FEB']
            ra_summary['NQC_List'][f'G{row_number}'].value = s.loc['MAR']
            ra_summary['NQC_List'][f'H{row_number}'].value = s.loc['APR']
            ra_summary['NQC_List'][f'I{row_number}'].value = s.loc['MAY']
            ra_summary['NQC_List'][f'J{row_number}'].value = s.loc['JUN']
            ra_summary['NQC_List'][f'K{row_number}'].value = s.loc['JUL']
            ra_summary['NQC_List'][f'L{row_number}'].value = s.loc['AUG']
            ra_summary['NQC_List'][f'M{row_number}'].value = s.loc['SEP']
            ra_summary['NQC_List'][f'N{row_number}'].value = s.loc['OCT']
            ra_summary['NQC_List'][f'O{row_number}'].value = s.loc['NOV']
            ra_summary['NQC_List'][f'P{row_number}'].value = s.loc['DEC']
            ra_summary['NQC_List'][f'Q{row_number}'].value = s.loc['Dispatchable']
            ra_summary['NQC_List'][f'R{row_number}'].value = s.loc['Deliverability Status']
            ra_summary['NQC_List'][f'S{row_number}'].value = s.loc['Deliverable MW']
            ra_summary['NQC_List'][f'T{row_number}'].value = s.loc['Comments']
            row_number+=1

def read_supply_plan(config:ConfigurationOptions,supply_plan_type:str,date:ts=None,version:int=None):
    '''
    loads data from caiso supply plan workbooks returns the data as a 2-tuple of
    dataframes.

    parameters:
        config - an instance of the ConfigurationOptions class
        date - an optional datetime object representing the filing month; if
            left blank, the filing month defined in the paths object will be applied
        version - an optional integer version number; if left blank, the
            highest numbered version of a filing for the given load-serving
            entity and month will be used
    '''
    path = config.paths.get_path(supply_plan_type,date=date,version=version)
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    if supply_plan_type=='supply_plan_system':
        columns = [
            'validation_status',
            'supplier',
            'resource_id',
            'local_resource_adequacy',
            'system_resource_adequacy',
            'total_capacity',
            'start_date',
            'end_date',
            'organization_id_caiso',
            'errors_and_warnings',
        ]
        supply_plan = pd.DataFrame(columns=columns)
        for row_number in range(1,sheet.nrows):
            new_row = [sheet.cell_value(rowx=row_number,colx=column_number) for column_number in range(sheet.ncols)]
            supply_plan.loc[len(supply_plan)] = new_row
    elif supply_plan_type=='supply_plan_flexible':
        columns = [
            'validation_status',
            'supplier',
            'resource_id',
            'category',
            'flex_capacity',
            'start_date',
            'end_date',
            'organization_id_caiso',
            'errors_and_warnings',
        ]
        supply_plan = pd.DataFrame(columns=columns)
        sheet = workbook.sheet_by_index(0)
        for row_number in range(1,sheet.nrows):
            new_row = [sheet.cell_value(rowx=row_number,colx=column_number) for column_number in range(sheet.ncols)]
            supply_plan.loc[len(supply_plan)] = new_row

    return supply_plan

def rename_locality(locality:str):
    '''
    reads a non-standard local area and returns a standardized locality name.

    parameters:
        locality - string containing a non-standard local area name
    '''
    if not pd.isnull(locality):
        if re.match('.*bay.*',locality.lower()):
            standard_locality_name = 'bay_area'
        elif re.match('.*fres.*',locality.lower()):
            standard_locality_name = 'fresno'
        elif re.match('.*humb.*',locality.lower()):
            standard_locality_name = 'humboldt'
        elif re.match('.*kern.*',locality.lower()):
            standard_locality_name = 'kern'
        elif re.match('.*basi.*',locality.lower()):
            standard_locality_name = 'los_angeles'
        elif re.match('.*ncnb.*',locality.lower()):
            standard_locality_name = 'northern_california'
        elif re.match('.*?(dieg|sdge).*',locality.lower()):
            standard_locality_name = 'san_diego'
        elif re.match('.*sier.*',locality.lower()):
            standard_locality_name = 'sierra'
        elif re.match('.*stoc.*',locality.lower()):
            standard_locality_name = 'stockton'
        elif re.match('.*vent.*',locality.lower()):
            standard_locality_name = 'ventura'
        else:
            standard_locality_name = 'caiso_system'
    else:
        standard_locality_name = 'caiso_system'
    return standard_locality_name

def parse_date(date_string):
    if isinstance(date_string,str):
        excel_date = re.match(r'(\d{1,2})/(\d{1,2})/(\d+)',date_string)
        if excel_date:
            month_string,day_string,year_string = excel_date.groups()
            if int(year_string)<1850:
                year_string = '1850'
            elif int(year_string)>2199:
                year_string = '2199'
            else:
                pass
            if int(month_string)<1:
                month_string = '1'
            elif int(month_string)>12:
                month_string = '12'
            next_month = int(month_string)%12+1
            next_year = int(year_string) + int(int(month_string)/12)
            last_day_of_month = (ts(next_year,next_month,1) - td(days=1)).day
            if int(day_string)<1:
                day_string = '1'
            elif int(day_string)>last_day_of_month:
                day_string = str(last_day_of_month)
            else:
                pass
            date = ts(int(year_string),int(month_string),int(day_string))
        else:
            try:
                date = ts(date_string)
            except:
                date = ts('NaT')
    elif isinstance(date_string,time):
        date = ts(1900,1,1)
    else:
        date = ts('NaT')
    return date
def location_renamer(s:str):
    if isinstance(s,str):
        s = s.lower()
        if 'basin' in s:
            s_out = 'los_angeles'
        elif 'vent' in s:
            s_out = 'ventura'
        elif 'diego' in s or 'sdge' in s:
            s_out = 'san_diego'
        elif 'bay' in s:
            s_out = 'bay_area'
        elif 'fresno' in s:
            s_out = 'fresno'
        elif 'sierra' in s:
            s_out = 'sierra'
        elif 'stock' in s:
            s_out = 'stockton'
        elif 'kern' in s:
            s_out = 'kern'
        elif 'humb' in s:
            s_out = 'humboldt'
        elif 'ncnb' in s:
            s_out = 'northern_california'
        else:
            s_out = 'non-lcr'
    else:
        s_out = 'non-lcr'
    return s_out