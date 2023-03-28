from re import I
import pandas as pd
from pandas import Timestamp as ts
from openpyxl.styles import Alignment

from ra_logging import TextLogger,EmailLogger,AttachmentLogger,ConsolidationLogger
from configuration_options import ConfigurationOptions
from data_extraction import *

regions_to_service_territories = {
    'SCE' : ['los_angeles','ventura','outside_lca'],
    'PGE' : ['bay_area','other_pge_areas','non-lcr','fresno','sierra','stockton','kern','humboldt','northern_california'],
    'SDGE' : ['san_diego'],
    'CAISO' : ['caiso_system'],
}

class DataExporter:
    '''
    this class assists in extracting data from source workbooks and exporting
    them into tables for inclusion in EZDB.
    '''
    def __init__(self,configuration_path:Path,filing_month:ts=None):
        '''
        initializes an instance of the ra_consolidator class
        parameters:
            configuration_path - path object pointing to a yaml file containing
                configuration options for the ra_consolidator object
            filing_month - an optional filing month timestamp to overwrite the
                date in the configuration options yaml file
        '''
        self.config = ConfigurationOptions(configuration_path,filing_month=filing_month)
        self.logger = TextLogger(
            self.config.get_option('cli_logging_criticalities'),
            self.config.get_option('file_logging_criticalities'),
            self.config.paths.get_path('log')
        )
        self.email_logger = EmailLogger(self.config.paths.get_path('email_log'))
        self.attachment_logger = AttachmentLogger(self.config.paths.get_path('attachment_log'))
        self.consolidation_logger = ConsolidationLogger(self.config.paths.get_path('consolidation_log'))

    def write_organizations(self):
        '''
        transforms the organizations dictionary into a dataframe transforms to a
        data table format for ezdb, and writes to csv.
        '''
        organizations = pd.DataFrame(self.config.organizations.data).transpose().reset_index()
        columns = [
            'CPUCAbbreviation',
            'CAISOAbbreviation',
            'PALCorporationID',
            'FullName',
            'Aliases',
            'Type',
            'Comment',
        ]
        organizations.rename({
            'index':'CPUCAbbreviation',
            'scid':'CAISOAbbreviation',
            'pal_id':'PALCorporationID',
            'aliases':'Aliases',
            'type':'Type'
        },axis='columns',inplace=True)
        organizations.loc[:,'FullName'] = organizations.loc[:,'Aliases'].map(lambda l: l[0])
        organizations.loc[:,'Aliases'] = organizations.loc[:,'Aliases'].map(lambda l: ';'.join(l))
        organizations.loc[:,'Comment'] = ''
        organizations = organizations.loc[:,columns]

        out_path = self.config.paths.get_path('ezdb_organizations')
        out_path.parent.mkdir(parents=True,exist_ok=True)
        self.logger.log('Exporting Organizations for EZDB to {}'.format(out_path.name),'INFORMATION')
        organizations.to_csv(out_path,index=False,sep=',',quotechar='"')

    def write_data_sources(self):
        data_sources = self.consolidation_logger.data.merge(self.attachment_logger.data.merge(self.email_logger.data,on='email_id',suffixes=('','_emails')),on='attachment_id',suffixes=('','_attachments'))
        data_sources.rename({
            'email_id':'EmailID',
            'subject':'EmailSubject',
            'sender':'EmailSender',
            'receipt_date':'EmailReceiptDatetime',
            'attachment_id':'AttachmentID',
            'download_path':'DownloadPath',
            'archive_path':'ArchivePath',
            'ra_category':'Category',
            'organization_id':'SourceOrganization',
            'effective_date':'EffectiveDate'
        },axis='columns',inplace=True)
        def get_version_number(r):
            archive_path = r.loc['ArchivePath']
            path_id = r.loc['Category']
            organization = self.config.organizations.get_organization(r.loc['SourceOrganization'])
            date = self.config.filing_month
            return self.config.paths.get_version_number(archive_path,path_id,organization,date)
        data_sources.loc[:,'Version'] = data_sources.apply(get_version_number,axis='columns')
        data_sources.loc[:,'Comment'] = ''
        data_sources = data_sources.loc[:,[
            'EmailID',
            'EmailSubject',
            'EmailSender',
            'EmailReceiptDatetime',
            'AttachmentID',
            'DownloadPath',
            'ArchivePath',
            'Category',
            'SourceOrganization',
            'EffectiveDate',
            'Version',
            'Comment'
        ]]

        # add summary and caiso cross-check files:
        ra_categories = ['ra_summary','caiso_cross_check']
        archive_paths = [self.config.paths.get_path(ra_category) for ra_category in ra_categories]
        attachment_ids = [self.attachment_logger.data.loc[self.attachment_logger.data.loc[:,'archive_path']==str(archive_path),'attachment_id'].iloc[0] for archive_path in archive_paths]
        email_ids = [self.attachment_logger.data.loc[self.attachment_logger.data.loc[:,'archive_path']==str(archive_path),'email_id'].iloc[0] for archive_path in archive_paths]
        email_senders = [self.email_logger.data.loc[self.email_logger.data.loc[:,'email_id']==email_id,'sender'].iloc[0] for email_id in email_ids]
        email_receipt_datetimes = [self.email_logger.data.loc[self.email_logger.data.loc[:,'email_id']==email_id,'receipt_date'].iloc[0] for email_id in email_ids]
        data_sources = data_sources.append(pd.DataFrame({
            'EmailID' : email_ids,
            'EmailSubject' : ['',''],
            'EmailSender' : email_senders,
            'EmailReceiptDatetime' : email_receipt_datetimes,
            'AttachmentID' : attachment_ids,
            'DownloadPath' : ['',''],
            'ArchivePath' : archive_paths,
            'Category' : ra_categories,
            'SourceOrganization' : ['CPUC','CPUC'],
            'EffectiveDate' : [self.config.filing_month] * 2,
            'Version' : [0,0],
            'Comment' : ['',''],
        }))

        # reformat date column:
        data_sources.loc[:,'EffectiveDate'] = pd.to_datetime(data_sources.loc[:,'EffectiveDate']).dt.date

        out_path = self.config.paths.get_path('ezdb_data_sources')
        out_path.parent.mkdir(parents=True,exist_ok=True)
        self.logger.log('Exporting Data Sources for EZDB to {}'.format(out_path.name),'INFORMATION')
        data_sources.to_csv(out_path,index=False,sep=',',quotechar='"')

    def write_requirements(self):
        columns = [
            'LoadServingEntity',
            'ServiceTerritory',
            'FilingMonth',
            'Locality',
            'Path26Region',
            'RequirementType',
            'RequirementValue',
            'DataSource',
            'Version',
            'Comment',
        ]
        requirements = pd.DataFrame(columns=columns)

        # year-ahead:
        archive_path = self.config.paths.get_path('year_ahead')
        if archive_path:
            attachment_id = self.attachment_logger.data.loc[self.attachment_logger.data.loc[:,'archive_path']==str(archive_path),'attachment_id']
            version = self.config.paths.get_version_number(archive_path,'year_ahead')
            wb = open_workbook(archive_path)
            load_forecast_input_data,demand_response_allocation,cam_credits,flexibility_requirements,flexibility_rmr,flexibility_cme,local_rar,total_lcr,cam_system = get_year_ahead_tables(wb,self.config)

            # load forecasts:
            load_forecast_input_data.reset_index(inplace=True)
            load_forecast_input_data.rename({
                'organization_id' : 'LoadServingEntity',
                'iou_territory' : 'ServiceTerritory',
                'month' : 'FilingMonth',
            },axis='columns',inplace=True)
            load_forecast_input_data.drop(columns=['lse_type'],inplace=True)
            load_forecast_input_data = pd.melt(
                load_forecast_input_data,
                id_vars=['LoadServingEntity','ServiceTerritory','FilingMonth'],
                var_name='RequirementType',
                value_name='RequirementValue',
                ignore_index=True
            )
            load_forecast_input_data.loc[:,['Locality','Path26Region','Comment']] = ''
            load_forecast_input_data.loc[:,'DataSource'] = [attachment_id] * len(load_forecast_input_data)
            load_forecast_input_data.loc[:,'Version'] = version
            requirements = requirements.append(load_forecast_input_data,ignore_index=True)

            # demand response:
            demand_response_allocation.reset_index(inplace=True)
            demand_response_allocation.rename({
                'location' : 'Locality',
                'month' : 'FilingMonth',
                'allocation_type' : 'RequirementType',
                'allocation' : 'RequirementValue',
            },axis='columns',inplace=True)
            demand_response_allocation.loc[:,'RequirementType'] = demand_response_allocation.loc[:,'RequirementType'].map(lambda s: s+'_demand_response')
            demand_response_allocation.loc[:,'ServiceTerritory'] = demand_response_allocation.loc[:,'Locality'].map(lambda s:[organization_id for organization_id,localities in regions_to_service_territories.items() if s in localities][0])
            demand_response_allocation.loc[:,['LoadServingEntity','Path26Region','Comment']] = ''
            demand_response_allocation.loc[:,'Version'] = version
            demand_response_allocation.loc[:,'DataSource'] = [attachment_id] * len(demand_response_allocation)
            requirements = requirements.append(demand_response_allocation,ignore_index=True)

            # cam credits:
            cam_credits.reset_index(inplace=True)
            cam_credits.rename({
                'iou_territory' : 'ServiceTerritory',
                'category' : 'RequirementType',
                'month' : 'FilingMonth',
                'cam_credit' : 'RequirementValue',
            },axis='columns',inplace=True)
            cam_credits.loc[:,'RequirementType'] = cam_credits.loc[:,'RequirementType'].map(lambda x: 'cam_category{:.0f}'.format(x))
            cam_credits.loc[:,['LoadServingEntity','Locality','Path26Region']] = ''
            cam_credits.loc[:,'DataSource'] = [attachment_id] * len(cam_credits)
            cam_credits.loc[:,'Version'] = version
            requirements = requirements.append(cam_credits,ignore_index=True)

            # flexibility requirements:
            flexibility_requirements.reset_index(inplace=True)
            flexibility_requirements.rename({
                'organization_id' : 'LoadServingEntity',
                'flex_category' : 'RequirementType',
                'month' : 'FilingMonth',
                'flexibility_requirement' : 'RequirementValue',
            },axis='columns',inplace=True)
            flexibility_requirements.loc[:,'RequirementType'] = flexibility_requirements.loc[:,'RequirementType'].map(lambda x: 'flex_category{:.0f}'.format(x))
            flexibility_requirements.loc[:,['ServiceTerritory','Locality','Path26Region','Comment']] = ''
            flexibility_requirements.loc[:,'DataSource'] = [attachment_id] * len(flexibility_requirements)
            flexibility_requirements.loc[:,'Version'] = version
            requirements = requirements.append(flexibility_requirements,ignore_index=True)

            # flexibility_rmr:
            flexibility_rmr.reset_index(inplace=True)
            flexibility_rmr.rename({
                'organization_id' : 'LoadServingEntity',
                'month' : 'FilingMonth',
                'flexibility_rmr' : 'RequirementValue',
            },axis='columns',inplace=True)
            flexibility_rmr.loc[:,'RequirementType'] = 'flex_rmr'
            flexibility_rmr.loc[:,['ServiceTerritory','Locality','Path26Region','Comment']] = ''
            flexibility_rmr.loc[:,'DataSource'] = [attachment_id] * len(flexibility_rmr)
            flexibility_rmr.loc[:,'Version'] = version
            requirements = requirements.append(flexibility_rmr,ignore_index=True)

            # local rar:
            local_rar.reset_index(inplace=True)
            local_rar.rename({
                'organization_id' : 'LoadServingEntity',
            },axis='columns',inplace=True)
            local_rar = pd.melt(
                local_rar,
                id_vars=['LoadServingEntity'],
                var_name='Locality',
                value_name='RequirementValue',
                ignore_index=True
            )
            local_rar.loc[:,'ServiceTerritory'] = local_rar.loc[:,'Locality'].map(lambda s:[organization_id for organization_id,localities in regions_to_service_territories.items() if s in localities][0])
            local_rar.loc[:,'FilingMonth'] = self.config.filing_month.replace(month=1)
            local_rar.loc[:,['Path26Region','Comment']] = ''
            local_rar.loc[:,'RequirementType'] = 'local_rar'
            local_rar.loc[:,'DataSource'] = [attachment_id] * len(local_rar)
            local_rar.loc[:,'Version'] = version
            requirements = requirements.append(local_rar,ignore_index=True)

            # total lcr:
            total_lcr = total_lcr.transpose().reset_index()
            total_lcr.rename({
                'index' : 'Locality',
                0 : 'RequirementValue',
            },axis='columns',inplace=True)
            total_lcr.loc[:,'ServiceTerritory'] = total_lcr.loc[:,'Locality'].map(lambda s:[organization_id for organization_id,localities in regions_to_service_territories.items() if s in localities][0])
            total_lcr.loc[:,'FilingMonth'] = 'NaT'
            total_lcr.loc[:,['LoadServingEntity','Path26Region','Comment']] = ''
            total_lcr.loc[:,'RequirementType'] = 'total_lcr'
            total_lcr.loc[:,'DataSource'] = [attachment_id] * len(total_lcr)
            total_lcr.loc[:,'Version'] = version
            requirements = requirements.append(total_lcr,ignore_index=True)

        archive_path = self.config.paths.get_path('incremental_local')
        if archive_path:
            attachment_id = self.attachment_logger.data.loc[self.attachment_logger.data.loc[:,'archive_path']==str(archive_path),'attachment_id']
            version = self.config.paths.get_version_number(archive_path,'incremental_local')
            wb = open_workbook(archive_path)
            incremental_flex,incremental_local_load,local_rar_trueup = get_incremental_local_tables(wb,self.config)

            # incremental flex:
            incremental_flex.reset_index(inplace=True)
            incremental_flex.rename({
                'organization_id' : 'LoadServingEntity',
                'category' : 'RequirementType',
                'flexibility_requirement' : 'RequirementValue',
            },axis='columns',inplace=True)
            incremental_flex.loc[:,'RequirementType'] = incremental_flex.loc[:,'RequirementType'].map(lambda x:'incremental_flex_category{:.0f}'.format(x))
            incremental_flex.loc[:,['ServiceTerritory','Locality','Path26Region','Comment']] = ''
            incremental_flex.loc[:,'FilingMonth'] = self.config.filing_month.replace(month=7)
            incremental_flex.loc[:,'DataSource'] = [attachment_id] * len(incremental_flex)
            incremental_flex.loc[:,'Version'] = version
            requirements = requirements.append(incremental_flex,ignore_index=True)

            # incremental local load:
            incremental_local_load.reset_index(inplace=True)
            incremental_local_load.rename({
                'organization_id' : 'LoadServingEntity',
                'location' : 'Locality',
                'incremental_load' : 'RequirementValue',
            },axis='columns',inplace=True)
            incremental_local_load.loc[:,'ServiceTerritory'] = incremental_local_load.loc[:,'Locality'].map(lambda s:[organization_id for organization_id,localities in regions_to_service_territories.items() if s in localities][0])
            incremental_local_load.loc[:,'FilingMonth'] = self.config.filing_month.replace(month=7)
            incremental_local_load.loc[:,['LoadServingEntity','Path26Region','Comment']] = ''
            incremental_local_load.loc[:,'RequirementType'] = ['incremental_local_load'] * len(incremental_local_load)
            incremental_local_load.loc[:,'DataSource'] = [attachment_id] * len(incremental_local_load)
            incremental_local_load.loc[:,'Version'] = version
            requirements = requirements.append(incremental_local_load,ignore_index=True)

            # local rar trueup:
            local_rar_trueup.reset_index(inplace=True)
            local_rar_trueup.rename({
                'organization_id' : 'LoadServingEntity',
                'location' : 'Locality',
                'local_rar_trueup' : 'RequirementValue',
            },axis='columns',inplace=True)
            local_rar_trueup = local_rar_trueup.loc[(local_rar_trueup.loc[:,'LoadServingEntity'].map(lambda s:s in [x['id'] for x in self.config.organizations.list_load_serving_entities()])),:]
            local_rar_trueup.loc[:,'FilingMonth'] = self.config.filing_month.replace(month=7)
            local_rar_trueup.loc[:,['Path26Region','Comment']] = ''
            local_rar_trueup.loc[:,'DataSource'] = [attachment_id] * len(local_rar_trueup)
            local_rar_trueup.loc[:,'Version'] = version
            regional_rar_trueup = local_rar_trueup.loc[(local_rar_trueup.loc[:,'Locality'].map(lambda s:s not in [x for l in regions_to_service_territories.values() for x in l])),:]
            regional_rar_trueup.loc[:,'ServiceTerritory'] = ''
            regional_rar_trueup.loc[:,'RequirementType'] = regional_rar_trueup.loc[:,'Locality']
            regional_rar_trueup.loc[:,'Locality'] = ''
            local_rar_trueup = local_rar_trueup.loc[(local_rar_trueup.loc[:,'Locality'].map(lambda s:s in [x for l in regions_to_service_territories.values() for x in l])),:]
            local_rar_trueup.loc[:,'ServiceTerritory'] = local_rar_trueup.loc[:,'Locality'].map(lambda s:[organization_id for organization_id,localities in regions_to_service_territories.items() if s in localities][0])
            local_rar_trueup.loc[:,'RequirementType'] = ['local_rar_trueup'] * len(local_rar_trueup)
            requirements = requirements.append(local_rar_trueup,ignore_index=True)
            requirements = requirements.append(regional_rar_trueup,ignore_index=True)

        archive_path = self.config.paths.get_path('month_ahead')
        if archive_path:
            attachment_id = self.attachment_logger.data.loc[self.attachment_logger.data.loc[:,'archive_path']==str(archive_path),'attachment_id']
            version = self.config.paths.get_version_number(archive_path,'month_ahead')
            wb = open_workbook(archive_path)
            month_ahead_forecasts = get_month_ahead_tables(wb,self.config)
            month_ahead_forecasts.reset_index(inplace=True)
            month_ahead_forecasts.rename({
                'organization_id' : 'LoadServingEntity',
                'month' : 'FilingMonth',
            },axis='columns',inplace=True)
            total_columns = list(filter(lambda s: 'total' in s,month_ahead_forecasts.columns))
            month_ahead_forecasts.drop(columns=['lse_type','jurisdiction','lse_lu','id_and_date']+total_columns,inplace=True)
            month_ahead_forecasts = pd.melt(
                month_ahead_forecasts,
                id_vars=['LoadServingEntity','FilingMonth'],
                var_name='RequirementType',
                value_name='RequirementValue',
                ignore_index=True
            )
            month_ahead_forecasts.loc[:,'ServiceTerritory'] = month_ahead_forecasts.loc[:,'RequirementType'].map(lambda s: s.split('_')[0].upper())
            month_ahead_forecasts.loc[:,'RequirementType'] = month_ahead_forecasts.loc[:,'RequirementType'].map(lambda s: 'month_ahead_'+'_'.join(s.split('_')[1:]))
            month_ahead_forecasts.loc[:,['Locality','Path26Region','Comment']] = ''
            month_ahead_forecasts.loc[:,'DataSource'] = [attachment_id] * len(month_ahead_forecasts)
            month_ahead_forecasts.loc[:,'Version'] = version
            requirements = requirements.append(month_ahead_forecasts,ignore_index=True)

        archive_path = self.config.paths.get_path('cam_rmr')
        if archive_path:
            attachment_id = self.attachment_logger.data.loc[self.attachment_logger.data.loc[:,'archive_path']==str(archive_path),'attachment_id']
            version = self.config.paths.get_version_number(archive_path,'cam_rmr')
            wb = open_workbook(archive_path)
            cam_rmr_monthly_tracking,total_cam_rmr = get_cam_rmr_tables(wb)

            # cam_rmr_monthly_tracking:
            cam_rmr_monthly_tracking.reset_index(inplace=True)
            cam_rmr_monthly_tracking.rename({
                'organization_id' : 'LoadServingEntity',
                'month' : 'FilingMonth',
            },axis='columns',inplace=True)
            total_columns = list(filter(lambda s: 'total' in s,cam_rmr_monthly_tracking.columns))
            cam_rmr_monthly_tracking.drop(columns=['lse_type','jurisdiction','lse_lu','id_and_date']+total_columns,inplace=True)
            cam_rmr_monthly_tracking = pd.melt(
                cam_rmr_monthly_tracking,
                id_vars=['LoadServingEntity','FilingMonth'],
                var_name='RequirementType',
                value_name='RequirementValue',
                ignore_index=True
            )
            cam_rmr_monthly_tracking.loc[:,'ServiceTerritory'] = cam_rmr_monthly_tracking.loc[:,'RequirementType'].map(lambda s: s.split('_')[0].upper())
            cam_rmr_monthly_tracking.loc[:,'RequirementType'] = cam_rmr_monthly_tracking.loc[:,'RequirementType'].map(lambda s: 'cam_rmr_'+'_'.join(s.split('_')[1:]))
            cam_rmr_monthly_tracking.loc[:,['Locality','Path26Region','Comment']] = ''
            cam_rmr_monthly_tracking.loc[:,'DataSource'] = [attachment_id] * len(cam_rmr_monthly_tracking)
            cam_rmr_monthly_tracking.loc[:,'Version'] = version
            requirements = requirements.append(cam_rmr_monthly_tracking,ignore_index=True)

            # total_cam_rmr:
            total_cam_rmr = pd.DataFrame(total_cam_rmr).reset_index()
            total_cam_rmr.rename({
                'index' : 'RequirementType',
                0 : 'RequirementValue',
            },axis='columns',inplace=True)
            total_cam_rmr.loc[:,'RequirementType'] = total_cam_rmr.loc[:,'RequirementType'].map(lambda s: 'total_cam_rmr_'+s)
            total_cam_rmr.loc[:,['LoadServingEntity','ServiceTerritory','Locality','Path26Region','Comment']] = ''
            total_cam_rmr.loc[:,'FilingMonth'] = self.config.filing_month
            total_cam_rmr.loc[:,'DataSource'] = [attachment_id] * len(total_cam_rmr)
            total_cam_rmr.loc[:,'Version'] = version
            requirements = requirements.append(total_cam_rmr,ignore_index=True)

        # reformat date column:
        requirements.loc[:,'FilingMonth'] = requirements.loc[:,'FilingMonth'].map(lambda x: ts('1900-01-01')+td(days=x) if isinstance(x,int) else x)
        requirements.loc[:,'FilingMonth'] = pd.to_datetime(requirements.loc[:,'FilingMonth']).dt.date

        out_path = self.config.paths.get_path('ezdb_requirements')
        out_path.parent.mkdir(parents=True,exist_ok=True)
        self.logger.log('Exporting Resource Adequacy Requirements for EZDB to {}'.format(out_path.name),'INFORMATION')
        requirements.loc[:,columns].to_csv(out_path,index=False)

    def write_resources(self):
        columns = [
            'LoadServingEntity',
            'ContractID',
            'ResourceID',
            'Operator',
            'Locality',
            'ServiceTerritory',
            'Path26Region',
            'MCCBucket',
            'ContinuousAvailability',
            'Start',
            'End',
            'CapacityType',
            'CapacityValue',
            'DataSource',
            'Version',
            'Comment',
        ]
        resources = pd.DataFrame(columns=columns)

        monthly_filings = self.consolidation_logger.data.loc[
            (self.consolidation_logger.data.loc[:,'ra_category']=='ra_monthly_filing') & \
            (
                (self.consolidation_logger.data.loc[:,'status']=='Ready') | \
                (self.consolidation_logger.data.loc[:,'status']=='Late')
            ),:
        ]
        for _,monthly_filing in monthly_filings.iterrows():
            archive_path = Path(monthly_filing.loc['archive_path'])
            if archive_path.is_file():
                attachment_id = monthly_filing.loc['attachment_id']
                organization = self.config.organizations.get_organization(monthly_filing.loc['organization_id'])
                summary,physical_resources,demand_response = read_ra_monthly_filing(organization,self.config,self.logger)
                version = self.config.paths.get_version_number(archive_path,'ra_monthly_filing',organization)

                # physical resources:
                if len(physical_resources)>0:
                    physical_resources.reset_index(inplace=True)
                    physical_resources.rename({
                        'organization_id' : 'LoadServingEntity',
                        'contract_id' : 'ContractID',
                        'resource_id' : 'ResourceID',
                        'resource_adequacy_system' : 'System',
                        'resource_adequacy_local' : 'Local',
                        'resource_mcc_bucket' : 'MCCBucket',
                        'continuous_availability' : 'ContinuousAvailability',
                        'resource_adequacy_committed_flexible' : 'Flexible',
                        'resource_adequacy_flexibility_category' : 'FlexibleCategory',
                        'start_date' : 'Start',
                        'end_date' : 'End',
                        'scid' : 'Operator',
                        'zone' : 'Path26Region',
                    },axis='columns',inplace=True)
                    physical_resources.loc[:,'Locality'] = physical_resources.loc[:,'local_area'].map(location_renamer)
                    physical_resources.drop(columns=['index','local_area'],inplace=True)
                    physical_resources.loc[:,'ServiceTerritory'] = physical_resources.loc[:,'Locality'].map(lambda s:[organization_id for organization_id,localities in regions_to_service_territories.items() if s in localities][0])
                    physical_resources = pd.melt(
                        physical_resources,
                        id_vars=['LoadServingEntity','ContractID','ResourceID','MCCBucket','ContinuousAvailability','FlexibleCategory','Start','End','Operator','Path26Region','Locality','ServiceTerritory'],
                        var_name='CapacityType',
                        value_name='CapacityValue',
                        ignore_index=True
                    )
                    physical_resources.dropna(axis='index',subset=['CapacityValue'],inplace=True)
                    physical_resources.loc[:,'FlexibleCategory'] = physical_resources.loc[:,'FlexibleCategory'].map(lambda x: None if x=='' else x).fillna(0).astype(int)
                    physical_resources.loc[:,'CapacityType'] = physical_resources.apply(lambda r: 'Flexible Category {}'.format(str(r.loc['FlexibleCategory'])) if r.loc['CapacityType']=='Flexible' else r.loc['CapacityType'],axis='columns')
                    physical_resources.loc[:,'DataSource'] = attachment_id
                    physical_resources.loc[:,'Version'] = version
                    physical_resources.loc[:,'Comment'] = ''
                    resources = resources.append(physical_resources,ignore_index=True)
                else:
                    pass

                # demand response resources:
                if len(demand_response)>0:
                    demand_response.reset_index(inplace=True)
                    demand_response.rename({
                        'organization_id' : 'LoadServingEntity',
                        'contract_id' : 'ContractID',
                        'program_id' : 'ResourceID',
                        'resource_adequacy_system' : 'System',
                        'resource_adequacy_local' : 'Local',
                        'resource_mcc_bucket' : 'MCCBucket',
                        'resource_adequacy_committed_flexible' : 'Flexible',
                        'resource_adequacy_flexibility_category' : 'FlexibleCategory',
                        'start_date' : 'Start',
                        'end_date' : 'End',
                        'operator' : 'Operator',
                        'zone' : 'Path26Region',
                    },axis='columns',inplace=True)
                    demand_response.loc[:,'Locality'] = demand_response.loc[:,'local_area'].map(rename_locality)
                    demand_response.drop(columns=['index','third_party_program','local_area'],inplace=True)
                    demand_response.loc[:,'ServiceTerritory'] = demand_response.loc[:,'Locality'].map(lambda s:[organization_id for organization_id,localities in regions_to_service_territories.items() if s in localities][0])
                    demand_response = pd.melt(
                        demand_response,
                        id_vars=['LoadServingEntity','ContractID','ResourceID','MCCBucket','FlexibleCategory','Start','End','Locality','Operator','Path26Region','ServiceTerritory'],
                        var_name='CapacityType',
                        value_name='CapacityValue',
                        ignore_index=True
                    )
                    demand_response.dropna(axis='index',subset=['CapacityValue'],inplace=True)
                    demand_response.loc[:,'FlexibleCategory'] = demand_response.loc[:,'FlexibleCategory'].map(lambda x: None if x=='' else x).fillna(0).astype(int)
                    demand_response.loc[:,'CapacityType'] = demand_response.apply(lambda r: 'Flexible Category {}'.format(str(r.loc['FlexibleCategory'])) if r.loc['CapacityType']=='Flexible' else r.loc['CapacityType'],axis='columns')
                    demand_response.loc[:,'ContinuousAvailability'] = False
                    demand_response.loc[:,'DataSource'] = [attachment_id] * len(demand_response)
                    demand_response.loc[:,'Version'] = version
                    demand_response.loc[:,'Comment'] = ''
                    resources = resources.append(demand_response,ignore_index=True)
                else:
                    pass
            else:
                pass
        # reformat date columns:
        resources.loc[:,'Start'] = pd.to_datetime(resources.loc[:,'Start']).dt.date
        resources.loc[:,'End'] = pd.to_datetime(resources.loc[:,'End']).dt.date

        out_path = self.config.paths.get_path('ezdb_resources')
        out_path.parent.mkdir(parents=True,exist_ok=True)
        self.logger.log('Exporting Physical and DR Resources for EZDB to {}'.format(out_path.name),'INFORMATION')
        resources.loc[:,columns].to_csv(out_path,index=False,sep=',',quotechar='"')

    def write_supply_plans(self):
        columns = [
            'LoadServingEntity',
            'Operator',
            'ResourceID',
            'Locality',
            'Start',
            'End',
            'CapacityType',
            'CapacityValue',
            'ErrorsAndWarnings',
            'DataSource',
            'Version',
            'Comment',
        ]
        supply_plans = pd.DataFrame(columns=columns)

        supply_plan_system_information = self.consolidation_logger.data.loc[
            (self.consolidation_logger.data.loc[:,'ra_category']=='supply_plan_system'),
            :
        ].iloc[0]
        version = self.config.paths.get_version_number(supply_plan_system_information.loc['archive_path'],'supply_plan_system')

        def scid_to_organization_id(scid:str):
            organization_id = self.config.organizations.lookup_id(scid)
            if not organization_id:
                organization_id=scid
            else:
                pass
            return organization_id

        ra_summary = open_workbook(self.config.paths.get_path('ra_summary'))
        nqc_list = get_nqc_list(ra_summary,self.config)

        # system and local capacity supply plan:
        supply_plan_system = read_supply_plan(
            self.config,
            'supply_plan_system',
            supply_plan_system_information.loc['effective_date'],
            version
        )
        supply_plan_system = supply_plan_system.merge(nqc_list.loc[:,['resource_id','local_area']],on='resource_id')
        supply_plan_system.rename({
            'supplier' : 'Operator',
            'resource_id' : 'ResourceID',
            'local_resource_adequacy' : 'local',
            'system_resource_adequacy' : 'system',
            'start_date' : 'Start',
            'end_date' : 'End',
            'organization_id_caiso' : 'LoadServingEntity',
            'errors_and_warnings' : 'ErrorsAndWarnings',
            'local_area' : 'Locality',
        },axis='columns',inplace=True)
        supply_plan_system.drop(columns=['validation_status','total_capacity'],inplace=True)
        supply_plan_system.loc[:,'Operator'] = supply_plan_system.loc[:,'Operator'].map(scid_to_organization_id)
        supply_plan_system.loc[:,'LoadServingEntity'] = supply_plan_system.loc[:,'LoadServingEntity'].map(scid_to_organization_id)
        supply_plan_system = pd.melt(
            supply_plan_system,
            id_vars=['Operator','ResourceID','Start','End','LoadServingEntity','ErrorsAndWarnings','Locality'],
            var_name='CapacityType',
            value_name='CapacityValue',
            ignore_index=True
        )
        supply_plan_system.loc[:,'DataSource'] = supply_plan_system_information.loc['attachment_id']
        supply_plan_system.loc[:,'Version'] = version
        supply_plan_system.loc[:,'Comment'] = ''
        supply_plans = supply_plans.append(supply_plan_system.loc[:,columns],ignore_index=True)

        # flexible capacity supply plan:
        supply_plan_flexible_information = self.consolidation_logger.data.loc[
            (self.consolidation_logger.data.loc[:,'ra_category']=='supply_plan_flexible'),
            :
        ].iloc[0]
        version = self.config.paths.get_version_number(supply_plan_flexible_information.loc['archive_path'],'supply_plan_flexible')

        supply_plan_flexible = read_supply_plan(
            self.config,
            'supply_plan_flexible',
            supply_plan_flexible_information.loc['effective_date'],
            version
        )
        supply_plan_flexible = supply_plan_flexible.merge(nqc_list.loc[:,['resource_id','local_area']],on='resource_id')
        supply_plan_flexible.rename({
            'supplier' : 'Operator',
            'resource_id' : 'ResourceID',
            'category' : 'CapacityType',
            'flex_capacity' : 'CapacityValue',
            'start_date' : 'Start',
            'end_date' : 'End',
            'organization_id_caiso' : 'LoadServingEntity',
            'errors_and_warnings' : 'ErrorsAndWarnings',
            'local_area' : 'Locality',
        },axis='columns',inplace=True)
        supply_plan_flexible.drop(columns=['validation_status'],inplace=True)
        supply_plan_flexible.loc[:,'Operator'] = supply_plan_flexible.loc[:,'Operator'].map(scid_to_organization_id)
        supply_plan_flexible.loc[:,'LoadServingEntity'] = supply_plan_flexible.loc[:,'LoadServingEntity'].map(scid_to_organization_id)
        supply_plan_flexible.loc[:,'CapacityType'] = supply_plan_flexible.loc[:,'CapacityType'].map(lambda c: 'flexible_category_{}'.format(c))
        supply_plan_flexible.loc[:,'DataSource'] = supply_plan_flexible_information.loc['attachment_id']
        supply_plan_flexible.loc[:,'Version'] = version
        supply_plan_flexible.loc[:,'Comment'] = ''
        supply_plans = supply_plans.append(supply_plan_flexible.loc[:,columns],ignore_index=True)

        # reformat date columns:
        supply_plans.loc[:,'Start'] = pd.to_datetime(supply_plans.loc[:,'Start']).dt.date
        supply_plans.loc[:,'End'] = pd.to_datetime(supply_plans.loc[:,'End']).dt.date

        # save data to file:
        out_path = self.config.paths.get_path('ezdb_supply_plans')
        out_path.parent.mkdir(parents=True,exist_ok=True)
        self.logger.log('Exporting Supply Plans for EZDB to {}'.format(out_path.name),'INFORMATION')
        supply_plans.to_csv(out_path,index=False,sep=',',quotechar='"')

    def write_summaries(self):
        columns = [
            'LoadServingEntity',
            'FilingMonth',
            'Parameter',
            'Value',
            'DataSource',
            'Comment',
        ]
        summaries = pd.DataFrame(columns=columns)
        archive_path = self.config.paths.get_path('caiso_cross_check')
        attachment_id = self.attachment_logger.data.loc[(self.attachment_logger.data.loc[:,'archive_path']==str(archive_path)),'attachment_id']
        caiso_cross_check = load_workbook(archive_path)
        system_requirements,flexibility_requirements = get_cross_check_tables(caiso_cross_check,self.config)

        system_requirements.reset_index(inplace=True)
        system_requirements.rename({
            'organization_id' : 'LoadServingEntity',
        },axis='columns',inplace=True)
        system_requirements.drop(columns=[
            'percent_required_resources_available',
            'total_resources_available'
        ],inplace=True)
        system_requirements = pd.melt(
            system_requirements,
            id_vars=['LoadServingEntity'],
            var_name='Parameter',
            value_name='Value',
            ignore_index=True
        )
        system_requirements.loc[:,'FilingMonth'] = self.config.filing_month
        system_requirements.loc[:,'DataSource'] = [attachment_id] * len(system_requirements)
        system_requirements.loc[:,'Comment'] = [''] * len(system_requirements)
        summaries = summaries.append(system_requirements,ignore_index=True)

        flexibility_requirements.reset_index(inplace=True)
        flexibility_requirements.rename({
            'organization_id' : 'LoadServingEntity',
            'flexibility_category_1' : 'flexibility_category_1_required',
            'flexibility_category_1_countable' : 'flexibility_category_1_available',
            'flexibility_category_2' : 'flexibility_category_2_required',
            'flexibility_category_2_countable' : 'flexibility_category_2_available',
            'flexibility_category_3' : 'flexibility_category_3_required',
            'flexibility_category_3_countable' : 'flexibility_category_3_available',
        },axis='columns',inplace=True)
        flexibility_requirements.drop(columns=[
            'flexibility_requirements',
            'flexibility_available',
            'percent_flexibility_requirements_available',
            'year_ahead_flexibility_category_1',
            'year_ahead_flexibility_category_2',
            'year_ahead_flexibility_category_3',
            'year_ahead_flex_total'
        ],inplace=True)
        flexibility_requirements = pd.melt(
            flexibility_requirements,
            id_vars=['LoadServingEntity'],
            var_name='Parameter',
            value_name='Value',
            ignore_index=True
        )
        flexibility_requirements.loc[:,'FilingMonth'] = self.config.filing_month
        flexibility_requirements.loc[:,'DataSource'] = [attachment_id] * len(flexibility_requirements)
        flexibility_requirements.loc[:,'Comment'] = [''] * len(flexibility_requirements)
        summaries = summaries.append(flexibility_requirements,ignore_index=True)

        # reformat date column:
        summaries.loc[:,'FilingMonth'] = pd.to_datetime(summaries.loc[:,'FilingMonth']).dt.date

        # save data to file:
        out_path = self.config.paths.get_path('ezdb_summaries')
        self.logger.log('Exporting Summaries for EZDB to {}'.format(out_path.name),'INFORMATION')
        out_path.parent.mkdir(parents=True,exist_ok=True)
        summaries.to_csv(out_path,index=False,sep=',',quotechar='"')

    def write_all(self):
        '''
        calls each of the write table methods in sequence, exporting all tables.
        '''
        self.write_data_sources()
        self.write_organizations()
        self.write_requirements()
        self.write_resources()
        self.write_supply_plans()
        self.write_summaries()

    def update_master_lookup_table(self):
        master_lookup_path = self.config.paths.get_path('ezdb_master_lookup')
        wb = open_workbook(master_lookup_path,data_only=True,read_only=False)
        columns = [cell.value for cell in wb['TablesMaster']['A1:L1'][0]]
        data_range = wb['TablesMaster']['A2:L{}'.format(wb['TablesMaster'].max_row)]
        tables_master = data_range_to_dataframe(columns,data_range)
        row_number = wb['TablesMaster'].max_row
        for table_name in ['organizations','data_sources','requirements','resources','supply_plans','summaries']:
            filename = self.config.paths.get_path('ezdb_'+table_name)
            if filename not in list(tables_master.loc[:,'FileName']):
                row_number += 1
                wb['TablesMaster']['A{}'.format(row_number)].value = row_number - 1
                wb['TablesMaster']['B{}'.format(row_number)].value = table_name
                wb['TablesMaster']['C{}'.format(row_number)].value = table_name
                wb['TablesMaster']['D{}'.format(row_number)].value = self.config.filing_month.strftime('%YM%m')
                wb['TablesMaster']['E{}'.format(row_number)].value = '{} Filings'.format(self.config.filing_month.strftime('%B %Y'))
                wb['TablesMaster']['G{}'.format(row_number)].value = 0
                wb['TablesMaster']['G{}'.format(row_number)].alignment = Alignment(horizontal='center')
                wb['TablesMaster']['I{}'.format(row_number)].value = filename.name
        self.logger.log('Writing Updated Master Lookup Table to {}'.format(str(master_lookup_path)),'INFORMATION')
        wb.save(str(master_lookup_path))
        wb.close()

if __name__=='__main__':
    de = DataExporter(Path('M:/Users/svc_energyRA/ra_filings/config/ra_filings_config.yaml'))
    de.write_all()
    de.update_master_lookup_table()