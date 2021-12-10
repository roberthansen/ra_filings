import io
import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter,column_index_from_string
from pathlib import Path
from yaml import load, safe_load
from datetime import datetime as dt
from logger import logger

class ra_consolidator:
    def __init__(self,configuration_path: Path):
        self.configuration_options = {
            'filing_month' : None,
            'lse_map_file' : None,
            'report_directory' : None,
            'ra_monthly_filing_filename_template' : None,
            'incremental_local_filename_template' : None,
            'cam_rmr_filename_template' : None,
            'ra_summary_filename_template' : None,
            'month_ahead_filename_template' : None,
            'year_ahead_filename_template' : None,
            'log_file' : Path.cwd() / 'download_organizer.log',
            'cli_logging_criticalities' : ['INFORMATION','WARNING','ERROR',],
            'file_logging_criticalities' : ['INFORMATION','WARNING','ERROR',],
        }
        self.logger = logger()
        self.lse_list = list()
        self.lse_map = dict()
        self.paths = dict()
        self.set_configuration_options(configuration_path)
        
    # get list of load serving entities:
    def set_lse_list(self,p: Path):
        if p.is_file():
            self.lse_map_file = p
            with self.lse_map_file.open(mode='r') as f:
                d = safe_load(f)
            # flip keys and list values
            self.lse_list = d.keys()
            self.logger.log('Loaded LSE List from {}'.format(p),'INFORMATION')
        else:
            self.lse_map_file = None
            self.logger.log('No LSE Mapping File Found at {}'.format(p),'ERROR')
    
    # read file with mapping from load serving entities' full names and
    # variations to abbreviations:
    def set_lse_map(self,p: Path):
        if p.is_file():
            self.lse_map_file = p
            with self.lse_map_file.open(mode='r') as f:
                self.lse_map = safe_load(f)
            # flip keys and list values
            # self.lse_map = dict()
            # for key in d.keys():
            #     for value in d[key]:
            #         self.lse_map[value] = key
            self.logger.log('Loaded LSE Map from {}'.format(p),'INFORMATION')
        else:
            self.lse_map_file = None
            self.logger.log('No LSE Mapping File Found at {}'.format(p),'ERROR')

    # set the paths of each file:
    def set_paths(self):
        # passing through configuration options until pattern matching is implemented:
        self.paths = {
            'ra_monthly_filing' : lambda lse_abbrev: self.parse_filename_template(self.configuration_options['ra_monthly_filing_filename_template'],lse_abbrev),
            'incremental_local' : self.parse_filename_template(self.configuration_options['incremental_local_filename_template']),
            'cam_rmr' : self.parse_filename_template(self.configuration_options['cam_rmr_filename_template']),
            'ra_summary' : self.parse_filename_template(self.configuration_options['ra_summary_filename_template']),
            'month_ahead' : self.parse_filename_template(self.configuration_options['month_ahead_filename_template']),
            'year_ahead' : self.parse_filename_template(self.configuration_options['year_ahead_filename_template']),
        }

    # read configuration file and apply relevant options:
    def set_configuration_options(self,p: Path):
        if p.is_file():
            self.configuration_path = p
            with self.configuration_path.open(mode='r') as f:
                d = safe_load(f)
                for key in d.keys():
                    if key in self.configuration_options.keys():
                        value = d[key]
                        if 'criticalities' in key:
                            self.configuration_options[key] = value.split(',')
                        elif key=='filing_month':
                            self.configuration_options[key] = dt.strptime(value,'%b %Y')
                        elif ('file' in key or 'directory' in key) and 'template' not in key:
                            self.configuration_options[key] = Path(value)
                        else:
                            self.configuration_options[key] = value
            self.logger.log('Applying Configuration Options from {}'.format(p),'INFORMATION')
        else:
            self.configuration_path = None
            self.logger.log('Applying Default Configuration---Unable to Load Options from {}'.format(p),'WARNING')
        self.logger.set_log_path(self.configuration_options['log_file'])
        self.logger.set_cli_logging_criticalities(self.configuration_options['cli_logging_criticalities'])
        self.logger.set_file_logging_criticalities(self.configuration_options['file_logging_criticalities'])
        self.set_lse_list(Path(self.configuration_options['lse_map_file']))
        self.set_lse_map(Path(self.configuration_options['lse_map_file']))
        self.set_paths()

    def clear_data_ranges(self):
        # open ra_summary file into memory (read-only, will overwrite later):
        ra_summary = self.open_workbook('ra_summary',data_only=False)
        
        # specify data ranges in summary sheet to be cleared:
        data_ranges = [
            self.get_data_range(ra_summary['NP26'],'B','C'),
            self.get_data_range(ra_summary['NP26'],'B','H'),
            self.get_data_range(ra_summary['SP26'],'B','C'),
            self.get_data_range(ra_summary['SP26'],'B','H'),
            self.get_data_range(ra_summary['FlexRAR'],'B','C'),
            self.get_data_range(ra_summary['FlexRAR'],'B','F'),
            self.get_data_range(ra_summary['FlexRAR'],'B','H'),
            self.get_data_range(ra_summary['FlexRAR'],'B','J'),
            ra_summary['PhysRes']['A2:AA{}'.format(ra_summary['PhysRes'].max_row)],
            self.get_data_range(ra_summary['CertifyingOfficers'],'B','F'),
            self.get_data_range(ra_summary['CertifyingOfficers'],'B','H'),
        ]

        # clear each data range:
        for data_range in data_ranges:
            for data_range_row in data_range:
                for cell in data_range_row:
                    sheet_name = cell.parent.title
                    cell_coordinate = cell.coordinate
                    cell_value = cell.value
                    #self.logger.log('In Sheet {}, Clearing Cell {} [Original Value: {}]'.format(sheet_name,cell_coordinate,cell_value),'INFORMATION')
                    cell.value = None

        # save and close summary file:
        ra_summary.save(str(self.paths['ra_summary']))
        ra_summary.close()
        self.logger.log('Cleared Data from {}'.format(self.paths['ra_summary'].name),'INFORMATION')

    # open each applicable file and copy data into summary sheet:
    def consolidate_allocations(self):
        # start timer:
        init_time = dt.now()

        # get source data from year ahead file:
        year_ahead = self.open_workbook('year_ahead')
        columns = [
            'iou_territory',
            'month',
            'lse_abbrev',
            'lse_type',
            'submitted_forecast',
            'coincidence_adjustment',
            'coincident_peak_forecast',
            'lse_specific_total',
            'copkadj_with_lseadj',
            'eelmdr_adjustment',
            'adjusted_with_lmdr',
            'pro_rata_adjustment',
            'final_coincident_peak_forecast',
        ]
        data_range = year_ahead['loadforecastinputdata']['B2:N{}'.format(year_ahead['loadforecastinputdata'].max_row)]
        load_forecast_input_data = self.data_range_to_dataframe(columns,data_range)
        load_forecast_input_data.dropna('index',inplace=True)
        load_forecast_input_data['month'] = load_forecast_input_data.loc[:,'month'].map(lambda s: dt(self.configuration_options['filing_month'].year,int(s),1))
        load_forecast_input_data.set_index(['iou_territory','lse_abbrev','month'],inplace=True)
        load_forecast_input_data.sort_index(inplace=True)
        # accommodate errant entries:
        load_forecast_input_data.loc[:,columns[4:]] = load_forecast_input_data.loc[:,columns[4:]].applymap(lambda x: 0 if isinstance(x,str) else x)

        month_columns = [dt(self.configuration_options['filing_month'].year,month,1) for month in range(1,13)]
        columns = [
            'location',
        ] + month_columns
        data_range = year_ahead['DRforAllocation']['D2:P32']
        demand_response_allocation = self.data_range_to_dataframe(columns,data_range)
        demand_response_allocation['location'] = demand_response_allocation.loc[:,'location'].map(lambda x: str(x).lower().replace(' ','_'))
        demand_response_allocation = demand_response_allocation.melt(id_vars=['location'],var_name='month',value_name='allocation')
        demand_response_allocation.set_index(['location','month'],inplace=True)
        demand_response_allocation.sort_index(inplace=True)

        # cam credits table:
        columns = [
            'iou_territory',
            'category',
        ] + month_columns
        data_range = year_ahead['Flexrequirements']['R4:AE12']
        cam_credits = self.data_range_to_dataframe(columns,data_range)
        cam_credits = cam_credits.melt(id_vars=['iou_territory','category'],var_name='month',value_name='cam_credit')
        cam_credits.set_index(['iou_territory','category','month'],inplace=True)
        cam_credits.sort_index(inplace=True)

        # cpuc flexibility requirements:
        table_header_text = 'Flex Requirements net CAM'
        table_header_offset = {
            'rows' : 1,
            'columns' : -1,
        }
        columns = [
            'lse_abbrev',
            'flex_category',
        ] + month_columns
        cpuc_flexibility_requirements = self.get_table(year_ahead['Flexrequirements'],table_header_text,table_header_offset,columns)
        cpuc_flexibility_requirements['flex_category'] = cpuc_flexibility_requirements.loc[:,'flex_category'].map(lambda s: int(s[-1]))
        cpuc_flexibility_requirements = cpuc_flexibility_requirements.melt(id_vars=['lse_abbrev','flex_category'],var_name='month',value_name='flexibility_requirement')
        cpuc_flexibility_requirements.set_index(['lse_abbrev','flex_category','month'],inplace=True)
        cpuc_flexibility_requirements.sort_index(inplace=True)


        # total lcr table:
        data_range = year_ahead['Local RA-CAM-{}'.format(self.configuration_options['filing_month'].year)]['C2:L2']
        columns = [cell.value for cell in data_range[0]]
        data_range = year_ahead['Local RA-CAM-{}'.format(self.configuration_options['filing_month'].year)]['C4:L4']
        total_lcr = self.data_range_to_dataframe(columns,data_range)

        year_ahead.close()

        # get source data from month ahead file:
        month_ahead = self.open_workbook('month_ahead')
        columns = [
            'lse_abbrev',
            'lse_type',
            'jurisdiction',
            'lse_lu',
            'month',
            'id_and_date',
            'sce_year_ahead_forecast',
            'sdge_year_ahead_forecast',
            'pge_year_ahead_forecast',
            'total_year_ahead_forecast',
            'sce_esps_migrating_load',
            'sce_cca_migrating_load',
            'sdge_esps_migrating_load',
            'sdge_cca_migrating_load',
            'pge_esps_migrating_load',
            'pge_cca_migrating_load',
            'sce_migration_adjustment',
            'sdge_migration_adjustment',
            'pge_migration_adjustment',
            'total_migration_adjustment',
            'sce_revised_monthly_forecast',
            'sdge_revised_monthly_forecast',
            'pge_revised_monthly_forecast',
            'total_revised_monthly_forecast',
            'sce_revised_noncoincident_monthly_forecast',
            'sdge_revised_noncoincident_monthly_forecast',
            'pge_revised_noncoincident_monthly_forecast',
            'total_revised_noncoincident_monthly_forecast',
            'sce_revised_nonjurisdictional_load_share',
            'sdge_revised_nonjurisdictional_load_share',
            'pge_revised_nonjurisdictional_load_share',
            'total_revised_nonjurisdictional_load_share',
            'sce_revised_jurisdictional_load_share',
            'sdge_revised_jurisdictional_load_share',
            'pge_revised_jurisdictional_load_share',
            'total_revised_jurisdictional_load_share',
        ]
        data_range = month_ahead['Monthly Tracking for CPUC']['B5:AK{}'.format(month_ahead['Monthly Tracking for CPUC'].max_row)]
        month_ahead_forecasts = self.data_range_to_dataframe(columns,data_range)
        month_ahead_forecasts.set_index(['lse_abbrev','month'],inplace=True)
        month_ahead_forecasts.sort_index(inplace=True)
        month_ahead.close()

        # get cam-rmr data tables:
        cam_rmr = self.open_workbook('cam_rmr')
        columns = [
            'lse_abbrev',
            'lse_type',
            'jurisdiction',
            'lse_lu',
            'month',
            'id_and_date',
            'sce_year_ahead_forecast',
            'sdge_year_ahead_forecast',
            'pge_year_ahead_forecast',
            'total_year_ahead_forecast',
            'sce_esps_migrating_load',
            'sce_cca_migrating_load',
            'sdge_esps_migrating_load',
            'sdge_cca_migrating_load',
            'pge_esps_migrating_load',
            'pge_cca_migrating_load',
            'sce_load_migration_adjustment',
            'sdge_load_migration_adjustment',
            'pge_load_migration_adjustment',
            'total_load_migration_adjustment',
            'sce_revised_monthly_forecast',
            'sdge_revised_monthly_forecast',
            'pge_revised_monthly_forecast',
            'total_revised_monthly_forecast',
            'sce_revised_noncoincident_monthly_forecast',
            'sdge_revised_noncoincident_monthly_forecast',
            'pge_revised_noncoincident_monthly_forecast',
            'total_revised_noncoincident_monthly_forecast',
            'sce_revised_nonjurisdictional_load_share',
            'sdge_revised_nonjurisdictional_load_share',
            'pge_revised_nonjurisdictional_load_share',
            'total_revised_nonjurisdictional_load_share',
            'blank',
            'sce_revised_jurisdictional_load_share',
            'sdge_revised_jurisdictional_load_share',
            'pge_revised_jurisdictional_load_share',
            'total_revised_jurisdictional_load_share',
        ]
        data_range = cam_rmr['monthlytracking']['B5:AL{}'.format(cam_rmr['monthlytracking'].max_row)]
        cam_rmr_monthly_tracking = self.data_range_to_dataframe(columns,data_range)
        cam_rmr_monthly_tracking['lse_abbrev'] = cam_rmr_monthly_tracking.loc[:,'lse_abbrev'].map(lambda s: s.replace('Total','').strip())
        cam_rmr_monthly_tracking.set_index(['lse_abbrev','month'],inplace=True)
        cam_rmr_monthly_tracking.sort_index(inplace=True)

        # total cam-rmr table:
        columns = [
            'np26_cam',
            'sp26_cam',
            'np26_rmr',
            'sp26_rmr',
            'system_rmr',
            'sce_preferred_lcr_credit',
            'sdge_cam',
            'sce_cam',
        ]
        data_range = cam_rmr['CAMRMR']['B4:I4']
        total_cam_rmr = pd.Series(data=[cell.value for cell in data_range[0]],index=columns)
        cam_rmr.close()

        # incremental local flex table:
        incremental_local = self.open_workbook('incremental_local')
        columns = [
            'lse_abbrev',
            1,
            2,
            3,
        ]
        data_range = self.get_data_range(incremental_local['IncrementalLocal'],'A','MNO')
        incremental_flex = self.data_range_to_dataframe(columns,data_range)
        incremental_flex = incremental_flex.melt(id_vars=['lse_abbrev'],var_name='category',value_name='flexibility_requirement')
        incremental_flex.set_index(['lse_abbrev','category'],inplace=True)
        incremental_flex.sort_index(inplace=True)

        # incremental local_load table:
        columns = [
            'lse_abbrev',
            'los_angeles',
            'ventura',
            'san_diego',
            'bay_area',
            'fresno',
            'sierra',
            'stockton',
            'kern',
            'humboldt',
            'northern_california',
        ]
        data_range = self.get_data_range(incremental_local['IncrementalLocal'],'A','BCDEFGHIJK')
        incremental_local_load = self.data_range_to_dataframe(columns,data_range)
        incremental_local_load = incremental_local_load.melt(id_vars=['lse_abbrev'],var_name='location',value_name='incremental_load')
        incremental_local_load.set_index(['lse_abbrev','location'],inplace=True)
        incremental_local_load.sort_index()

        # incremental local rar true-up:
        columns = [
            'lse_abbrev',
            'los_angeles',
            'ventura',
            'san_diego',
            'bay_area',
            'fresno',
            'sierra',
            'stockton',
            'kern',
            'humboldt',
            'northern_california',
            'sp26_cam_capacity',
            'sp26_condition2_rmr',
            'np26_condition2_rmr',
            'path26_ns',
            'path26_sn',
            'np26_cam_capacity',
        ]
        local_rar_trueup = self.get_table(incremental_local['Local Trueup'],'YA Local RAR Allocations',{'rows':6,'columns':1},columns)
        local_rar_trueup = local_rar_trueup.melt(id_vars=['lse_abbrev'],var_name='location',value_name='local_rar_trueup')
        local_rar_trueup.set_index(['lse_abbrev','location'],inplace=True)
        local_rar_trueup.sort_index(inplace=True)
        incremental_local.close()

        # initialize summary table:
        ra_summary = self.open_workbook('ra_summary',data_only=False)
        columns = [
            'lse_abbrev',
        ]
        data_range = self.get_data_range(ra_summary['Summary'],'B','')
        summary = self.data_range_to_dataframe(columns,data_range)
        def calculate_summary(row):
            lse_abbrev = row.loc['lse_abbrev']
            month = self.configuration_options['filing_month']
            # NP26 summary sheet:
            try:
                np26_ra_obligation = np.round(
                    (
                        # PGEload
                        1.15 * month_ahead_forecasts.loc[(lse_abbrev,month),'pge_revised_monthly_forecast']
                        # NP26CAM
                        -total_cam_rmr.loc['np26_cam'] * cam_rmr_monthly_tracking.loc[(lse_abbrev,month),'pge_revised_nonjurisdictional_load_share']
                        # NP26RMR
                        -total_cam_rmr.loc['np26_rmr'] * cam_rmr_monthly_tracking.loc[(lse_abbrev,month),'pge_revised_nonjurisdictional_load_share']
                    ),
                    0
                ) - total_cam_rmr.loc['system_rmr'] * cam_rmr_monthly_tracking.loc[(lse_abbrev,month),'total_revised_jurisdictional_load_share']
            except:
                np26_ra_obligation = 0
            sn_path26_allocation = 0
            # SP26 summary sheet:
            try:
                sp26_ra_obligation = np.round(
                    (
                        1.15 * (
                            month_ahead_forecasts.loc[(lse_abbrev,month),'sce_revised_monthly_forecast']
                            +month_ahead_forecasts.loc[(lse_abbrev,month),'sdge_revised_monthly_forecast']
                        )
                        -total_cam_rmr.loc['sp26_cam'] * cam_rmr_monthly_tracking.loc[(lse_abbrev,month),'sce_revised_nonjurisdictional_load_share']
                        -total_cam_rmr.loc['sp26_cam'] * cam_rmr_monthly_tracking.loc[(lse_abbrev,month),'sdge_revised_nonjurisdictional_load_share']
                        -total_cam_rmr.loc['sp26_rmr'] * cam_rmr_monthly_tracking.loc[(lse_abbrev,month),'sce_revised_nonjurisdictional_load_share']
                        -total_cam_rmr.loc['sp26_rmr'] * cam_rmr_monthly_tracking.loc[(lse_abbrev,month),'sdge_revised_nonjurisdictional_load_share']
                        -total_cam_rmr.loc['sce_preferred_lcr_credit'] * cam_rmr_monthly_tracking.loc[(lse_abbrev,month),'sce_revised_jurisdictional_load_share']
                    ),
                    0
                )
            except:
                sp26_ra_obligation = 0
            ns_path26_allocation = 0
            def incremental_flex_by_category(lse_abbrev: str,category: int):
                if self.configuration_options['filing_month'].month>6:
                    flex = incremental_flex.loc[(lse_abbrev,category),'flexibility_requirement']
                else:
                    flex = 0
                return flex
            def august_demand_response(iou_territory: str,location: str,lse_abbrev: str):
                if self.configuration_options['filing_month'].month>6:
                    month = dt(self.configuration_options['filing_month'].year,8,1)
                    if (iou_territory,lse_abbrev,month) in load_forecast_input_data.index and (location,month) in demand_response_allocation.index:
                        august_forecast_lse = load_forecast_input_data.loc[(iou_territory,lse_abbrev,month),'final_coincident_peak_forecast']
                        lse_abbrev_indices = list(dict.fromkeys(load_forecast_input_data.loc[(iou_territory),:].index.get_level_values(0)))
                        august_forecast_local = load_forecast_input_data.loc[(iou_territory,lse_abbrev_indices,month),'final_coincident_peak_forecast'].sum()
                        august_demand_response_allocation = demand_response_allocation.loc[(location,month),'allocation'].sum()
                        if august_forecast_local>0:
                            demand_response = august_forecast_lse / august_forecast_local * august_demand_response_allocation
                        else:
                            demand_response = 0
                    else:
                        demand_response = 0
                else:
                    demand_response = 0
                return demand_response

            return pd.Series({
                'lse_abbrev' : lse_abbrev,
                'np26_ra_obligation' : np26_ra_obligation,
                'sn_path26_allocation' : sn_path26_allocation,
                'sp26_ra_obligation' : sp26_ra_obligation,
                'ns_path26_allocation' : ns_path26_allocation,
                'year_ahead_flex_rar_category1' : cpuc_flexibility_requirements.loc[(lse_abbrev,1,month),'flexibility_requirement'],
                'year_ahead_flex_rar_category2' : cpuc_flexibility_requirements.loc[(lse_abbrev,2,month),'flexibility_requirement'],
                'year_ahead_flex_rar_category3' : cpuc_flexibility_requirements.loc[(lse_abbrev,3,month),'flexibility_requirement'],
                'year_ahead_flex_rar_total' : cpuc_flexibility_requirements.loc[(lse_abbrev,(1,2,3),month),'flexibility_requirement'].sum(),
                'year_ahead_flex_incremental_category1' : incremental_flex_by_category(lse_abbrev,1),
                'year_ahead_flex_incremental_category2' : incremental_flex_by_category(lse_abbrev,2),
                'year_ahead_flex_incremental_category3' : incremental_flex_by_category(lse_abbrev,3),
                'los_angeles_rar_trueup' : local_rar_trueup.loc[(lse_abbrev,'los_angeles'),'local_rar_trueup'],
                'ventura_rar_trueup' : local_rar_trueup.loc[(lse_abbrev,'ventura'),'local_rar_trueup'],
                'san_diego_rar_trueup' : local_rar_trueup.loc[(lse_abbrev,'san_diego'),'local_rar_trueup'],
                'bay_area_rar_trueup' : local_rar_trueup.loc[(lse_abbrev,'bay_area'),'local_rar_trueup'],
                'fresno_rar_trueup' : local_rar_trueup.loc[(lse_abbrev,'fresno'),'local_rar_trueup'],
                'sierra_rar_trueup' : local_rar_trueup.loc[(lse_abbrev,'sierra'),'local_rar_trueup'],
                'stockton_rar_trueup' : local_rar_trueup.loc[(lse_abbrev,'stockton'),'local_rar_trueup'],
                'kern_rar_trueup' : local_rar_trueup.loc[(lse_abbrev,'kern'),'local_rar_trueup'],
                'humboldt_rar_trueup' : local_rar_trueup.loc[(lse_abbrev,'humboldt'),'local_rar_trueup'],
                'northern_california_rar_trueup' : local_rar_trueup.loc[(lse_abbrev,'northern_california'),'local_rar_trueup'],
                'los_angeles_august_demand_response' : august_demand_response('SCE','la_basin1',lse_abbrev),
                'ventura_august_demand_response' : august_demand_response('SCE','big_creek/ventura1',lse_abbrev),
                'san_diego_august_demand_response' : august_demand_response('SDGE','sdge1',lse_abbrev),
                'bay_area_august_demand_response' : august_demand_response('PGE','bay_area1',lse_abbrev),
                'fresno_august_demand_response' : august_demand_response('PGE','fresno',lse_abbrev),
                'sierra_august_demand_response' : august_demand_response('PGE','sierra',lse_abbrev),
                'stockton_august_demand_response' : august_demand_response('PGE','stockton',lse_abbrev),
                'kern_august_demand_response' : august_demand_response('PGE','kern',lse_abbrev),
                'humboldt_august_demand_response' : august_demand_response('PGE','humboldt',lse_abbrev),
                'northern_california_august_demand_response' : august_demand_response('PGE','ncnb',lse_abbrev),
                'los_angeles_incremental_load' : incremental_local_load.loc[(lse_abbrev,'los_angeles'),'incremental_load'],
                'ventura_incremental_load' : incremental_local_load.loc[(lse_abbrev,'ventura'),'incremental_load'],
                'san_diego_incremental_load' : incremental_local_load.loc[(lse_abbrev,'san_diego'),'incremental_load'],
                'bay_area_incremental_load' : incremental_local_load.loc[(lse_abbrev,'bay_area'),'incremental_load'],
                'fresno_incremental_load' : incremental_local_load.loc[(lse_abbrev,'fresno'),'incremental_load'],
                'sierra_incremental_load' : incremental_local_load.loc[(lse_abbrev,'sierra'),'incremental_load'],
                'stockton_incremental_load' : incremental_local_load.loc[(lse_abbrev,'stockton'),'incremental_load'],
                'kern_incremental_load' : incremental_local_load.loc[(lse_abbrev,'kern'),'incremental_load'],
                'humboldt_incremental_load' : incremental_local_load.loc[(lse_abbrev,'humboldt'),'incremental_load'],
                'northern_california_incremental_load' : incremental_local_load.loc[(lse_abbrev,'northern_california'),'incremental_load'],
            })
        summary = summary.merge(summary.apply(calculate_summary,axis='columns'),on='lse_abbrev')

        #  copy summary dataframe to worksheet:
        row_number = 6
        for row in summary.iterrows():
            ra_summary['NP26']['B{}'.format(row_number)].value = row[1]['lse_abbrev']
            ra_summary['NP26']['C{}'.format(row_number)].value = row[1]['np26_ra_obligation']
            ra_summary['NP26']['F{}'.format(row_number)].value = row[1]['sn_path26_allocation']
            ra_summary['SP26']['B{}'.format(row_number)].value = row[1]['lse_abbrev']
            ra_summary['SP26']['C{}'.format(row_number)].value = row[1]['sp26_ra_obligation']
            ra_summary['SP26']['F{}'.format(row_number)].value = row[1]['ns_path26_allocation']
            ra_summary['FlexRAR']['B{}'.format(row_number)].value = row[1]['lse_abbrev']
            ra_summary['FlexRAR']['M{}'.format(row_number)].value = row[1]['year_ahead_flex_rar_category1']
            ra_summary['FlexRAR']['N{}'.format(row_number)].value = row[1]['year_ahead_flex_rar_category2']
            ra_summary['FlexRAR']['O{}'.format(row_number)].value = row[1]['year_ahead_flex_rar_category3']
            ra_summary['FlexRAR']['P{}'.format(row_number)].value = row[1]['year_ahead_flex_rar_total']
            ra_summary['FlexRAR']['R{}'.format(row_number)].value = row[1]['year_ahead_flex_incremental_category1']
            ra_summary['FlexRAR']['S{}'.format(row_number)].value = row[1]['year_ahead_flex_incremental_category3']
            ra_summary['FlexRAR']['T{}'.format(row_number)].value = row[1]['year_ahead_flex_incremental_category3']
            ra_summary['Localtrueup']['A{}'.format(row_number)].value = row[1]['lse_abbrev']
            ra_summary['Localtrueup']['C{}'.format(row_number)].value = row[1]['los_angeles_rar_trueup']
            ra_summary['Localtrueup']['H{}'.format(row_number)].value = row[1]['ventura_rar_trueup']
            ra_summary['Localtrueup']['M{}'.format(row_number)].value = row[1]['san_diego_rar_trueup']
            ra_summary['Localtrueup']['R{}'.format(row_number)].value = row[1]['bay_area_rar_trueup']
            ra_summary['Localtrueup']['W{}'.format(row_number)].value = row[1]['humboldt_rar_trueup']
            ra_summary['Localtrueup']['AB{}'.format(row_number)].value = row[1]['sierra_rar_trueup']
            ra_summary['Localtrueup']['AG{}'.format(row_number)].value = row[1]['stockton_rar_trueup']
            ra_summary['Localtrueup']['AL{}'.format(row_number)].value = row[1]['northern_california_rar_trueup']
            ra_summary['Localtrueup']['AQ{}'.format(row_number)].value = row[1]['fresno_rar_trueup']
            ra_summary['Localtrueup']['AV{}'.format(row_number)].value = row[1]['kern_rar_trueup']
            ra_summary['Localtrueup']['E{}'.format(row_number)].value = row[1]['los_angeles_august_demand_response']
            ra_summary['Localtrueup']['J{}'.format(row_number)].value = row[1]['ventura_august_demand_response']
            ra_summary['Localtrueup']['O{}'.format(row_number)].value = row[1]['san_diego_august_demand_response']
            ra_summary['Localtrueup']['T{}'.format(row_number)].value = row[1]['bay_area_august_demand_response']
            ra_summary['Localtrueup']['Y{}'.format(row_number)].value = row[1]['humboldt_august_demand_response']
            ra_summary['Localtrueup']['AD{}'.format(row_number)].value = row[1]['sierra_august_demand_response']
            ra_summary['Localtrueup']['AI{}'.format(row_number)].value = row[1]['stockton_august_demand_response']
            ra_summary['Localtrueup']['AN{}'.format(row_number)].value = row[1]['northern_california_august_demand_response']
            ra_summary['Localtrueup']['AS{}'.format(row_number)].value = row[1]['fresno_august_demand_response']
            ra_summary['Localtrueup']['AX{}'.format(row_number)].value = row[1]['kern_august_demand_response']
            ra_summary['Localtrueup']['D{}'.format(row_number)].value = row[1]['los_angeles_incremental_load']
            ra_summary['Localtrueup']['I{}'.format(row_number)].value = row[1]['ventura_incremental_load']
            ra_summary['Localtrueup']['N{}'.format(row_number)].value = row[1]['san_diego_incremental_load']
            ra_summary['Localtrueup']['S{}'.format(row_number)].value = row[1]['bay_area_incremental_load']
            ra_summary['Localtrueup']['X{}'.format(row_number)].value = row[1]['fresno_incremental_load']
            ra_summary['Localtrueup']['AC{}'.format(row_number)].value = row[1]['sierra_incremental_load']
            ra_summary['Localtrueup']['AH{}'.format(row_number)].value = row[1]['stockton_incremental_load']
            ra_summary['Localtrueup']['AM{}'.format(row_number)].value = row[1]['kern_incremental_load']
            ra_summary['Localtrueup']['AR{}'.format(row_number)].value = row[1]['humboldt_incremental_load']
            ra_summary['Localtrueup']['AW{}'.format(row_number)].value = row[1]['northern_california_incremental_load']
            row_number += 1

        # save and close summary file and close all other files:
        ra_summary.save(str(self.paths['ra_summary']))
        ra_summary.close()
        # ra_monthly_filing.close()

        # check time and report:
        run_time = (dt.now() - init_time).total_seconds()
        self.logger.log('Retrieved Allocations in {:02.0f}:{:02.0f}:{:02.2f}'.format(int(run_time/3600),int((run_time%3600)/60),run_time%60),'INFORMATION')

        # return all dataframes for testing:
        return {
            'load_forecast_input_data' : load_forecast_input_data,
            'total_lcr' : total_lcr,
            'demand_response_allocation' : demand_response_allocation,
            'summary' : summary,
            'cam_rmr_monthly_tracking' : cam_rmr_monthly_tracking,
            'total_cam_rmr' : total_cam_rmr,
            'month_ahead_forecasts' : month_ahead_forecasts,
            'cpuc_flexibility_requirements' : cpuc_flexibility_requirements,
            'incremental_local_load' : incremental_local_load,
            'incremental_flex' : incremental_flex,
        }

    # collect data from each of the monthly lse filings:
    def consolidate_filings(self):
        # start timer:
        init_time = dt.now()

        # get list of active load serving entities from summary sheet:
        ra_summary = self.open_workbook('ra_summary',data_only=False)
        data_range = self.get_data_range(ra_summary['Summary'],'B','')
        active_lses = [row[0].value for row in data_range]

        # initialize summary table:
        summary_columns = [
            'lse_abbrev',
            'lse_officer_name',
            'lse_officer_title',
            'np26dr',
            'sp26dr',
        ]
        summary = pd.DataFrame(columns = summary_columns)

        # initialize physical resource table:
        physical_resources_columns=[
            'lse_abbrev',
            'contract_id',
            'resource_id',
            'resource_adequacy_system',
            'resource_adequacy_local',
            'resource_mcc_bucket',
            'continuous_availability',
            'resource_adequacy_committed_flexible',
            'resource_adequacy_flexibility_category'
        ]
        physical_resources = pd.DataFrame(columns=physical_resources_columns)

        # initialize demand response table:
        demand_response_columns = [
            'lse_abbrev',
            'contract_id',
            'program_id',
            'resource_adequacy_system',
            'resource_adequacy_local',
            'resource_mcc_bucket',
            'third_party_program',
            'resource_adequacy_committed_flexible',
            'resource_adequacy_flexibility_category',
        ]
        demand_response = pd.DataFrame(columns=demand_response_columns)

        # combine data tables from each lse filing:
        for lse_abbrev in active_lses:
            lse_filing = self.open_workbook('ra_monthly_filing',lse_abbrev=lse_abbrev,data_only=True)

            # retrieve values for summary table:
            summary_lse = pd.DataFrame({
                'lse_abbrev' : [lse_abbrev],
                'lse_officer_name' : [lse_filing['Certification']['B21'].value],
                'lse_officer_title' : [lse_filing['Certification']['B22'].value],
                'np26dr' : [
                        lse_filing['III_Demand_Response']['O5'].value +
                        lse_filing['III_Demand_Response']['S13'].value * 0.097 / 1.097
                    ],
                'sp26dr' : [
                        lse_filing['III_Demand_Response']['P5'].value +
                        lse_filing['III_Demand_Response']['U8'].value * 0.076 / 1.076 +
                        lse_filing['III_Demand_Response']['W6'].value * 0.096 / 1.096
                    ],
            })
            summary = pd.concat([summary,summary_lse],axis='index',ignore_index=True)

            # retrieve physical resources table:
            for row_number in range(5,lse_filing['I_Phys_Res_Import_RA_Res'].max_row+1):
                if str(lse_filing['I_Phys_Res_Import_RA_Res']['C{}'.format(row_number)].value).lower()=='none':
                    last_row = row_number
                    break
                else:
                    pass
            if last_row > 5:
                physical_resources_lse = self.data_range_to_dataframe(physical_resources_columns,lse_filing['I_Phys_Res_Import_RA_Res']['A5:I{}'.format(last_row)])
            else:
                physical_resources_lse = pd.DataFrame(columns=physical_resources_columns)
            physical_resources_lse.loc[:,'lse_abbrev'] = lse_abbrev
            physical_resources_lse.loc[:,'continuous_availability'] = physical_resources_lse.loc[:,'continuous_availability'].map(lambda s: True if s=='Y' else False)
            physical_resources = pd.concat([physical_resources,physical_resources_lse],axis='index',ignore_index=True)

            # retrieve demand response table:
            for row_number in range(17,lse_filing['III_Demand_Response'].max_row+1):
                if str(lse_filing['III_Demand_Response']['C{}'.format(row_number)].value).lower()=='none':
                    last_row = row_number
                    break
                else:
                    pass
            if last_row > 17:
                demand_response_lse = self.data_range_to_dataframe(demand_response_columns,lse_filing['III_Demand_Response']['A17:I{}'.format(last_row)])
            else:
                demand_response_lse = pd.DataFrame(columns=demand_response_columns)
            demand_response_lse.loc[:,'lse_abbrev'] = lse_abbrev
            demand_response_lse.loc[:,'third_party_program'] = demand_response_lse.loc[:,'third_party_program'].map(lambda s: True if s=='Y' else False)
            demand_response = pd.concat([demand_response,demand_response_lse],axis='index',ignore_index=True)
        
        # set summary table index:
        summary.set_index('lse_abbrev',inplace=True)
        summary.sort_index(inplace=True)
            
        # write to summary file:
        row_number_summary = 6
        row_number_physical_resources = 2
        for lse_abbrev in active_lses:
            ra_summary['NP26']['H{}'.format(row_number_summary)].value = summary.loc[lse_abbrev,'np26dr']
            ra_summary['SP26']['H{}'.format(row_number_summary)].value = summary.loc[lse_abbrev,'sp26dr']
            ra_summary['CertifyingOfficers']['F{}'.format(row_number_summary-2)].value = summary.loc[lse_abbrev,'lse_officer_name']
            ra_summary['CertifyingOfficers']['H{}'.format(row_number_summary-2)].value = summary.loc[lse_abbrev,'lse_officer_title']
            row_number += 1
            for row in physical_resources.loc[(physical_resources.loc[:,'lse_abbrev']==lse_abbrev),:].iterrows():
                ra_summary['PhysRes']['A{}'.format(row_number_physical_resources)].value = lse_abbrev
                ra_summary['PhysRes']['B{}'.format(row_number_physical_resources)].value = row[1].loc['contract_id']
                ra_summary['PhysRes']['C{}'.format(row_number_physical_resources)].value = row[1].loc['resource_id']
                ra_summary['PhysRes']['D{}'.format(row_number_physical_resources)].value = row[1].loc['resource_adequacy_system']
                ra_summary['PhysRes']['E{}'.format(row_number_physical_resources)].value = row[1].loc['resource_adequacy_local']
                ra_summary['PhysRes']['F{}'.format(row_number_physical_resources)].value = row[1].loc['resource_mcc_bucket']
                ra_summary['PhysRes']['G{}'.format(row_number_physical_resources)].value = row[1].loc['continuous_availability']
                ra_summary['PhysRes']['H{}'.format(row_number_physical_resources)].value = row[1].loc['resource_adequacy_committed_flexible']
                ra_summary['PhysRes']['I{}'.format(row_number_physical_resources)].value = row[1].loc['resource_adequacy_flexibility_category']
                row_number_physical_resources += 1
            for row in demand_response.loc[(demand_response.loc[:,'lse_abbrev']==lse_abbrev),:].iterrows():
                ra_summary['PhysRes']['A{}'.format(row_number_physical_resources)].value = lse_abbrev
                ra_summary['PhysRes']['B{}'.format(row_number_physical_resources)].value = row[1].loc['contract_id']
                ra_summary['PhysRes']['C{}'.format(row_number_physical_resources)].value = row[1].loc['program_id']
                ra_summary['PhysRes']['D{}'.format(row_number_physical_resources)].value = row[1].loc['resource_adequacy_system']
                ra_summary['PhysRes']['E{}'.format(row_number_physical_resources)].value = row[1].loc['resource_adequacy_local']
                ra_summary['PhysRes']['F{}'.format(row_number_physical_resources)].value = row[1].loc['resource_mcc_bucket']
                ra_summary['PhysRes']['G{}'.format(row_number_physical_resources)].value = 0
                ra_summary['PhysRes']['H{}'.format(row_number_physical_resources)].value = row[1].loc['resource_adequacy_committed_flexible']
                ra_summary['PhysRes']['I{}'.format(row_number_physical_resources)].value = row[1].loc['resource_adequacy_flexibility_category']
                row_number_physical_resources += 1
        ra_summary.save(str(self.paths['ra_summary']))
        ra_summary.close()

        # check time and report:
        run_time = (dt.now() - init_time).total_seconds()
        self.logger.log('Retrieved LSE Filings in {:02.0f}:{:02.0f}:{:02.2f}'.format(int(run_time/3600),int((run_time%3600)/60),run_time%60),'INFORMATION')

        # return all dataframes for testing:
        return {
            'summary' : summary,
            'physical_resources' : physical_resources,
            'demand_response' : demand_response,
        }

    # load excel file into memory:
    def open_workbook(self,file_id: str,lse_abbrev: str='',data_only:bool=True):
        if file_id in self.paths.keys():
            if file_id=='ra_monthly_filing':
                path = self.paths[file_id](lse_abbrev)
            else:
                path=self.paths[file_id]
            if path.is_file():
                with path.open('rb') as f:
                    in_mem_file = io.BytesIO(f.read())
                    in_mem_workbook = load_workbook(in_mem_file,data_only=data_only)
                self.logger.log('Loaded Workbook {}'.format(path.name),'INFORMATION')
            else:
                in_mem_workbook = None
                self.logger.log('File Not Found: {}'.format(path.name),'WARNING')
        else:
            in_mem_workbook = None
            self.logger.log('Unable to Load File for {}'.format(file_id),'WARNING')
        return in_mem_workbook

    # get indices of range of data in ra_summary file:
    def get_data_range(self,worksheet,lse_column:str,data_columns:str):
        first_row = 0
        last_row = 0
        for row_number in range(1,worksheet.max_row+1):
            if first_row==0 and str(worksheet['{}{}'.format(lse_column,row_number)].value).lower() in map(lambda s: s.lower(),self.lse_list):
                first_row = row_number
            if first_row>0 and last_row==0 and str(worksheet['{}{}'.format(lse_column,row_number+1)].value).lower() in ('none','total'):
                last_row = row_number
                break
        if first_row>0 and last_row>0:
            # return a list of lists containing excel cell objects:
            data_range = []
            for row_number in range(first_row,last_row+1):
                data_range.append([worksheet['{}{}'.format(lse_column,row_number)]]+[worksheet['{}{}'.format(data_column,row_number)] for data_column in data_columns])
            return data_range
        else:
            # return a list containing an empty list;
            return [[]]
    
    # read data range into a pandas dataframe:
    def data_range_to_dataframe(self,columns,data_range):
        data_array = []
        for data_range_row in data_range:
            data_array_row = []
            for data_range_cell in data_range_row:
                data_array_row.append(data_range_cell.value)
            data_array.append(data_array_row)
        return pd.DataFrame(data_array,columns=columns)

    # find and return data range for the flex requirements net cam table:
    def get_table(self,worksheet,table_header_text:str,table_header_offset:dict,columns:list):
        table_bounds = {
            'top' : worksheet.max_row,
            'bottom' : worksheet.max_row,
            'left' : get_column_letter(worksheet.max_column),
            'right' : get_column_letter(worksheet.max_column),
        }
        for row in worksheet['A1:{right}{bottom}'.format(**table_bounds)]:
            if row[0].row<table_bounds['bottom']:
                for cell in row:
                    if str(cell.value).lower()==table_header_text.lower():
                        # found header, setting upper, left, and right boundaries of table:
                        table_bounds['top'] = cell.row + table_header_offset['rows']
                        table_bounds['left'] = get_column_letter(column_index_from_string(cell.column) + table_header_offset['columns'])
                        table_bounds['right'] = get_column_letter(column_index_from_string(cell.column) + table_header_offset['columns'] + len(columns) - 1)
                    if cell.row>table_bounds['top'] and cell.column==table_bounds['left'] and cell.value is None:
                        # found last row in table:
                        table_bounds['bottom'] = cell.row - 1
            else:
                break
        data_range = worksheet['{left}{top}:{right}{bottom}'.format(**table_bounds)]
        return self.data_range_to_dataframe(columns,data_range)

    # convert filename template to actual filename based on filing month specified in configuration file:
    def parse_filename_template(self,filename_template: str,lse_abbrev: str=''):
        filename = filename_template
        replacements = {
            '[yy]' : self.configuration_options['filing_month'].strftime('%y'),
            '[yyyy]' : self.configuration_options['filing_month'].strftime('%Y'),
            '[mm]' : self.configuration_options['filing_month'].strftime('%m'),
            '[mmm]' : self.configuration_options['filing_month'].strftime('%b'),
            '[mmmm]' : self.configuration_options['filing_month'].strftime('%B'),
            '[lse_abbrev]' : lse_abbrev,
            '[lse_full]' : self.lse_map[lse_abbrev][0] if lse_abbrev else '',
        }
        for key in re.findall(r'\[\w[_A-Za-z]*\]',filename):
            if key in replacements.keys():
                filename = filename.replace(key,replacements[key])
            else:
                pass
        return Path(filename)